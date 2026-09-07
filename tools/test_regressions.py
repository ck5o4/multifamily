"""Regression tests for bugs fixed in the 2026-08-09 deep sweep.

Each test reproduces a bug that shipped and was caught by audit; a failure
here means the fix regressed. Run: python3 tools/test_regressions.py
"""

import sys

import pymodel


FAILURES = []


def check(name, cond, detail=""):
    status = "PASS" if cond else "FAIL"
    print(f"  [{status}] {name}" + (f"  ({detail})" if detail and not cond else ""))
    if not cond:
        FAILURES.append(name)


def base_inputs():
    return dict(pymodel._load_deal("baker-trails"),
                price=1_000_000, taxes_annual=12_000)


def test_solve_not_false_unreachable():
    """A midpoint IRR within tol just below target must not end the search."""
    inputs = base_inputs()
    asking = inputs["price"]
    mid = (max(50_000, asking * 0.35) + asking) / 2
    irr_mid = pymodel.run(dict(inputs, price=mid, exit_cap=None))["levered_irr"]
    res = pymodel.solve_price(dict(inputs), irr_mid + 0.00005)
    check("solve_price: no false 'unreachable' on tol-below midpoint",
          res is not None and res.get("price", 0) > 0,
          f"returned {res}")


def test_no_phantom_debt_service():
    """After the loan fully amortizes, debt service must be zero."""
    r = pymodel.run(dict(base_inputs(), amort_years=3, hold_years=5,
                         exit_cap=None))
    check("no phantom DS after amortization",
          r["total_ds"][4] == 0.0 and r["total_ds"][5] == 0.0,
          f"yr4={r['total_ds'][4]:,.0f} yr5={r['total_ds'][5]:,.0f}")
    check("final amortization year still charges its full 12 payments",
          abs(r["total_ds"][3] - r["total_ds"][1]) < 1.0,
          f"yr3={r['total_ds'][3]:,.0f} vs yr1={r['total_ds'][1]:,.0f}")


def test_waterfall_invalid_flag():
    """Negative distributable years must null LP/GP metrics and set the flag."""
    r = pymodel.run(dict(base_inputs(), refi_year=3, refi_valuation_cap=0.11,
                         refi_cost_pct=0.03, exit_cap=None))
    if not r.get("waterfall_neg_years"):
        check("waterfall invalid-flag scenario still produces a deficit year",
              False, "scenario no longer triggers; rebuild the trigger")
        return
    check("waterfall: LP/GP nulled when distributable cash is negative",
          r.get("waterfall_invalid") is True and r["lp_irr"] is None
          and r["gp_irr"] is None)


def test_mc_vacancy_centered_on_underwriting():
    """MC vacancy process must center on the deal's own vacancy + bad debt."""
    eden = pymodel._load_deal("eden-church-mhp")
    det = pymodel.run(eden)["levered_irr"]
    mc = pymodel.monte_carlo(eden, n=1000, seed=42,
                             deal_name="eden-church-mhp")
    # Pre-fix: P50 ~18.5% vs det 9.9% (vacancy swapped 13.7% -> 8.6%).
    #
    # Directional since the 2026-08-24 sweep. The failure mode this test exists
    # to catch is MC coming out OPTIMISTIC against the deal's own underwriting,
    # and a symmetric band cannot express that: once rent/expense growth were
    # also recentered, eden's honest P50 settled 4.2pts BELOW deterministic
    # (left skew from the insurance shock, the integer-unit vacancy process and
    # the exit-cap spread), which the old |diff| < 4pts band read as a failure.
    # Assert the direction that matters, and keep a loose floor for sanity.
    check("MC P50 does not exceed deterministic on eden (optimism guard)",
          mc["p50"] - det < 0.01,
          f"P50 {mc['p50']:.1%} vs det {det:.1%} (+{(mc['p50']-det)*100:.1f}pts)")
    check("MC P50 within 8pts below deterministic on eden (sanity floor)",
          det - mc["p50"] < 0.08,
          f"P50 {mc['p50']:.1%} vs det {det:.1%}")
    check("MC reports the vacancy recentering note",
          bool(mc.get("vacancy_note")))


def test_rentcast_matcher_requires_all_tokens():
    """A different property sharing one name token must not match the cache."""
    good = pymodel._load_rentcast_mult("eden-church-mhp", 1180.0)
    bad = pymodel._load_rentcast_mult("church-street-8plex", 1000.0)
    check("rentcast: eden still matches its own cache", good is not None)
    check("rentcast: cross-property token match rejected", bad is None)


def test_discovery_prefers_newest_and_reports_alternates():
    """Multiple candidates for one slot: newest wins, losers are reported.

    2026-08-24 sweep: discover() took the alphabetically FIRST match and
    silently dropped the rest. baker-trails therefore resolved to
    rentroll_baker_trails_ESTIMATED.csv (GPR $118,800) instead of the seller's
    rentroll_baker_trails_OM_2026-08-13.csv (GPR $114,000 in-place, real
    8x2BR/4x3BR mix), with no warning that a second roll existed.
    """
    import intake
    from pathlib import Path

    baker = Path(intake.INTAKE) / "baker-trails"
    found, alt = intake.discover(baker)
    check("discovery: baker-trails resolves to the OM roll, not ESTIMATED",
          found.get("rent_roll") is not None
          and found["rent_roll"].name == "rentroll_baker_trails_OM_2026-08-13.csv",
          f"picked {found.get('rent_roll')}")
    check("discovery: the superseded roll is reported, not dropped",
          any(p.name == "rentroll_baker_trails_ESTIMATED.csv"
              for p in alt.get("rent_roll", [])),
          f"alternates {alt.get('rent_roll')}")

    hwy = Path(intake.INTAKE) / "hwy42-mhp"
    found2, alt2 = intake.discover(hwy)
    check("discovery: hwy42 T-12 picks YTD 2026 over P&L 2025",
          found2.get("t12") is not None and "2026" in found2["t12"].name,
          f"picked {found2.get('t12')}")
    check("discovery: hwy42's two rejected T-12s are both reported",
          len(alt2.get("t12", [])) == 2,
          f"alternates {alt2.get('t12')}")

    # A folder with one candidate per slot must report no alternates at all.
    eden = Path(intake.INTAKE) / "eden-church-mhp"
    _, alt3 = intake.discover(eden)
    check("discovery: no false alternates when each slot is unambiguous",
          alt3 == {}, f"alternates {alt3}")


def test_solver_returns_a_price_that_clears_its_own_target():
    """Every solved rung must actually hit the IRR it is labelled with.

    2026-08-24: the bisection break tested the last midpoint probe instead of
    the best clearing price it would return, so a near-miss midpoint ended the
    search while `best` still held a far lower price (hwy42 @16% reported
    $2,269,000 at 16.54% when $2,298,000 clears). And the result was ROUNDED to
    the nearest $1,000, which rounds UP past the boundary so the reported price
    no longer clears (treme 16% -> $664,000 @ 15.997%).
    """
    for deal in ("eden-church-mhp", "treme-gov-nicholls", "hwy42-mhp",
                 "covington-2nd", "baker-trails", "weber-city-mhp"):
        inputs = pymodel._load_deal(deal)
        for target in (0.13, 0.16, 0.22):
            res = pymodel.solve_price(dict(inputs), target)
            if not res or not res.get("price"):
                continue
            check(f"solve_price: {deal} @{target:.0%} clears its own target",
                  res["irr"] >= target,
                  f"${res['price']:,} -> {res['irr']:.4%} < {target:.0%}")

    # covington @16% was reported unreachable when it is reachable.
    res = pymodel.solve_price(dict(pymodel._load_deal("covington-2nd")), 0.16)
    check("solve_price: covington-2nd 16% is reachable, not 'unreachable'",
          res is not None and res.get("price") is not None,
          f"returned {res}")


def test_irr_requires_a_sign_change():
    """No sign change means no IRR. Returning a number there is fabrication.

    2026-08-24: all-zero flows returned the bisection bracket midpoint (450.05%)
    and [0,...,0,X] returned the Newton clamp (5000%). Both are reachable in one
    step from lp_pct=1.0 (investor funds all equity -> gp_capital == 0), which
    printed "GP IRR: 450.05%".
    """
    check("_irr: all-zero flows have no IRR", pymodel._irr([0, 0, 0, 0, 0, 0]) is None)
    check("_irr: inflow-only flows have no IRR",
          pymodel._irr([0, 0, 0, 0, 0, 17683]) is None)
    check("_irr: a normal flow still solves",
          pymodel._irr([-100, 10, 10, 10, 120]) is not None)
    r = pymodel.run(dict(pymodel._load_deal("covington-2nd"), lp_pct=1.0))
    check("GP IRR is None (not 450%) when the LP funds all equity",
          r["gp_irr"] is None and r["gp_capital"] == 0,
          f"gp_irr={r['gp_irr']} gp_capital={r['gp_capital']}")


def test_mc_growth_recentered_on_underwriting():
    """MC must not draw looser growth than the deal underwrites.

    2026-08-24: vacancy was recentered in the 2026-08-09 sweep but rent growth
    (fitted P50 3.215% vs a 2.0% underwrite) and expense growth (2.147% vs
    2.5%) were not, so MC P50 came out ABOVE the deterministic IRR on all seven
    deals and every quoted P(IRR>=13%) inherited the optimism.
    """
    for deal in ("eden-church-mhp", "treme-gov-nicholls", "baker-trails"):
        inputs = pymodel._load_deal(deal)
        det = pymodel.run(inputs)["levered_irr"]
        mc = pymodel.monte_carlo(inputs, n=1000, seed=42, deal_name=deal)
        check(f"MC P50 does not exceed deterministic on {deal}",
              mc["p50"] - det < 0.01,
              f"P50 {mc['p50']:.2%} vs det {det:.2%}")
    mc = pymodel.monte_carlo(pymodel._load_deal("eden-church-mhp"), n=200, seed=1,
                             deal_name="eden-church-mhp")
    check("MC reports the growth recentering note", bool(mc.get("growth_note")))


def test_part_year_t12_is_flagged_not_read_as_annual():
    """A 6-month statement must never be reported as an annual figure.

    2026-08-24: hwy42's 'P&L YTD 2026.xlsx' (JAN-JUN populated) parsed every
    expense at half its annual size with basis still reading 'under total
    header' - insurance $12,750 against a true $25,500. This became reachable
    when discovery started preferring the newest file.
    """
    import parsers
    from pathlib import Path
    import intake

    ytd = Path(intake.INTAKE) / "hwy42-mhp" / "P&L YTD 2026.xlsx"
    lines, notes = parsers.parse_t12(ytd, units=30)
    check("part-year T-12 raises a PART-YEAR note",
          any("PART-YEAR" in n for n in notes), f"notes={notes}")
    check("part-year T-12 stamps every line's basis",
          all("partial year" in v["basis"] for v in lines.values()),
          f"{[v['basis'] for v in lines.values()][:3]}")

    for full in ("hwy42-mhp/P&L 2025.xlsx", "eden-church-mhp/PL_2025_ACTUALS.xlsx"):
        _, n = parsers.parse_t12(Path(intake.INTAKE) / full, units=18)
        check(f"full-year statement is NOT flagged part-year ({full.split('/')[0]})",
              not any("PART-YEAR" in x for x in n))


def test_rent_roll_reports_its_own_implied_vacancy():
    """A roll with vacant units must surface the vacancy it implies.

    2026-08-24: vacant units' rents are imputed from the type median so they
    land in gross potential income, while the model was written the 7% template
    default. On baker-trails' OM roll that is 4 of 12 units - 33.3% vs 7%,
    roughly $30k/yr of NOI on a $750k asset.
    """
    import parsers
    from pathlib import Path
    import intake

    roll = (Path(intake.INTAKE) / "baker-trails"
            / "rentroll_baker_trails_OM_2026-08-13.csv")
    _, notes = parsers.parse_rent_roll(roll)
    hit = [n for n in notes if "IMPLIED PHYSICAL VACANCY" in n]
    check("rent roll reports implied physical vacancy", bool(hit), f"notes={notes}")
    check("implied vacancy on the baker OM roll reads 33.3%",
          bool(hit) and "33.3%" in hit[0], f"{hit}")


def test_flood_degrades_to_unknown_not_clear():
    """A failed or unstudied flood lookup is UNKNOWN, never 'no flood risk'.

    2026-08-24: interpret([]) returned sfha=False (so icmemo's `if sfha` gate
    dropped the flood line and asserted "SFHA: No"); FEMA's 'AREA NOT INCLUDED'
    and 'OPEN WATER' domain values fell through to "minimal hazard"; and
    fetch_json raised SystemExit (BaseException), making every `except Exception`
    guard dead code so a service failure killed the whole run.
    """
    import flood
    check("flood: FloodLookupError is a catchable Exception",
          issubclass(flood.FloodLookupError, Exception))
    check("flood: unmapped parcel is sfha UNKNOWN (None), not False",
          flood.interpret([])[1] is None)
    for z in ("AREA NOT INCLUDED", "OPEN WATER", "SOMETHING NEW"):
        zone, sfha, note = flood.interpret(
            [{"FLD_ZONE": z, "ZONE_SUBTY": None, "SFHA_TF": "F"}])
        check(f"flood: {z!r} is undetermined, not minimal",
              sfha is None and "UNDETERMINED" in note,
              f"sfha={sfha} note={note!r}")
    check("flood: plain X is still minimal and not SFHA",
          flood.interpret([{"FLD_ZONE": "X", "ZONE_SUBTY": None, "SFHA_TF": "F"}])[1] is False)
    check("flood: AE is still SFHA",
          flood.interpret([{"FLD_ZONE": "AE", "ZONE_SUBTY": None, "SFHA_TF": "T"}])[1] is True)


def test_rentcast_matches_by_address_hint():
    """A deal named for its neighbourhood still finds its address-keyed cache.

    2026-08-24: the slug-token matcher required every >3-char token of the deal
    slug to appear in the cache filename. Cache files are named for the street,
    so 'treme-gov-nicholls' (no shared token with '1429-governor-nicholls-st')
    silently missed its own paid cache and the MC ran with no rent-level risk.
    """
    good = pymodel._load_rentcast_mult("treme-gov-nicholls", 1400.0)
    check("rentcast: treme matches via the deals.json address hint",
          good is not None, f"got {good}")
    baker = pymodel._load_rentcast_mult("baker-trails", 800.0)
    check("rentcast: baker matches via the deals.json address hint",
          baker is not None, f"got {baker}")
    # A deal with no cache and no hint must produce a LOUD note, not silence.
    mc = pymodel.monte_carlo(pymodel._load_deal("weber-city-mhp"), n=200, seed=1,
                             deal_name="no-such-deal-xyz")
    check("rentcast: an unmatched deal emits a loud rent_level_note",
          "no RentCast cache matched" in (mc.get("rent_level_note") or ""),
          f"note={mc.get('rent_level_note')!r}")


def test_beats_index_is_deterministic():
    """The house gate must not depend on an arbitrary RNG seed.

    2026-08-24: pairing each IRR sample with one rng.gauss() draw put +/-3pp of
    pure RNG noise on a rule whose threshold is exactly 50% (eden ranged
    52.1%-57.7% across twenty seed choices). Closed form removes that term.
    """
    import board
    samples = [0.02, 0.08, 0.10, 0.12, 0.20, -0.05]
    a = board._beats_index(samples)
    b = board._beats_index(samples)
    check("beats_index is deterministic", a == b, f"{a} vs {b}")
    check("beats_index of a far-above-market sample set approaches 1",
          board._beats_index([0.60] * 50) > 0.99)
    check("beats_index of a far-below-market sample set approaches 0",
          board._beats_index([-0.40] * 50) < 0.01)
    check("beats_index of an at-market sample set is ~0.5",
          abs(board._beats_index([0.10] * 50) - 0.5) < 1e-9)
    check("beats_index returns None on no samples", board._beats_index([]) is None)


def test_unknown_model_input_is_refused():
    """An input key the engine does not read must raise, not be ignored.

    2026-09-07: `run({**deal, "purchase_price": 1_251_000})` returned the
    ask-price answer (eden 9.87%) instead of the $1.251M answer (32.91%) with
    no diagnostic, because _merge_defaults did `dict(defaults); update(inputs)`
    and nothing ever read the stray key. The repo's own history records a price
    ladder published and retracted over this class of slip (baker-trails,
    2026-08-03 "RED-TEAM CORRECTION"). Silence is the bug; the raise is the fix.
    """
    base = pymodel._load_deal("eden-church-mhp")
    ok = pymodel.run(dict(base))
    check("unknown-key guard: a valid input set still runs",
          ok.get("levered_irr") is not None)

    for bad_key in ("purchase_price", "totally_bogus_key_xyz"):
        try:
            pymodel.run({**base, bad_key: 1_251_000})
            check(f"unknown model input {bad_key!r} raises", False,
                  "no exception raised")
        except ValueError as exc:
            check(f"unknown model input {bad_key!r} raises",
                  bad_key in str(exc), f"message did not name the key: {exc}")

    # The correctly-spelled key must still change the answer.
    at_ask = pymodel.run(dict(base))["levered_irr"]
    at_pursue = pymodel.run({**base, "price": 1_251_000})["levered_irr"]
    check("unknown-key guard: 'price' still overrides",
          at_pursue > at_ask + 0.10,
          f"ask {at_ask:.4f} vs pursue {at_pursue:.4f}")

    # Keys that are legitimate but absent from the defaults table (supplied by
    # _load_deal / callers) must NOT trip the guard.
    for legit in ("location", "unit_mix", "price"):
        check(f"unknown-key guard: {legit!r} is accepted",
              legit in base and pymodel.run(dict(base)) is not None)


def test_insurance_gate_needs_an_artifact_not_prose():
    """Prose naming an OPEN insurance gate must not close the gate.

    2026-09-07: _insurance_noted returned True on the substring "bindable".
    Eden's history says 'GATE: bindable habitational insurance <=~1600/u
    (Apartment Guard follow-up drafted)' — the sentence that says the quote is
    outstanding — so the IC memo printed '✓ Insurance quote noted' and the bank
    package labelled the seller's carried $1,060/unit premium
    '(Louisiana-adjusted)' to a lender. One of CLAUDE.md's four offer-stage
    hard gates, reported clear while open, on the Priority 1 deal.
    """
    import json as _json
    import tempfile
    from pathlib import Path as _Path
    import icmemo

    root = _Path(__file__).resolve().parent.parent
    deals = _json.loads((root / "portfolio" / "deals.json").read_text())

    for deal in ("eden-church-mhp", "treme-gov-nicholls", "baker-trails"):
        rec = deals.get(deal, {})
        hist = " ".join(h.get("note", "") for h in rec.get("history", []))
        noted, _ev = icmemo._insurance_noted(root / "deal-intake" / deal, hist, rec)
        check(f"insurance gate: {deal} has no quote artifact, so it reads UNMET",
              noted is False, f"got {noted}")

    # The exact sentence that used to close the gate.
    with tempfile.TemporaryDirectory() as td:
        empty = _Path(td)
        open_gate = ("GATE: bindable habitational insurance <=~1600/u "
                     "(Apartment Guard follow-up drafted)")
        noted, _ = icmemo._insurance_noted(empty, open_gate, {})
        check("insurance gate: the word 'bindable' alone does not close it",
              noted is False, f"got {noted}")
        for phrase in ("insurance quote requested", "flood quote pending"):
            noted, _ = icmemo._insurance_noted(empty, phrase, {})
            check(f"insurance gate: {phrase!r} does not close it", noted is False)

        # Positive control 1: a quote document filed in the deal folder.
        (empty / "insurance_quote_apartmentguard.pdf").write_text("x")
        noted, ev = icmemo._insurance_noted(empty, "", {})
        check("insurance gate: a filed quote document closes it",
              noted is True and "insurance_quote_apartmentguard.pdf" in ev,
              f"got {noted}, {ev!r}")

    # Positive control 2: an explicit record in deals.json.
    with tempfile.TemporaryDirectory() as td:
        noted, ev = icmemo._insurance_noted(
            _Path(td), "",
            {"insurance_quote": {"carrier": "Apartment Guard",
                                 "per_unit": 1550, "date": "2026-09-01"}})
        check("insurance gate: a deals.json insurance_quote record closes it",
              noted is True and "Apartment Guard" in ev, f"got {noted}, {ev!r}")


def test_bankpackage_diligence_gates_fail_closed():
    """A failure in the gap check must not produce a clean bank package.

    2026-09-07: the check sat under `except Exception: pass` with defaults of
    rent-roll-actual / T-12-present / insurance-quoted, so any error handed the
    lender a package with every gate silently reported clear.
    """
    import bankpackage
    import icmemo

    # Break the gap check the way a real failure would, and confirm the package
    # refuses instead of sailing through with every gate reported clear.
    original = icmemo._insurance_noted

    def _boom(*a, **k):
        raise RuntimeError("simulated gap-check failure")

    icmemo._insurance_noted = _boom
    try:
        raised = False
        try:
            bankpackage.gather("eden-church-mhp", force=False)
        except SystemExit:
            raised = True          # refused, which is the fail-closed behaviour
        except Exception:
            raised = True
        check("bankpackage: a broken gap check refuses rather than passing clean",
              raised, "gather() returned normally with the check broken")
    finally:
        icmemo._insurance_noted = original


def test_intake_apply_end_to_end_preserves_the_workbook():
    """`intake.py --apply` must write the deal, and never destroy it on failure.

    2026-09-07: intake.py:310 called `defaults.resolve(args.model)` (4 args
    required) behind an `hasattr(defaults, "resolve")` guard that is always
    True, so every --apply on a roll containing a vacant unit raised TypeError.
    clone_model() had already copied the blank master over the deal workbook,
    and w.save() never ran: baker-trails' underwriting silently became the
    master's $600,000 / 4x1BR+4x2BR demo deal. Two weeks live, because no test
    exercised intake.main() — only parsers.parse_rent_roll.

    Runs against a throwaway copy of the repo; touches nothing real.
    """
    import shutil
    import subprocess
    import tempfile
    from pathlib import Path as _Path
    import openpyxl

    root = _Path(__file__).resolve().parent.parent
    with tempfile.TemporaryDirectory() as td:
        sandbox = _Path(td) / "repo"
        shutil.copytree(root, sandbox, ignore=shutil.ignore_patterns(
            ".git", "market-data", "docs", "reference", "*.pdf"))
        wb_path = sandbox / "deal-intake" / "baker-trails" / "baker-trails_acq.xlsx"

        def mix(path):
            ws = openpyxl.load_workbook(path)["Inputs"]
            return (ws["B2"].value,
                    [(ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 8).value)
                     for r in (3, 4)])

        before = mix(wb_path)
        proc = subprocess.run(
            [sys.executable, str(sandbox / "tools" / "intake.py"),
             "--deal", "baker-trails", "--price", "750000",
             "--location", "Baker", "--apply"],
            capture_output=True, text=True, timeout=600)
        out = proc.stdout + proc.stderr

        check("intake --apply: does not raise TypeError on a roll with vacancies",
              "TypeError" not in out, out[-300:])
        check("intake --apply: writes the workbook",
              "WROTE" in out and proc.returncode == 0,
              f"rc={proc.returncode}")

        after = mix(wb_path)
        # The master demo deal is $600,000 / 4x'1 BR/ 1 BA' - never the result.
        check("intake --apply: workbook is not replaced by the master demo deal",
              after[0] == 750000 and after[1][0][1] != "1 BR/ 1 BA",
              f"before={before} after={after}")
        check("intake --apply: the OM roll's real mix lands in the workbook",
              after[1] == [(8, "2 BR/ 1 BA", 725), (4, "3 BR/ 1 BA", 925)],
              f"got {after[1]}")
        check("intake --apply: the roll's own 33.3% vacancy is surfaced",
              "33.3% physical vacancy" in out)
        check("intake --apply: no staging file is left behind",
              not list(wb_path.parent.glob("*staging*")))

        # A failure after the clone must leave the ORIGINAL workbook intact.
        good = mix(wb_path)
        proc = subprocess.run(
            [sys.executable, str(sandbox / "tools" / "intake.py"),
             "--deal", "baker-trails", "--price", "750000",
             "--set", "no_such_input_key=1", "--apply"],
            capture_output=True, text=True, timeout=600)
        check("intake --apply: a failed run leaves the previous workbook intact",
              mix(wb_path) == good and proc.returncode != 0,
              f"rc={proc.returncode}, workbook now {mix(wb_path)}")


def test_solve_price_returns_the_basis_it_certified():
    """A rung must be re-scorable on the basis that certified it.

    2026-09-07: solve_price re-derived taxes and the exit cap at every trial
    price, then returned {price, irr} alone. Consumers re-ran the rung with the
    ASKING price's tax bill and exit cap still attached, understating the
    rung's own IRR and its P(IRR>=target) by 3.0-5.5pp — on eden's 22% rung,
    which is the live pursue basis.
    """
    for deal, target in (("eden-church-mhp", 0.22), ("treme-gov-nicholls", 0.13)):
        inputs = pymodel._load_deal(deal)
        res = pymodel.solve_price(dict(inputs), target)
        check(f"solve_price: {deal} @{target:.0%} solves", res is not None)
        if not res:
            continue
        check(f"solve_price: {deal} returns the tax basis it used",
              res.get("taxes_annual") is not None
              and abs(res["taxes_annual"] - inputs["taxes_annual"]) > 1,
              f"returned {res.get('taxes_annual')}, ask basis {inputs['taxes_annual']}")
        check(f"solve_price: {deal} returns the exit cap it used",
              res.get("exit_cap") is not None)
        # Re-scored on the returned basis, the rung reproduces its own label.
        rescored = pymodel.run(dict(inputs, price=res["price"],
                                    taxes_annual=res["taxes_annual"],
                                    exit_cap=res["exit_cap"]))["levered_irr"]
        check(f"solve_price: {deal} @{target:.0%} re-scores to its own label",
              abs(rescored - target) < 0.002,
              f"re-scored {rescored:.4f} vs target {target}")


def test_commercial_share_is_a_valid_model_input():
    """The unknown-key guard must not reject a key _load_deal injects.

    2026-09-07: the guard added earlier the same day omitted commercial_share,
    which _load_deal copies out of deals.json exactly as it copies location.
    Every mixed-use deal — supported, documented, and inside the NOLA buy box —
    would have raised in run/tornado/monte_carlo while solve_price (which pops
    the key) kept working: a confusing half-failure.
    """
    inputs = pymodel._load_deal("treme-gov-nicholls")
    inputs["commercial_share"] = 0.3
    r = pymodel.run({k: v for k, v in inputs.items() if k != "location"})
    check("commercial_share is accepted by run()",
          r.get("levered_irr") is not None)
    try:
        pymodel.run({**inputs, "comercial_share": 0.3})
        check("a misspelled commercial_share is still rejected", False, "no raise")
    except ValueError:
        check("a misspelled commercial_share is still rejected", True)


def test_icmemo_tornado_keeps_every_downside_row():
    """The vacancy stress must survive into the memo the house rule cites."""
    import icmemo
    inputs = pymodel._load_deal("treme-gov-nicholls")
    rows = pymodel.tornado(inputs)
    downside = [r for r in rows if (r["delta_irr"] or 0) < 0]
    check("tornado: treme has a vacancy downside row",
          any("vacancy" in r["factor"].lower() for r in downside),
          f"factors: {[r['factor'] for r in downside]}")
    check("tornado: more downside rows exist than the old top-3 slice showed",
          len(downside) > 3, f"{len(downside)} downside rows")
    memo = "\n".join(icmemo.build_memo_lines("treme-gov-nicholls")) \
        if hasattr(icmemo, "build_memo_lines") else None
    if memo is not None:
        check("icmemo: the vacancy stress row reaches the memo",
              "vacancy" in memo.lower())


def test_hedonic_market_comes_from_the_parish():
    """The comp market must follow the deal's parish, not its history prose."""
    import hedonic
    cases = [("new orleans", "new orleans"), ("gretna", "new orleans"),
             ("chalmette", "new orleans"), ("covington", "northshore"),
             ("hammond", "northshore"), ("baker", "baton rouge"),
             ("denham springs", "baton rouge"), ("gonzales", "baton rouge"),
             ("lafayette", "lafayette")]
    for loc, want in cases:
        got, how = hedonic.market_for_location(loc)
        check(f"hedonic: {loc} -> {want}", got == want, f"got {got} ({how})")
    got, why = hedonic.market_for_location("Nowheresville")
    check("hedonic: an unresolvable location fails rather than defaulting to "
          "baton rouge", got is None, f"got {got}")
    got, why = hedonic.market_for_location(None)
    check("hedonic: a missing location fails loudly", got is None)


def test_hedonic_has_no_year_regressor():
    """The year term measured market mix, not appreciation — and was applied.

    2026-09-07: fit() printed "Do NOT read it as a time trend" and predict()
    then multiplied every estimate by exp(-0.0881 * (year-2021)) = 0.644 at
    2026. Baton Rouge read $41,664/unit instead of $59,100; a 12-unit Baker
    priced at $499,969 instead of $709,198, so the market approach appeared to
    corroborate a sub-$500K ladder on the strength of an acknowledged artifact.
    """
    import hedonic
    fr = hedonic.fit(verbose=False)
    check("hedonic: design matrix has no year column", fr["k"] == 5, f"k={fr['k']}")
    a = hedonic.predict("baton rouge", 12, year=2021, fit_result=fr, loud=False)
    b = hedonic.predict("baton rouge", 12, year=2026, fit_result=fr, loud=False)
    check("hedonic: the year argument no longer moves the estimate",
          a["point_per_unit"] == b["point_per_unit"],
          f"{a['point_per_unit']} vs {b['point_per_unit']}")
    check("hedonic: baton rouge is back above the artifact-suppressed level",
          b["point_per_unit"] > 55_000, f"got {b['point_per_unit']}")
    ns = hedonic.predict("northshore", 18, fit_result=fr, loud=False)
    check("hedonic: a one-sale market coefficient is flagged as thin support",
          any("THIN SUPPORT" in c for c in ns["caveats"]), f"{ns['caveats']}")
    nola = hedonic.predict("new orleans", 8, fit_result=fr, loud=False)
    check("hedonic: a well-supported market is not flagged thin",
          not any("THIN SUPPORT" in c for c in nola["caveats"]))


def test_flood_cli_does_not_print_no_for_undetermined():
    """`SFHA: no` is the line that gets copied into memos."""
    import io
    import contextlib
    import flood
    for zone, sfha, want in (("X", False, "no"), ("AE", True, "YES"),
                             ("UNMAPPED", None, "UNDETERMINED"),
                             ("D", None, "UNDETERMINED")):
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            flood._print_result({"address": "a", "zone": zone, "sfha": sfha,
                                 "note": "n"}, as_json=False)
        line = next(l for l in buf.getvalue().splitlines() if l.startswith("SFHA"))
        check(f"flood CLI: {zone} (sfha={sfha}) prints {want}",
              want in line, f"printed {line!r}")


def test_latax_rejects_a_percentage_typed_as_a_percent():
    """`--commercial-share 30` meaning 30% used to return a 16x tax bill."""
    import latax
    ok, _ = latax.estimate_tax(1_500_000, "Gonzales", 0.3)
    check("latax: a valid fraction still works", ok is not None and ok > 0)
    for bad in (30, -0.5, 1.5):
        try:
            latax.estimate_tax(1_500_000, "Gonzales", bad)
            check(f"latax: commercial_share={bad} is rejected", False, "no raise")
        except ValueError:
            check(f"latax: commercial_share={bad} is rejected", True)


def test_market_parish_coverage_matches_latax():
    """market.py's docstring claimed parity with latax; it was short two."""
    import latax
    import market
    missing = sorted(set(latax.MILLAGE) - set(market.PARISHES))
    check("market.py covers every latax parish", not missing, f"missing {missing}")
    check("market.py covers st. bernard (in the buy box)",
          "st. bernard" in market.PARISHES)


def test_irr_verdict_applies_the_pursue_floor():
    """12.5% is inside the realistic band and below the house floor."""
    import report
    check("irr_verdict: 12.5% is called out as below the pursue floor",
          "BELOW PURSUE FLOOR" in report.irr_verdict(0.125),
          report.irr_verdict(0.125))
    check("irr_verdict: 12.9% likewise",
          "BELOW PURSUE FLOOR" in report.irr_verdict(0.129))
    check("irr_verdict: 13.0% clears",
          "BELOW PURSUE FLOOR" not in report.irr_verdict(0.130),
          report.irr_verdict(0.130))


def test_portfolio_sim_charges_idle_capital_once():
    """Year 0 debited the whole cheque; later buys debited their equity again.

    2026-09-07: baker-then-fourplex printed IRR -1.1% / -$23,995 profit beside
    "Cash returned $437,208 (1.46x)" because $143,747 of idle year-0 cash was
    charged twice. On a capital-called convention the same sequence is +9.0%.
    """
    import portfolio_sim
    res = portfolio_sim.simulate(
        portfolio_sim._load_scenario("baker-then-fourplex"))
    cf = res["portfolio_cf"]
    check("portfolio_sim: a profitable sequence does not print as a loss",
          res["total_profit"] > 0, f"profit {res['total_profit']:,.0f}")
    check("portfolio_sim: portfolio IRR is positive on this scenario",
          (res["portfolio_irr"] or 0) > 0, f"IRR {res['portfolio_irr']}")
    check("portfolio_sim: year 0 charges deployed capital, not the whole cheque",
          abs(cf[0]) < res["starting_equity"] - 1,
          f"cf[0]={cf[0]:,.0f} vs cheque {res['starting_equity']:,.0f}")
    check("portfolio_sim: sum of flows equals reported profit",
          abs(sum(cf) - res["total_profit"]) < 1.0)


def test_portfolio_status_excludes_non_operating_credits():
    """A loan draw is not rent."""
    import portfolio
    acts = [
        {"date": "2027-01-05", "desc": "Rent", "amount": 1200, "cat": "rent"},
        {"date": "2027-01-06", "desc": "Construction loan draw",
         "amount": 25000, "cat": "debt"},
        {"date": "2027-01-15", "desc": "Repairs", "amount": -180, "cat": "repairs"},
    ]
    import io
    import contextlib
    import json as _json
    import shutil
    import tempfile
    from pathlib import Path as _Path

    root = _Path(__file__).resolve().parent.parent
    original = portfolio.DEALS
    with tempfile.TemporaryDirectory() as td:
        fake = _Path(td) / "deals.json"
        rec = {"testdeal": {"stage": "owned", "created": "2027-01-01",
                            "history": [], "actuals": acts, "payback": [],
                            "plan": {"monthly_noi_target": 11956}}}
        fake.write_text(_json.dumps(rec))
        portfolio.DEALS = fake
        try:
            buf = io.StringIO()
            with contextlib.redirect_stdout(buf):
                portfolio.cmd_status("testdeal")
            out = buf.getvalue()
        finally:
            portfolio.DEALS = original
    check("portfolio status: a $25,000 loan draw is not booked as collections",
          "$29,020" not in out and "$1,200" in out, out)
    check("portfolio status: the month reads NO against a $11,956 target",
          "NO" in out, out)
    check("portfolio status: the excluded credit is reported, not hidden",
          "25,000" in out, out)


def main():
    print("REGRESSION TESTS (2026-08-09 sweep)")
    test_solve_not_false_unreachable()
    test_no_phantom_debt_service()
    test_waterfall_invalid_flag()
    test_mc_vacancy_centered_on_underwriting()
    test_rentcast_matcher_requires_all_tokens()
    print("REGRESSION TESTS (2026-08-24 sweep)")
    test_discovery_prefers_newest_and_reports_alternates()
    test_solver_returns_a_price_that_clears_its_own_target()
    test_irr_requires_a_sign_change()
    test_mc_growth_recentered_on_underwriting()
    test_part_year_t12_is_flagged_not_read_as_annual()
    test_rent_roll_reports_its_own_implied_vacancy()
    test_flood_degrades_to_unknown_not_clear()
    test_rentcast_matches_by_address_hint()
    test_beats_index_is_deterministic()
    print("REGRESSION TESTS (2026-09-07 sweep)")
    test_unknown_model_input_is_refused()
    test_insurance_gate_needs_an_artifact_not_prose()
    test_bankpackage_diligence_gates_fail_closed()
    test_intake_apply_end_to_end_preserves_the_workbook()
    test_solve_price_returns_the_basis_it_certified()
    test_commercial_share_is_a_valid_model_input()
    test_hedonic_market_comes_from_the_parish()
    test_hedonic_has_no_year_regressor()
    test_flood_cli_does_not_print_no_for_undetermined()
    test_latax_rejects_a_percentage_typed_as_a_percent()
    test_market_parish_coverage_matches_latax()
    test_irr_verdict_applies_the_pursue_floor()
    test_icmemo_tornado_keeps_every_downside_row()
    test_portfolio_sim_charges_idle_capital_once()
    test_portfolio_status_excludes_non_operating_credits()
    if FAILURES:
        print(f"\nRESULT: {len(FAILURES)} FAILED: {FAILURES}")
        sys.exit(1)
    print("\nRESULT: all regression tests passed")


if __name__ == "__main__":
    main()
