"""Recalculate a workbook with LibreOffice headless, then scan for formula
errors and read the Checks tab. openpyxl drops cached values on save, so this
step is mandatory before any number is trusted."""

import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl

from cellmap import ERROR_TOKENS

SOFFICE_CANDIDATES = [
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/opt/homebrew/bin/soffice",
    "/usr/local/bin/soffice",
]

INSTALL_HINT = (
    "LibreOffice not found. Install it (required to recalculate formulas):\n"
    "  brew install --cask libreoffice\n"
    "or download from https://www.libreoffice.org/download/\n"
    "Without it, written inputs are saved but every formula reads as empty."
)


class RecalcUnavailable(Exception):
    pass


def find_soffice():
    for p in SOFFICE_CANDIDATES:
        if Path(p).exists():
            return p
    for name in ("soffice", "libreoffice"):
        w = shutil.which(name)
        if w:
            return w
    return None


def recalc(path, timeout=180):
    """Recalculate in place. Returns the path. Raises RecalcUnavailable."""
    soffice = find_soffice()
    if not soffice:
        raise RecalcUnavailable(INSTALL_HINT)
    path = Path(path).resolve()
    with tempfile.TemporaryDirectory() as td:
        cmd = [soffice, "--headless", "--norestore", "--convert-to", "xlsx",
               "--outdir", td, str(path)]
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        produced = Path(td) / path.name
        if not produced.exists():
            raise RecalcUnavailable(
                f"LibreOffice produced no output.\nstdout: {proc.stdout}\nstderr: {proc.stderr}"
            )
        shutil.copy2(produced, path)
    return path


def scan_errors(path):
    """-> list of 'Sheet!Cell = #ERR' strings."""
    wb = openpyxl.load_workbook(path, data_only=True)
    found = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.strip() in ERROR_TOKENS:
                    found.append(f"{ws.title}!{c.coordinate} = {v.strip()}")
    return found


def read_checks(path, spec):
    sheet, r0, r1 = spec["checks"]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    out = []
    for r in range(r0, r1 + 1):
        out.append((ws.cell(r, 1).value, ws.cell(r, 2).value))
    return out


def read_values(path, refs):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for name, ref in refs.items():
        sheet, coord = ref.split("!")
        out[name] = wb[sheet][coord].value
    return out
