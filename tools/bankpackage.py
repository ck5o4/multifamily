#!/usr/bin/env python3
"""Bank package generator - a professional loan-request document for a deal.

Modeled on the Stoa bank package structure (stats-forward, per-unit metrics,
timeline, clear repayment story), scaled to a single-property acquisition
request. Every number is pulled live from the deal's recalculated workbook -
nothing is typed twice, nothing can drift from the underwriting.

Output is print-ready HTML: open it, Cmd+P, save as PDF, send to the banker.
Personal items the system cannot know (bio, photos) are marked [FILL].

Refuses (exit 1) if the workbook has failing checks or formula errors - never
send a banker numbers that don't tie. --force overrides, stamping a prominent
warning into the HTML header.

Usage:
  python3 tools/bankpackage.py <deal-name> [--force]
  # requires deal-intake/<deal>/<deal>_acq.xlsx to exist and be recalculated
"""

import json
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import recalc as rc
from cellmap import ACQ

ROOT = Path(__file__).resolve().parent.parent


def money(v):
    return f"${v:,.0f}" if isinstance(v, (int, float)) else "-"


def pct(v, dp=2):
    return f"{v*100:.{dp}f}%" if isinstance(v, (int, float)) else "-"


def x2(v):
    return f"{v:.2f}x" if isinstance(v, (int, float)) else "-"


def n0(v):
    return f"{v:.0f}" if isinstance(v, (int, float)) else "-"


def gather(deal, force=False):
    wb = ROOT / "deal-intake" / deal / f"{deal}_acq.xlsx"
    if not wb.exists():
        sys.exit(f"no workbook at {wb} - run intake.py --deal {deal} ... --recalc first")
    v = rc.read_values(wb, ACQ["outputs"])
    s = rc.read_values(wb, ACQ["scalars"])
    checks = rc.read_checks(wb, ACQ)
    errors = rc.scan_errors(wb)
    failing = [f"check: {label} -> {st}" for label, st in checks if st != "OK"]
    failing += [f"formula error: {e}" for e in errors]
    if failing:
        print("Workbook has failing checks or formula errors - a banker can pick "
              "these numbers apart:")
        for f in failing:
            print(f"  {f}")
        if not force:
            sys.exit("REFUSING to generate. Fix the workbook (recalc), or rerun "
                     "with --force to generate anyway with a warning banner.")
        print("--force given: generating anyway with a prominent warning in the "
              "package header.")
    # rent roll straight from the inputs
    import openpyxl
    w = openpyxl.load_workbook(wb, data_only=True)["Inputs"]
    mix = []
    for r in range(3, 11):
        n = w.cell(r, 5).value
        if n:
            mix.append({"units": int(n), "type": w.cell(r, 6).value,
                        "sf": w.cell(r, 7).value, "rent": w.cell(r, 8).value})
    # 5-yr NOI / DSCR path from Annual CF
    cf = openpyxl.load_workbook(wb, data_only=True)["Annual CF"]
    years = []
    for j in range(4, 9):  # D..H = years 1..5
        years.append({"yr": cf.cell(2, j).value, "noi": cf.cell(25, j).value,
                      "ds": cf.cell(34, j).value, "dscr": cf.cell(38, j).value,
                      "cf": cf.cell(37, j).value})
    # portfolio record if tracked
    pd = {}
    pf = ROOT / "portfolio" / "deals.json"
    if pf.exists():
        pd = json.loads(pf.read_text()).get(deal, {})
    return wb, v, s, mix, years, pd, failing


def build_html(deal, v, s, mix, years, pd, failing=None):
    price = s.get("purchase_price")
    loan = v.get("loan_amount")
    reno = v.get("reno_budget") or 0
    equity = v.get("total_equity")
    units = v.get("total_units")
    today = date.today().strftime("%B %Y")

    def row(label, val, bold=False):
        w = "font-weight:700;" if bold else ""
        return f"<tr><td style='{w}'>{label}</td><td style='text-align:right;{w}'>{val}</td></tr>"

    mix_rows = "".join(
        f"<tr><td>{m['units']}</td><td>{m['type']}</td><td style='text-align:right'>{m['sf']:,}</td>"
        f"<td style='text-align:right'>{money(m['rent'])}</td></tr>" for m in mix)
    yr_rows = "".join(
        f"<tr><td>Year {y['yr']}</td><td style='text-align:right'>{money(y['noi'])}</td>"
        f"<td style='text-align:right'>{money(abs(y['ds'])) if y['ds'] else '-'}</td>"
        f"<td style='text-align:right'>{x2(y['dscr'])}</td>"
        f"<td style='text-align:right'>{money(y['cf'])}</td></tr>"
        for y in years if isinstance(y["noi"], (int, float)))

    reno_block = ""
    if reno:
        reno_block = f"""
  <h2>Business Plan - Value-Add Program</h2>
  <table>
    {row('Units to renovate', n0(s.get('reno_units')))}
    {row('Budget per unit', money(s.get('reno_cost_per_unit')))}
    {row('Total renovation budget (funded at close, in equity below)', money(reno), bold=True)}
    {row('Rent premium per unit', money(s.get('reno_premium_month')) + '/month')}
    {row('Pace', n0(s.get('reno_units_per_year')) + ' units/year starting year ' + n0(s.get('reno_start_year')))}
  </table>
  <p class="note">The renovation program is fully funded with cash at closing - no draw
  requests, no construction risk to the lender.</p>"""

    refi_block = ""
    if s.get("refi_year"):
        refi_block = f"""
  <h2>Repayment / Takeout Plan</h2>
  <table>
    {row('Planned refinance', 'Year ' + n0(s.get('refi_year')))}
    {row('Projected value at refi (NOI / ' + pct(s.get('refi_valuation_cap')) + ')', money(v.get('refi_value')))}
    {row('Projected takeout loan', money(v.get('refi_new_loan')))}
    {row('This loan repaid in full', money(v.get('refi_old_payoff')), bold=True)}
  </table>
  <p class="note">The refinance retires this note before any balloon. If market
  conditions do not support the refinance, year-5 sale at a
  {pct(s.get('exit_cap'))} exit cap also retires the note (see pro forma).</p>"""

    html = f"""<!doctype html><html><head><meta charset="utf-8">
<title>{deal} - Loan Request</title>
<style>
  @page {{ margin: 1in; }}
  html, body {{ background: #ffffff !important; }}
  body {{ font-family: Georgia, 'Times New Roman', serif; color: #1a1a1a;
         max-width: 7.5in; margin: 0 auto; line-height: 1.45; font-size: 11pt;
         padding: 0 24px; }}
  .cover {{ text-align:center; padding: 2.5in 0 1in; page-break-after: always; }}
  .cover h1 {{ font-size: 26pt; margin: 0 0 8px; letter-spacing: 0.5px; }}
  .cover .sub {{ font-size: 13pt; color:#555; }}
  .cover .ask {{ font-size: 18pt; margin-top: 40px; font-weight: 700; }}
  h2 {{ font-size: 13pt; text-transform: uppercase; letter-spacing: 1.5px;
       border-bottom: 2px solid #1a1a1a; padding-bottom: 4px; margin-top: 28px; }}
  table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
  td, th {{ padding: 5px 8px; border-bottom: 1px solid #ddd; font-size: 10.5pt; }}
  th {{ text-align:left; background:#f2f2f2; text-transform: uppercase;
       font-size: 8.5pt; letter-spacing: 1px; }}
  .note {{ font-size: 9.5pt; color: #555; font-style: italic; }}
  .fill {{ background: #fff3cd; padding: 2px 6px; font-family: monospace; font-size: 9pt; }}
  .kpis {{ display:flex; gap:16px; margin:16px 0; }}
  .kpi {{ flex:1; border:1px solid #ccc; padding:12px; text-align:center; }}
  .kpi .n {{ font-size: 16pt; font-weight:700; }}
  .kpi .l {{ font-size: 8.5pt; text-transform:uppercase; letter-spacing:1px; color:#666; }}
</style></head><body>
{('<div style="border:3px solid #b00020; color:#b00020; padding:12px 16px; '
  'margin:16px 0; font-weight:700;">WARNING - GENERATED WITH --force: the '
  'underwriting workbook has FAILING CHECKS / FORMULA ERRORS. Do NOT send to '
  'a banker until fixed.<br>' + '<br>'.join(failing) + '</div>') if failing else ''}
<div class="cover">
  <h1>{deal.replace('-', ' ').title()}</h1>
  <div class="sub">{n0(units)}-Unit Multifamily Acquisition &mdash; Louisiana</div>
  <div class="sub">{today}</div>
  <div class="ask">Loan Request: {money(loan)}</div>
  <div class="sub" style="margin-top:60px">Prepared by <span class="fill">[FILL: your name / entity LLC]</span></div>
</div>

<h2>The Request</h2>
<div class="kpis">
  <div class="kpi"><div class="n">{money(loan)}</div><div class="l">Loan Amount</div></div>
  <div class="kpi"><div class="n">{pct(s.get('ltv'), 0)}</div><div class="l">Max LTV</div></div>
  <div class="kpi"><div class="n">{x2(v.get('dscr_yr1'))}</div><div class="l">Year-1 DSCR</div></div>
  <div class="kpi"><div class="n">{money(equity)}</div><div class="l">Cash In (Sponsor Side)</div></div>
</div>
<table>
  {row('Purchase price', money(price))}
  {row('Requested loan', money(loan), bold=True)}
  {row('Assumed terms for sizing', pct(s.get('rate')) + ' / ' + n0(s.get('amort_years')) + '-yr amortization')}
  {row('Year-1 DSCR at these terms', x2(v.get('dscr_yr1')) + ' (post-reserve ' + x2(v.get('dscr_yr1_post_reserve')) + ')')}
  {row('Going-in cap rate', pct(v.get('going_in_cap')))}
</table>

<h2>Sponsor</h2>
<p><span class="fill">[FILL: 3-4 sentence bio - background, Stoa Group experience
(1,995-unit Gulf Coast developer/operator), why this market]</span></p>
<table>
  {row('Total cash committed to this acquisition', money(equity), bold=True)}
  {row('Capital source', 'Family investor equity + sponsor cash')}
  {row('Property management', '<span class="fill">[FILL: third-party PM company name]</span> (professional third-party management)')}
</table>

<h2>The Property</h2>
<p><span class="fill">[FILL: address, year built, construction type, 2-3 sentences on
location/condition]</span></p>
<table>
  <tr><th>Units</th><th>Type</th><th>SF</th><th>In-Place Rent</th></tr>
  {mix_rows}
</table>
{reno_block}

<h2>Sources &amp; Uses</h2>
<table>
  <tr><th colspan="2">Sources</th></tr>
  {row('Loan proceeds', money(loan))}
  {row('Sponsor equity', money(equity))}
  {row('Total sources', money((loan or 0) + (equity or 0)), bold=True)}
  <tr><th colspan="2">Uses</th></tr>
  {row('Purchase price', money(price))}
  {row('Closing costs', money((price or 0) * (s.get('closing_pct') or 0)))}
  {row('Loan fee', money((loan or 0) * (s.get('loan_fee_pct') or 0)))}
  {row('Renovation budget', money(reno))}
  {row('Operating / debt-service reserve', money(max(0, (loan or 0) + (equity or 0) - (price or 0) * (1 + (s.get('closing_pct') or 0)) - (loan or 0) * (s.get('loan_fee_pct') or 0) - reno)))}
</table>

<h2>Five-Year Pro Forma</h2>
<table>
  <tr><th></th><th>NOI</th><th>Debt Service</th><th>DSCR</th><th>Cash Flow</th></tr>
  {yr_rows}
</table>
<p class="note">Underwriting assumptions: {pct(s.get('vacancy'), 0)} vacancy,
{pct(s.get('rent_growth'), 0)} rent growth, {pct(s.get('expense_growth'), 1)} expense growth,
insurance {money(s.get('insurance'))}/unit/yr (Louisiana-adjusted), property taxes
reassessed to purchase price at the parish millage - not the seller's current bill.</p>
{refi_block}

<h2>Exhibits Available on Request</h2>
<p class="note">Rent roll &middot; trailing-12 operating statement &middot; purchase agreement &middot;
personal financial statement &middot; entity documents &middot; insurance quotes &middot; market comp survey</p>

</body></html>"""
    return html


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    force = "--force" in sys.argv[1:]
    if not args:
        print(__doc__)
        return
    deal = args[0]
    wb, v, s, mix, years, pd, failing = gather(deal, force=force)
    html = build_html(deal, v, s, mix, years, pd, failing=failing)
    out = wb.parent / f"{deal}_bank_package.html"
    out.write_text(html)
    print(f"bank package -> {out}")
    print("open it, Cmd+P, 'Save as PDF'. [FILL] items are highlighted yellow.")


if __name__ == "__main__":
    main()
