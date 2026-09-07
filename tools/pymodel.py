#!/usr/bin/env python3
"""Pure-Python mirror of Multifamily_Acquisition_Model.xlsx.

Replicates every calculation in the workbook with no LibreOffice dependency.
Verified against three real deals; see test_pymodel.py.

  run(inputs)           -> full model output dict
  solve_price(inputs, target_irr) -> bisect to target IRR
  tornado(inputs)       -> sensitivity by factor
  monte_carlo(inputs)   -> P10/P50/P90 and P(IRR>=13%)

CLI:
  python3 tools/pymodel.py --deal baker-trails
  python3 tools/pymodel.py --deal baker-trails --solve 0.16
  python3 tools/pymodel.py --deal baker-trails --tornado
  python3 tools/pymodel.py --deal baker-trails --mc
"""

import argparse
import json
import math
import random
import sys
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent))

import latax


# ---------------------------------------------------------------------------
# Finance primitives (stdlib only, no numpy)
# ---------------------------------------------------------------------------

def _pv_annuity(rate_per_period, n_periods, payment):
    """PV of level payments — Excel PV(rate, nper, pmt, 0, 0) semantics.

    Returns the magnitude (positive number), matching Excel's sign convention
    for loan sizing: PV(rate/12, years*12, -monthly_payment) > 0.
    """
    if rate_per_period == 0:
        return payment * n_periods
    return payment * (1 - (1 + rate_per_period) ** (-n_periods)) / rate_per_period


def _fv_annuity(rate_per_period, n_periods, payment, pv):
    """Remaining loan balance after n_periods monthly payments.

    Replicates the workbook: MAX(0, FV(rate/12, 12, monthly_pmt, -beg_bal))
    Excel FV sign convention: pmt flows in the opposite direction from pv.
    Here pv is the beginning balance (positive) and payment is the periodic
    payment amount (positive), both from the lender's perspective.

    Formula: beg_bal*(1+r)^n - pmt*((1+r)^n - 1)/r
    """
    factor = (1 + rate_per_period) ** n_periods
    if rate_per_period == 0:
        remaining = pv - payment * n_periods
    else:
        remaining = pv * factor - payment * (factor - 1) / rate_per_period
    return max(0.0, remaining)


def _pmt(rate_per_period, n_periods, pv):
    """Monthly payment for a loan — Excel PMT(rate, nper, -pv) semantics."""
    if rate_per_period == 0:
        return pv / n_periods
    return pv * rate_per_period / (1 - (1 + rate_per_period) ** (-n_periods))


def _irr(cash_flows, guess=0.1, max_iter=200, tol=1e-8):
    """Annual IRR matching Excel IRR() semantics (Newton-Raphson with bisect fallback).

    cash_flows: list starting at year 0 (negative = outflow).
    Returns None if not found.
    """
    # An IRR exists only if the flows change sign. Without this guard the
    # solver "converges" on garbage: all-zeros returned the bisection bracket
    # midpoint (450.05%) and [0,...,0,X] passed the relative-NPV convergence
    # test at the Newton clamp (5000%). That is reachable in one step from the
    # documented lp_pct=1.0 structure (investor funds all equity -> gp_capital
    # and gp_profit are both 0), which printed "GP IRR: 450.05%". (2026-08-24)
    if not any(cf < 0 for cf in cash_flows) or not any(cf > 0 for cf in cash_flows):
        return None

    def _npv(r):
        try:
            return sum(cf / (1 + r) ** t for t, cf in enumerate(cash_flows))
        except (OverflowError, ZeroDivisionError):
            return float('nan')

    # Newton-Raphson
    r = guess
    for _ in range(max_iter):
        npv = _npv(r)
        if math.isnan(npv):
            r = guess * 0.5
            continue
        h = 1e-6
        npv_h = _npv(r + h)
        if math.isnan(npv_h):
            break
        d_npv = (npv_h - npv) / h
        if d_npv == 0:
            break
        r_new = r - npv / d_npv
        # Clamp to avoid explosive divergence
        r_new = max(-0.999, min(r_new, 50.0))
        if abs(r_new - r) < tol:
            # A clamped fixed point is not a root: verify NPV before trusting
            # Newton (audit F2 2026-08-05 — all-negative CF vectors returned
            # the 50.0 clamp as a "converged" 5000% IRR).
            scale = sum(abs(c) for c in cash_flows) or 1.0
            if abs(_npv(r_new)) <= 1e-6 * scale:
                return r_new
            break
        r = r_new

    # Bisection fallback: bracket where sign changes
    lo, hi = -0.999, 10.0
    npv_lo = _npv(lo)
    npv_hi = _npv(hi)
    if math.isnan(npv_lo) or math.isnan(npv_hi):
        return None
    if npv_lo * npv_hi > 0:
        # Try a tighter range
        for lo_try in [-0.5, -0.2, 0.0]:
            if _npv(lo_try) * npv_hi <= 0:
                lo, npv_lo = lo_try, _npv(lo_try)
                break
        else:
            return None
    for _ in range(100):
        mid = (lo + hi) / 2
        npv_mid = _npv(mid)
        if math.isnan(npv_mid):
            return None
        if abs(npv_mid) < 1e-6 or (hi - lo) < tol:
            return mid
        if npv_lo * npv_mid < 0:
            hi = mid
        else:
            lo, npv_lo = mid, npv_mid
    return (lo + hi) / 2


# ---------------------------------------------------------------------------
# Core model
# ---------------------------------------------------------------------------

def run(inputs: dict) -> dict:
    """Run the acquisition model.  Returns a rich dict of outputs.

    Required inputs:
        price           purchase price ($)
        unit_mix        list of {units, sf, rent} dicts (rent = $/month in-place)
        taxes_annual    annual property tax ($, already reassessed or as-is)

    Optional (all have defaults matching defaults.py / the workbook):
        closing_pct, loan_fee_pct, vacancy, bad_debt, concessions,
        other_income_unit_mo, rent_growth, other_income_growth,
        payroll, ga, marketing, rm, contract_services, utilities,
        other, insurance, mgmt_pct, asset_mgmt_pct, capex_unit,
        expense_growth, ltv, min_dscr, rate, amort_years, io_years,
        hold_years, exit_cap,  cost_of_sale, lp_pct, pref, residual_lp,
        refi_year, refi_valuation_cap, refi_ltv, refi_min_dscr, refi_rate,
        refi_amort_years, refi_cost_pct,
        reno_units, reno_cost_per_unit, reno_premium_month, reno_units_per_year,
        reno_downtime_months, reno_start_year
    """
    p = _merge_defaults(inputs)

    # --- Rent roll totals (Inputs!E11, G11, H11) ---
    total_units = sum(g["units"] for g in p["unit_mix"])
    if total_units == 0:
        raise ValueError("unit_mix must have at least one row with units > 0")
    wt_rent = sum(g["units"] * g["rent"] for g in p["unit_mix"]) / total_units
    # wt_sf not needed directly but consistent with workbook

    price = p["price"]
    closing_pct = p["closing_pct"]
    loan_fee_pct = p["loan_fee_pct"]
    vacancy = p["vacancy"]
    bad_debt = p["bad_debt"]
    concessions = p["concessions"]
    other_income_unit_mo = p["other_income_unit_mo"]
    rent_growth = p["rent_growth"]
    other_income_growth = p["other_income_growth"]
    payroll = p["payroll"]
    ga = p["ga"]
    marketing = p["marketing"]
    rm = p["rm"]
    contract_services = p["contract_services"]
    utilities = p["utilities"]
    other_exp = p["other"]
    insurance = p["insurance"]
    taxes_annual = p["taxes_annual"]
    mgmt_pct = p["mgmt_pct"]
    asset_mgmt_pct = p["asset_mgmt_pct"]
    capex_unit = p["capex_unit"]
    expense_growth = p["expense_growth"]
    ltv = p["ltv"]
    min_dscr = p["min_dscr"]
    rate = p["rate"]
    amort_years = p["amort_years"]
    io_years = p["io_years"]
    hold_years = p["hold_years"]
    exit_cap = p["exit_cap"]
    cost_of_sale = p["cost_of_sale"]
    lp_pct = p["lp_pct"]
    pref_rate = p["pref"]
    residual_lp = p["residual_lp"]
    refi_year = p["refi_year"]
    refi_valuation_cap = p["refi_valuation_cap"]
    refi_ltv = p["refi_ltv"]
    refi_min_dscr = p["refi_min_dscr"]
    refi_rate = p["refi_rate"]
    refi_amort_years = p["refi_amort_years"]
    refi_cost_pct = p["refi_cost_pct"]
    reno_units = p["reno_units"]
    reno_cost_per_unit = p["reno_cost_per_unit"]
    reno_premium_month = p["reno_premium_month"]
    reno_units_per_year = p["reno_units_per_year"]
    reno_downtime_months = p["reno_downtime_months"]
    reno_start_year = p["reno_start_year"]

    # --- Gross potential rent base (year 1) ---
    # Formula: SUMPRODUCT(units*rent)*12 — does NOT compound in base; year
    # exponent is (year-1), so year 1 = base * (1+g)^0 = base.
    gpr_base = sum(g["units"] * g["rent"] for g in p["unit_mix"]) * 12  # Inputs!H11 * total_units

    # --- Renovation budget ---
    reno_budget = reno_units * reno_cost_per_unit

    # --- Year 1 NOI for loan sizing ---
    # Must compute year-1 NOI first to size the loan.
    noi_y1 = _noi(1, gpr_base, wt_rent, total_units,
                  rent_growth, vacancy, bad_debt, concessions,
                  other_income_unit_mo, other_income_growth,
                  payroll, ga, marketing, rm, contract_services,
                  utilities, other_exp, insurance, taxes_annual,
                  mgmt_pct, expense_growth,
                  reno_units, reno_units_per_year, reno_start_year,
                  reno_premium_month, reno_downtime_months)

    # --- Loan sizing: MIN(LTV*price, PV of DSCR-constrained payment) ---
    # Workbook: MIN(B32*B2, PV(B34/12, B35*12, -(NOI_y1/min_dscr)/12))
    ltv_loan = ltv * price
    monthly_rate = rate / 12
    n_payments = amort_years * 12
    dscr_pmt = (noi_y1 / min_dscr) / 12   # max monthly payment DSCR allows
    dscr_loan = _pv_annuity(monthly_rate, n_payments, dscr_pmt)
    loan_amount = max(0.0, min(ltv_loan, dscr_loan))
    if noi_y1 <= 0:
        print("WARNING [run]: Year-1 NOI <= 0 — no loan is supportable; "
              "outputs are all-equity nonsense. Check inputs.", file=sys.stderr)

    monthly_pmt = _pmt(monthly_rate, n_payments, loan_amount)
    annual_debt_service = monthly_pmt * 12  # used in non-IO years

    # --- Total equity required ---
    total_equity = price + price * closing_pct + loan_amount * loan_fee_pct - loan_amount + reno_budget
    lp_capital = total_equity * lp_pct
    gp_capital = total_equity - lp_capital

    # --- Going-in cap rate ---
    going_in_cap = noi_y1 / price if price > 0 else 0.0

    # --- Refi block (Inputs!F23:F28) ---
    refi_value = 0.0
    refi_new_loan = 0.0
    refi_old_payoff = 0.0
    refi_cash_out = 0.0
    refi_new_monthly_pmt = 0.0
    refi_new_annual_ds = 0.0

    if refi_year > 0:
        # Refi value = NOI at refi_year / refi_valuation_cap
        noi_refi = _noi(refi_year, gpr_base, wt_rent, total_units,
                        rent_growth, vacancy, bad_debt, concessions,
                        other_income_unit_mo, other_income_growth,
                        payroll, ga, marketing, rm, contract_services,
                        utilities, other_exp, insurance, taxes_annual,
                        mgmt_pct, expense_growth,
                        reno_units, reno_units_per_year, reno_start_year,
                        reno_premium_month, reno_downtime_months)
        refi_value = noi_refi / refi_valuation_cap if refi_valuation_cap > 0 else 0.0

        # New loan: MIN(LTV*value, PV of DSCR-constrained payment at refi NOI)
        refi_dscr_pmt = (noi_refi / refi_min_dscr) / 12
        refi_monthly_rate = refi_rate / 12
        refi_n_pmt = refi_amort_years * 12
        refi_dscr_loan = _pv_annuity(refi_monthly_rate, refi_n_pmt, refi_dscr_pmt)
        refi_new_loan = min(refi_ltv * refi_value, refi_dscr_loan)

        refi_new_monthly_pmt = _pmt(refi_monthly_rate, refi_n_pmt, refi_new_loan)
        refi_new_annual_ds = refi_new_monthly_pmt * 12

        # Old payoff = ending balance of original loan at end of refi_year
        refi_old_payoff = _loan_balance_at(loan_amount, monthly_rate, monthly_pmt, refi_year, io_years)

        refi_costs = refi_new_loan * refi_cost_pct
        refi_cash_out = refi_new_loan - refi_old_payoff - refi_costs

    # --- Annual cash flows (years 1..hold_years) ---
    years = list(range(1, hold_years + 1))

    # Debt schedule per year
    beg_bal = {}
    end_bal = {}
    interest_pmt = {}
    principal_pmt = {}
    total_ds = {}

    for yr in years:
        if yr == 1:
            bb = loan_amount
        elif refi_year > 0 and yr == refi_year + 1:
            bb = refi_new_loan
        else:
            bb = end_bal[yr - 1]
        beg_bal[yr] = bb

        # Determine which rate/payment applies (post-refi vs original)
        post_refi = refi_year > 0 and yr > refi_year
        if post_refi:
            yr_rate = refi_rate
            yr_pmt = refi_new_monthly_pmt
        else:
            yr_rate = rate
            yr_pmt = monthly_pmt

        # Interest-only check (original IO period only applies pre-refi)
        if (not post_refi) and yr <= io_years:
            # IO: balance stays flat; interest = balance * rate (annual)
            interest_pmt[yr] = bb * yr_rate
            principal_pmt[yr] = 0.0
            end_bal[yr] = bb
        elif bb <= 0.01:
            # Loan already retired: no phantom payments on a zero balance
            # (sweep 2026-08-09: amort_years < hold_years charged full DS forever).
            # Dollar-scaled epsilon, not 1e-9: _fv_annuity leaves floating-point
            # residue at the final amortization year that is routinely larger
            # than 1e-9, so the 1e-9 guard missed it and the NEXT year charged a
            # phantom month on a repaid loan (cannon $20,729 at amort+1). (2026-08-24)
            interest_pmt[yr] = 0.0
            principal_pmt[yr] = 0.0
            end_bal[yr] = 0.0
        else:
            # Amortizing: use FV to find ending balance
            if post_refi:
                eb = _fv_annuity(refi_rate / 12, 12, yr_pmt, bb)
            else:
                eb = _fv_annuity(rate / 12, 12, monthly_pmt, bb)
            end_bal[yr] = eb
            if eb <= 0.01:
                # Retires mid-year: charge only the payments actually made,
                # not 12 (12*pmt - bb overstates interest on the short year).
                r_m = yr_rate / 12
                if r_m > 0 and yr_pmt > bb * r_m:
                    n_pay = math.ceil(math.log(yr_pmt / (yr_pmt - bb * r_m))
                                      / math.log(1 + r_m))
                else:
                    n_pay = math.ceil(bb / yr_pmt) if yr_pmt > 0 else 0
                n_pay = min(max(n_pay, 0), 12)
                interest_pmt[yr] = max(n_pay * yr_pmt - bb, 0.0)
                principal_pmt[yr] = bb
                end_bal[yr] = 0.0
            else:
                # Interest = annual_pmt*12 - principal_reduction (or directly from FV formula)
                # Excel: interest = 12*pmt - (beg_bal - end_bal)
                interest_pmt[yr] = 12 * yr_pmt - (bb - eb)
                principal_pmt[yr] = bb - eb

        total_ds[yr] = interest_pmt[yr] + principal_pmt[yr]

    # Income and expense by year
    egr = {}
    nri = {}
    other_income = {}
    noi = {}
    capex_reserve = {}
    asset_mgmt = {}
    cfbds = {}   # cash flow before debt service
    cfads = {}   # cash flow after debt service
    dscr = {}
    dscr_post_capex = {}

    for yr in years:
        noi_yr = _noi(yr, gpr_base, wt_rent, total_units,
                      rent_growth, vacancy, bad_debt, concessions,
                      other_income_unit_mo, other_income_growth,
                      payroll, ga, marketing, rm, contract_services,
                      utilities, other_exp, insurance, taxes_annual,
                      mgmt_pct, expense_growth,
                      reno_units, reno_units_per_year, reno_start_year,
                      reno_premium_month, reno_downtime_months)
        noi[yr] = noi_yr

        # EGR and other income for reporting
        egr[yr] = _egr(yr, gpr_base, wt_rent, total_units,
                       rent_growth, vacancy, bad_debt, concessions,
                       other_income_unit_mo, other_income_growth,
                       reno_units, reno_units_per_year, reno_start_year,
                       reno_premium_month, reno_downtime_months)

        cr = -total_units * capex_unit * (1 + expense_growth) ** (yr - 1)
        capex_reserve[yr] = cr
        am = -egr[yr] * asset_mgmt_pct
        asset_mgmt[yr] = am
        cfbds[yr] = noi_yr + cr + am
        cfads[yr] = cfbds[yr] - total_ds[yr]
        dscr[yr] = noi_yr / total_ds[yr] if total_ds[yr] > 0 else 0.0
        dscr_post_capex[yr] = (noi_yr + cr) / total_ds[yr] if total_ds[yr] > 0 else 0.0

    # --- Exit (sale at end of hold_years) ---
    # Sale price = forward NOI (year hold+1) / exit_cap
    noi_fwd = _noi(hold_years + 1, gpr_base, wt_rent, total_units,
                   rent_growth, vacancy, bad_debt, concessions,
                   other_income_unit_mo, other_income_growth,
                   payroll, ga, marketing, rm, contract_services,
                   utilities, other_exp, insurance, taxes_annual,
                   mgmt_pct, expense_growth,
                   reno_units, reno_units_per_year, reno_start_year,
                   reno_premium_month, reno_downtime_months)
    sale_price = noi_fwd / exit_cap if exit_cap > 0 else 0.0
    cost_of_sale_amt = sale_price * cost_of_sale
    # Audit F1 2026-08-05: a refi landing IN the sale year never rolls the debt
    # schedule to the new loan — exit must pay off the refi loan, not the old
    # balance (which refi_cash_out already netted out). Without this, ~$271K of
    # phantom cash appeared on eden at refi_year=hold_years (+7 IRR pts).
    if refi_year > 0 and refi_year == hold_years:
        loan_payoff = refi_new_loan
    else:
        loan_payoff = end_bal[hold_years]
    if refi_year > hold_years:
        raise ValueError(f"refi_year ({refi_year}) is beyond hold_years "
                         f"({hold_years}) — refi would never occur; remove it "
                         "or shorten to within the hold.")
    net_sale_proceeds = sale_price - cost_of_sale_amt - loan_payoff

    # --- Refi proceeds in the waterfall year ---
    refi_proceeds = {}
    for yr in years:
        if refi_year > 0 and yr == refi_year:
            refi_proceeds[yr] = refi_cash_out
        else:
            refi_proceeds[yr] = 0.0

    # --- Levered cash flows (year 0 + years 1..hold) ---
    # Year 0: -(price + closing + loan_fee - loan + reno_budget)
    # Year t: cfads[t] + refi_proceeds[t]  (+ exit items in hold year)
    lev_cf = [0.0] * (hold_years + 1)
    lev_cf[0] = -(price * (1 + closing_pct) + loan_amount * loan_fee_pct - loan_amount + reno_budget)

    for yr in years:
        lev_cf[yr] = (cfads[yr] if yr <= hold_years else 0.0) + refi_proceeds[yr]
    # Add exit at hold year
    lev_cf[hold_years] += sale_price - cost_of_sale_amt - loan_payoff

    # --- Unlevered cash flows ---
    unlev_cf = [0.0] * (hold_years + 1)
    unlev_cf[0] = -(price * (1 + closing_pct) + reno_budget)  # C50 formula: C42+C43+C52
    for yr in years:
        unlev_cf[yr] = (cfbds[yr] if yr <= hold_years else 0.0)
    unlev_cf[hold_years] += sale_price - cost_of_sale_amt

    levered_irr = _irr(lev_cf)
    unlevered_irr = _irr(unlev_cf)
    total_profit = sum(lev_cf)
    equity_multiple = 1 + total_profit / (-lev_cf[0]) if lev_cf[0] < 0 else 0.0
    avg_coc = sum(cfads[yr] for yr in years) / hold_years / total_equity

    # --- Waterfall ---
    wf = _waterfall(lev_cf, lp_capital, gp_capital, pref_rate, residual_lp, hold_years)

    return {
        # Loan
        "loan_amount": loan_amount,
        "monthly_pmt": monthly_pmt,
        "total_equity": total_equity,
        "lp_capital": lp_capital,
        "gp_capital": gp_capital,
        "going_in_cap": going_in_cap,
        # NOI and cash flows by year
        "noi": noi,            # dict yr->value
        "egr": egr,
        "capex_reserve": capex_reserve,
        "cfbds": cfbds,
        "cfads": cfads,
        "total_ds": total_ds,
        "beg_bal": beg_bal,
        "end_bal": end_bal,
        "dscr": dscr,
        "dscr_post_capex": dscr_post_capex,
        # Exit
        "noi_fwd": noi_fwd,
        "sale_price": sale_price,
        "cost_of_sale_amt": cost_of_sale_amt,
        "loan_payoff": loan_payoff,
        "net_sale_proceeds": net_sale_proceeds,
        # Returns
        "levered_irr": levered_irr,
        "unlevered_irr": unlevered_irr,
        "equity_multiple": equity_multiple,
        "total_profit": total_profit,
        "avg_coc": avg_coc,
        "lev_cf": lev_cf,
        "unlev_cf": unlev_cf,
        # Refi
        "refi_value": refi_value,
        "refi_new_loan": refi_new_loan,
        "refi_old_payoff": refi_old_payoff,
        "refi_cash_out": refi_cash_out,
        # Waterfall
        "lp_irr": wf["lp_irr"],
        "lp_em": wf["lp_em"],
        "lp_profit": wf["lp_profit"],
        "gp_irr": wf["gp_irr"],
        "gp_em": wf["gp_em"],
        "gp_profit": wf["gp_profit"],
        "waterfall_invalid": wf.get("waterfall_invalid", False),
        "waterfall_neg_years": wf.get("waterfall_neg_years", []),
        "wf_detail": wf["detail"],
        # Reno
        "reno_budget": reno_budget,
    }


def _reno_units_cumulative(yr, reno_units, reno_units_per_year, reno_start_year):
    """Units renovated by end of year yr (workbook: MIN(reno_units, MAX(0,(yr-start+1)*per_yr)))."""
    if reno_units == 0 or reno_units_per_year == 0:
        return 0.0
    return min(reno_units, max(0.0, (yr - reno_start_year + 1) * reno_units_per_year))


def _reno_premium_this_year(yr, reno_units, reno_units_per_year, reno_start_year,
                             reno_premium_month, rent_growth, wt_rent,
                             reno_downtime_months):
    """Annual rent premium added by renovation, plus downtime deduction.

    Returns (premium_annual, downtime_loss_annual) — both positive magnitudes.
    Downtime loss is deducted from vacancy in the workbook (row 6 formula).
    """
    cum_now = _reno_units_cumulative(yr, reno_units, reno_units_per_year, reno_start_year)
    cum_prev = _reno_units_cumulative(yr - 1, reno_units, reno_units_per_year, reno_start_year)
    units_this_yr = cum_now - cum_prev  # newly renovated this year

    g = (1 + rent_growth) ** (yr - 1)
    # Premium on all completed renovations (same formula as row 63)
    premium = cum_now * reno_premium_month * 12 * g
    # Downtime: units being turned × downtime months × avg rent (this year's)
    downtime_loss = units_this_yr * reno_downtime_months * wt_rent * g

    return premium, downtime_loss


def _egr(yr, gpr_base, wt_rent, total_units,
          rent_growth, vacancy, bad_debt, concessions,
          other_income_unit_mo, other_income_growth,
          reno_units, reno_units_per_year, reno_start_year,
          reno_premium_month, reno_downtime_months):
    """Effective Gross Revenue = NRI + Other Income (rows 9+10)."""
    g = (1 + rent_growth) ** (yr - 1)

    # GPR including reno premium (workbook row 5)
    cum = _reno_units_cumulative(yr, reno_units, reno_units_per_year, reno_start_year)
    gpr = (gpr_base + cum * reno_premium_month * 12) * g

    # Vacancy + downtime (row 6): -(gpr * vacancy_pct + downtime_loss)
    cum_prev = _reno_units_cumulative(yr - 1, reno_units, reno_units_per_year, reno_start_year)
    units_turned = cum - cum_prev
    downtime_loss = units_turned * reno_downtime_months * wt_rent * g
    physical_vac = gpr * vacancy

    # Bad debt and concessions on base GPR (rows 7, 8) — workbook applies to gpr, not gpr+premium
    bad = gpr * bad_debt
    conc = gpr * concessions

    # NRI = gpr - vacancy - downtime - bad_debt - concessions
    nri = gpr - physical_vac - downtime_loss - bad - conc

    # Other income: units * oi_mo * 12 * (1 - vacancy) * (1+other_growth)^(yr-1)
    og = (1 + other_income_growth) ** (yr - 1)
    oi = total_units * other_income_unit_mo * 12 * (1 - vacancy) * og

    return nri + oi


def _noi(yr, gpr_base, wt_rent, total_units,
         rent_growth, vacancy, bad_debt, concessions,
         other_income_unit_mo, other_income_growth,
         payroll, ga, marketing, rm, contract_services,
         utilities, other_exp, insurance, taxes_annual,
         mgmt_pct, expense_growth,
         reno_units, reno_units_per_year, reno_start_year,
         reno_premium_month, reno_downtime_months):
    """Net Operating Income for year yr.  Matches workbook rows 11+24."""
    egr = _egr(yr, gpr_base, wt_rent, total_units,
               rent_growth, vacancy, bad_debt, concessions,
               other_income_unit_mo, other_income_growth,
               reno_units, reno_units_per_year, reno_start_year,
               reno_premium_month, reno_downtime_months)

    xg = (1 + expense_growth) ** (yr - 1)

    # Per-unit expenses (rows 14-21)
    exp_pu = (payroll + ga + marketing + rm + contract_services +
               utilities + other_exp + insurance) * total_units * xg

    # Property taxes: flat total grown at expense_growth (row 22)
    tax_exp = taxes_annual * xg

    # Management fee: % of EGR (row 23)
    mgmt = egr * mgmt_pct

    total_opex = -(exp_pu + tax_exp + mgmt)
    return egr + total_opex  # EGR + (negative) opex


def _loan_balance_at(loan_amount, monthly_rate, monthly_pmt, year, io_years):
    """Ending balance of the original loan at end of `year`."""
    bal = loan_amount
    for yr in range(1, year + 1):
        if yr <= io_years:
            pass  # IO: balance unchanged
        else:
            bal = _fv_annuity(monthly_rate, 12, monthly_pmt, bal)
    return bal


def _waterfall(lev_cf, lp_capital, gp_capital, pref_rate, residual_lp, hold_years):
    """LP/GP waterfall matching the Waterfall tab exactly.

    Tier 1: LP pref (8%) accrues on unreturned capital, paid from distributable CF
    Tier 2: Return of LP capital
    Tier 3: Return of GP capital
    Tier 4: Residual split (residual_lp / (1-residual_lp))

    distributable CF = lev_cf[yr] for yr >= 1 (year 0 is the equity outflow).
    LP initial investment = -lp_capital (outflow at year 0).
    """
    years = list(range(1, hold_years + 1))

    # Year 0 outflows for IRR calc
    lp_flows = [-lp_capital] + [0.0] * hold_years
    gp_flows = [-gp_capital] + [0.0] * hold_years

    # Waterfall state
    pref_owed_beg = {}
    pref_accrual = {}
    pref_paid = {}
    pref_owed_end = {}
    lp_unreturned_beg = {}
    lp_roc = {}
    lp_unreturned_end = {}
    gp_unreturned_beg = {}
    gp_roc = {}
    gp_unreturned_end = {}
    residual_cf = {}
    lp_residual = {}
    gp_residual = {}
    lp_net = {}
    gp_net = {}

    neg_years = [yr for yr in years if lev_cf[yr] < 0]
    if neg_years:
        # Audit F4 2026-08-05: tiers floor at 0, so deficit years vanish from
        # partner flows — LP/GP can be "paid" cash the project never generated.
        # (Faithful to the workbook's MIN(MAX(...,0)) formulas, but economically
        # wrong in distress.) Until deficits carry forward, flag loudly.
        print(f"WARNING [waterfall]: negative distributable cash in year(s) "
              f"{neg_years} — LP/GP splits ignore the deficit and overstate "
              "partner returns. Do not quote LP numbers from this run.",
              file=sys.stderr)

    for yr in years:
        dist = lev_cf[yr]   # distributable cash flow this year

        # Pref owed beginning
        if yr == 1:
            pref_owed_beg[yr] = 0.0
            lp_unreturned_beg[yr] = lp_capital
            gp_unreturned_beg[yr] = gp_capital
        else:
            pref_owed_beg[yr] = pref_owed_end[yr - 1]
            lp_unreturned_beg[yr] = lp_unreturned_end[yr - 1]
            gp_unreturned_beg[yr] = gp_unreturned_end[yr - 1]

        # Pref accrues on UNRETURNED LP capital (not original capital)
        pref_accrual[yr] = lp_unreturned_beg[yr] * pref_rate if yr <= hold_years else 0.0
        pref_paid[yr] = min(max(dist, 0.0), pref_owed_beg[yr] + pref_accrual[yr])
        pref_owed_end[yr] = pref_owed_beg[yr] + pref_accrual[yr] - pref_paid[yr]

        # Tier 2: LP return of capital
        lp_roc[yr] = min(max(dist - pref_paid[yr], 0.0), lp_unreturned_beg[yr])
        lp_unreturned_end[yr] = lp_unreturned_beg[yr] - lp_roc[yr]

        # Tier 3: GP return of capital
        gp_roc[yr] = min(max(dist - pref_paid[yr] - lp_roc[yr], 0.0), gp_unreturned_beg[yr])
        gp_unreturned_end[yr] = gp_unreturned_beg[yr] - gp_roc[yr]

        # Tier 4: Residual
        residual_cf[yr] = max(dist - pref_paid[yr] - lp_roc[yr] - gp_roc[yr], 0.0)
        lp_residual[yr] = residual_cf[yr] * residual_lp
        gp_residual[yr] = residual_cf[yr] * (1.0 - residual_lp)

        # Net distributions
        lp_net[yr] = pref_paid[yr] + lp_roc[yr] + lp_residual[yr]
        gp_net[yr] = gp_roc[yr] + gp_residual[yr]

        lp_flows[yr] = lp_net[yr]
        gp_flows[yr] = gp_net[yr]

    lp_irr = _irr(lp_flows)
    gp_irr = _irr(gp_flows)
    lp_profit = sum(lp_flows)
    gp_profit = sum(gp_flows)
    lp_em = 1 + lp_profit / lp_capital if lp_capital > 0 else 0.0
    gp_em = 1 + gp_profit / gp_capital if gp_capital > 0 else 0.0

    if neg_years:
        # The tier math floored the deficit years at 0, so partner flows contain
        # phantom cash the project never distributed. Null the metrics rather
        # than return numbers the warning says not to quote (sweep 2026-08-09).
        lp_irr = None
        gp_irr = None
        lp_em = None
        gp_em = None

    return {
        "waterfall_invalid": bool(neg_years),
        "waterfall_neg_years": neg_years,
        "lp_irr": lp_irr,
        "lp_em": lp_em,
        "lp_profit": lp_profit,
        "gp_irr": gp_irr,
        "gp_em": gp_em,
        "gp_profit": gp_profit,
        "detail": {
            "pref_owed_beg": pref_owed_beg,
            "pref_accrual": pref_accrual,
            "pref_paid": pref_paid,
            "pref_owed_end": pref_owed_end,
            "lp_unreturned_beg": lp_unreturned_beg,
            "lp_roc": lp_roc,
            "lp_unreturned_end": lp_unreturned_end,
            "gp_unreturned_beg": gp_unreturned_beg,
            "gp_roc": gp_roc,
            "gp_unreturned_end": gp_unreturned_end,
            "residual_cf": residual_cf,
            "lp_residual": lp_residual,
            "gp_residual": gp_residual,
            "lp_net": lp_net,
            "gp_net": gp_net,
        },
    }


def _merge_defaults(inputs):
    """Fill in defaults for any keys not supplied by the caller."""
    defaults = {
        "closing_pct": 0.02,
        "loan_fee_pct": 0.01,
        "vacancy": 0.07,
        "bad_debt": 0.005,
        "concessions": 0.0,
        "other_income_unit_mo": 40,
        "rent_growth": 0.02,
        "other_income_growth": 0.02,
        "payroll": 0,
        "ga": 480,
        "marketing": 415,
        "rm": 515,
        "contract_services": 0,
        "utilities": 755,
        "other": 105,
        "insurance": 2000,
        "taxes_annual": 0,
        "mgmt_pct": 0.08,
        "asset_mgmt_pct": 0.0,
        "capex_unit": 250,
        "expense_growth": 0.025,
        "ltv": 0.75,
        "min_dscr": 1.30,
        "rate": 0.0675,
        "amort_years": 25,
        "io_years": 0,
        "hold_years": 5,
        "exit_cap": None,  # will be set to going_in + 50bps if None
        "cost_of_sale": 0.03,
        "lp_pct": 0.90,
        "pref": 0.08,
        "residual_lp": 0.70,
        "refi_year": 0,
        "refi_valuation_cap": 0.07,
        "refi_ltv": 0.75,
        "refi_min_dscr": 1.30,
        "refi_rate": 0.0675,
        "refi_amort_years": 25,
        "refi_cost_pct": 0.02,
        "reno_units": 0,
        "reno_cost_per_unit": 0,
        "reno_premium_month": 0,
        "reno_units_per_year": 0,
        "reno_downtime_months": 1,
        "reno_start_year": 1,
    }
    # A key the engine does not read is almost always a caller typo, and a typo
    # here is silent and expensive: `run({**deal, "purchase_price": 1_251_000})`
    # returns the ask-price answer with no diagnostic. That class of scripting
    # slip already published a wrong price ladder once (baker-trails,
    # 2026-08-03 "RED-TEAM CORRECTION" in portfolio/deals.json). Refuse instead.
    # `location` and `commercial_share` are consumed by solve_price's tax
    # re-derivation rather than by run() itself, but _load_deal injects both
    # from deals.json, so they are legitimate members of a deal input dict.
    allowed = set(defaults) | {"price", "unit_mix", "location", "commercial_share"}
    unknown = sorted(set(inputs) - allowed)
    if unknown:
        raise ValueError(
            "unknown model input(s): " + ", ".join(unknown)
            + ". The engine does not read these, so they would be silently "
              "ignored and the run would answer a different question than you "
              "asked. Known inputs: " + ", ".join(sorted(allowed))
        )

    p = dict(defaults)
    p.update(inputs)

    # exit_cap: if not provided, solve going-in cap + 50bps
    if p["exit_cap"] is None:
        # Need year-1 NOI to get going-in cap
        gpr_base = sum(g["units"] * g["rent"] for g in p["unit_mix"]) * 12
        wt_rent = (sum(g["units"] * g["rent"] for g in p["unit_mix"])
                   / sum(g["units"] for g in p["unit_mix"]))
        total_units = sum(g["units"] for g in p["unit_mix"])
        noi_y1 = _noi(1, gpr_base, wt_rent, total_units,
                      p["rent_growth"], p["vacancy"], p["bad_debt"], p["concessions"],
                      p["other_income_unit_mo"], p["other_income_growth"],
                      p["payroll"], p["ga"], p["marketing"], p["rm"], p["contract_services"],
                      p["utilities"], p["other"], p["insurance"], p["taxes_annual"],
                      p["mgmt_pct"], p["expense_growth"],
                      p["reno_units"], p["reno_units_per_year"], p["reno_start_year"],
                      p["reno_premium_month"], p["reno_downtime_months"])
        going_in = noi_y1 / p["price"] if p["price"] > 0 else 0.0
        p["exit_cap"] = round(going_in + 0.005, 5)

    return p


# ---------------------------------------------------------------------------
# Solve, tornado, monte_carlo
# ---------------------------------------------------------------------------

def solve_price(inputs: dict, target_irr: float, tol: float = 0.0001,
                max_iter: int = 50):
    """Bisect purchase price to hit target_irr.

    On every iteration: taxes are re-derived from price via latax (if
    location is in inputs), and exit_cap is re-derived as going-in + 50bps
    (inputs may not override exit_cap — it is always re-derived).

    Returns {price, irr} or None if unreachable.
    """
    base = dict(inputs)
    location = base.pop("location", None)
    commercial_share = base.pop("commercial_share", 0.0)
    base.pop("exit_cap", None)   # always re-derive
    if not location:
        print("WARNING [solve_price]: no 'location' in inputs - property taxes "
              "stay FROZEN at the loaded basis across all trial prices, biasing "
              "the solve. Pass location for per-price reassessment. "
              "(This exact omission skewed live-deal ladders on 2026-08-02/03.)",
              file=sys.stderr)
    else:
        # A location that latax cannot resolve is just as biasing as no location
        # at all, but used to drop through silently to frozen taxes - the very
        # bias the warning above exists to prevent. Probe once and say so.
        # (sweep 2026-08-24; matches solve.py's guard.)
        _probe, _exp = latax.estimate_tax(base.get("price", 0) or 1, location,
                                          commercial_share)
        if _probe is None:
            print(f"WARNING [solve_price]: location {location!r} did not resolve "
                  f"({_exp}) - property taxes stay FROZEN across all trial prices, "
                  "biasing the solve optimistic. Fix the town/parish name.",
                  file=sys.stderr)

    def _inputs_at(price):
        """The exact input set a trial price is scored on."""
        p = dict(base)
        p["price"] = price
        p["exit_cap"] = None   # force re-derive
        if location:
            tax, _ = latax.estimate_tax(price, location, commercial_share)
            if tax is not None:
                p["taxes_annual"] = tax
        return p

    def _irr_at(price):
        try:
            return run(_inputs_at(price))["levered_irr"]
        except Exception:
            return None

    def _basis_at(price):
        """The tax bill and exit cap the rung was certified on, so consumers
        can re-score it on the same basis instead of the asking price's."""
        p = _inputs_at(price)
        try:
            merged = _merge_defaults(dict(p))   # resolves exit_cap=None
        except Exception:
            return {}
        return {"taxes_annual": p.get("taxes_annual"),
                "exit_cap": merged.get("exit_cap"),
                "basis_note": (
                    f"scored at ${p.get('taxes_annual', 0):,.0f}/yr taxes "
                    f"(reassessed to ${price:,.0f})" if location else
                    "taxes frozen at the loaded basis - no location on the deal")}

    asking = inputs["price"]
    irr_hi = _irr_at(asking)
    if irr_hi is not None and irr_hi >= target_irr:
        return {"price": asking, "irr": irr_hi, "note": "already clears at asking"}

    lo = max(50_000, asking * 0.35)
    irr_lo = _irr_at(lo)
    if irr_lo is None or irr_lo < target_irr:
        return None

    hi = asking
    best = None
    for _ in range(max_iter):
        mid = (lo + hi) / 2
        irr_mid = _irr_at(mid)
        if irr_mid is None:
            hi = mid
            continue
        if irr_mid >= target_irr:
            best = {"price": mid, "irr": irr_mid}
            lo = mid
        else:
            hi = mid
        # Only stop on tolerance once a clearing price exists: a midpoint that
        # lands within tol *below* target would otherwise end the search with
        # best=None and report a reachable target as unreachable (sweep 2026-08-09).
        #
        # Test the accuracy of the answer we would RETURN, not of a probe we
        # discarded (sweep 2026-08-24). Testing irr_mid let a near-miss midpoint
        # end the search while `best` still held a far lower price from an
        # earlier iteration - hwy42 @16% reported $2,269,000 (16.54%) when
        # $2,298,000 clears it, a $30k understatement on a live board rung.
        if best is not None and abs(best["irr"] - target_irr) < tol:
            break

    if best:
        # Floor, never round: best["price"] is by construction the highest price
        # that clears the target, so rounding UP crosses the boundary and the
        # reported price no longer clears its own label (treme 16% reported
        # $664,000 @ 15.997%). Flooring keeps price and label consistent.
        best["price"] = int(best["price"] // 1000) * 1000
        best["irr"] = _irr_at(best["price"])
        # Hand back the basis the rung was certified on. A sale reassesses, so
        # every trial price above was scored on ITS OWN tax bill and its own
        # re-derived exit cap — but the solver used to return {price, irr}
        # alone, and consumers then re-scored the rung with the ASKING price's
        # taxes still attached. That understated P(IRR>=target) by 3.0-5.5pp
        # (eden's 22% rung, the live pursue basis, 33.9% -> 38.8%) and the
        # rung's own deterministic IRR by up to a full point. Returning the
        # basis kills the class rather than each instance (2026-09-07 sweep).
        best.update(_basis_at(best["price"]))
    return best


def tornado(inputs: dict) -> list[dict]:
    """Sensitivity table: IRR impact per stress factor, sorted by magnitude.

    Returns list of {factor, delta_irr, stressed_irr, direction}.
    """
    base = run(inputs)
    base_irr = base["levered_irr"]
    if base_irr is None:
        raise ValueError("base case IRR cannot be computed")

    stresses = [
        ("rate +50bps",   {"rate": inputs.get("rate", 0.0675) + 0.005}),
        ("rate -50bps",   {"rate": inputs.get("rate", 0.0675) - 0.005}),
        ("insurance +50%",{"insurance": inputs.get("insurance", 2000) * 1.5}),
        ("insurance -50%",{"insurance": inputs.get("insurance", 2000) * 0.5}),
        ("rents -5%",     {
            "unit_mix": [dict(g, rent=g["rent"] * 0.95) for g in inputs["unit_mix"]]
        }),
        ("exit_cap +100bps", {
            "exit_cap": (inputs.get("exit_cap") or base["going_in_cap"] + 0.005) + 0.01
        }),
        ("vacancy +3pts", {"vacancy": inputs.get("vacancy", 0.07) + 0.03}),
    ]

    # Audit F3 2026-08-05: freeze the exit cap at the BASE-case value for every
    # stress (except the cap stress itself). With exit_cap=None the stressed run
    # re-derived a richer multiple from the stressed going-in cap — the exit
    # cushioned the very shock being tested, understating downside 2.8-4.0 pts.
    frozen_cap = inputs.get("exit_cap") or round(base["going_in_cap"] + 0.005, 5)
    results = []
    for label, overrides in stresses:
        p = dict(inputs)
        p["exit_cap"] = frozen_cap
        p.update(overrides)
        try:
            r = run(p)
            stressed_irr = r["levered_irr"]
            if stressed_irr is None:
                continue
            delta = stressed_irr - base_irr
            results.append({
                "factor": label,
                "base_irr": base_irr,
                "stressed_irr": stressed_irr,
                "delta_irr": delta,
            })
        except Exception:
            pass

    results.sort(key=lambda x: abs(x["delta_irr"]), reverse=True)
    return results


# ---------------------------------------------------------------------------
# Monte Carlo internals (Gaussian copula, stdlib only)
# ---------------------------------------------------------------------------

# Hardcoded Cholesky L for the 6x6 correlation matrix:
#   order: [rent_growth, expense_growth, vacancy, insurance_mult, exit_cap_spread, turnover_rate]
#   rho_rg_eg=0.3, rho_rg_vac=-0.5, rho_ins_cap=0.3, rho_rg_tr=-0.3, all others=0
# Verified: L @ L.T reproduces the correlation matrix exactly.
# Row 5 (turnover_rate): L[5][0]=-0.3, L[5][5]=sqrt(1-0.09)=sqrt(0.91)=0.9539392014
_CHOL_L = [
    [1.0000000000, 0.0000000000, 0.0000000000, 0.0000000000, 0.0000000000, 0.0000000000],
    [0.3000000000, 0.9539392014, 0.0000000000, 0.0000000000, 0.0000000000, 0.0000000000],
    [-0.5000000000, 0.1572427255, 0.8516306273, 0.0000000000, 0.0000000000, 0.0000000000],
    [0.0000000000, 0.0000000000, 0.0000000000, 1.0000000000, 0.0000000000, 0.0000000000],
    [0.0000000000, 0.0000000000, 0.0000000000, 0.3000000000, 0.9539392014, 0.0000000000],
    [-0.3000000000, 0.0000000000, 0.0000000000, 0.0000000000, 0.0000000000, 0.9539392014],
]


def _norm_cdf(x: float) -> float:
    """Standard normal CDF via math.erf."""
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def _norm_ppf(u: float) -> float:
    """Inverse standard normal CDF (probit) — rational approximation.

    Based on Peter Acklam's algorithm (max error ~1.15e-9).
    Clamps u to (1e-12, 1-1e-12) to avoid infinities.
    """
    u = max(1e-12, min(1.0 - 1e-12, u))

    # Coefficients for the rational approximation
    a = [-3.969683028665376e+01,  2.209460984245205e+02,
         -2.759285104469687e+02,  1.383577518672690e+02,
         -3.066479806614716e+01,  2.506628277459239e+00]
    b = [-5.447609879822406e+01,  1.615858368580409e+02,
         -1.556989798598866e+02,  6.680131188771972e+01,
         -1.328068155288572e+01]
    c = [-7.784894002430293e-03, -3.223964580411365e-01,
         -2.400758277161838e+00, -2.549732539343734e+00,
          4.374664141464968e+00,  2.938163982698783e+00]
    d = [7.784695709041462e-03,  3.224671290700398e-01,
         2.445134137142996e+00,  3.754408661907416e+00]

    p_low = 0.02425
    p_high = 1.0 - p_low

    if u < p_low:
        q = math.sqrt(-2.0 * math.log(u))
        return (((((c[0] * q + c[1]) * q + c[2]) * q + c[3]) * q + c[4]) * q + c[5]) / \
               ((((d[0] * q + d[1]) * q + d[2]) * q + d[3]) * q + 1.0)
    elif u <= p_high:
        q = u - 0.5
        r = q * q
        return (((((a[0] * r + a[1]) * r + a[2]) * r + a[3]) * r + a[4]) * r + a[5]) * q / \
               (((((b[0] * r + b[1]) * r + b[2]) * r + b[3]) * r + b[4]) * r + 1.0)
    else:
        q = math.sqrt(-2.0 * math.log(1.0 - u))
        return -(((((c[0] * q + c[1]) * q + c[2]) * q + c[3]) * q + c[4]) * q + c[5]) / \
                ((((d[0] * q + d[1]) * q + d[2]) * q + d[3]) * q + 1.0)


def _tri_icdf(u: float, lo: float, mode: float, hi: float) -> float:
    """Inverse CDF of the triangular distribution (closed form)."""
    if hi == lo:
        return lo
    u = max(0.0, min(1.0, u))
    fc = (mode - lo) / (hi - lo)
    if u <= fc:
        return lo + math.sqrt(u * (hi - lo) * (mode - lo))
    else:
        return hi - math.sqrt((1.0 - u) * (hi - lo) * (hi - mode))


def _pw_icdf(u: float, p10: float, p50: float, p90: float) -> float:
    """Piecewise-uniform inverse CDF matching P10/P50/P90 exactly.

    Density is uniform on four segments: a left tail carrying 10% of mass
    (width continues the [p10,p50] segment's density), [p10,p50] with 40%,
    [p50,p90] with 40%, and a right tail with 10% (width continues the
    [p50,p90] density). Handles any skew a P10/P50/P90 triple can express;
    triangulars cannot (they max out near 3:1). Fable red-team 2026-08-03.
    """
    u = max(0.0, min(1.0, u))
    wl = 0.25 * (p50 - p10)   # left tail width at continued density
    wr = 0.25 * (p90 - p50)   # right tail width at continued density
    if u < 0.10:
        return (p10 - wl) + (u / 0.10) * wl
    if u < 0.50:
        return p10 + ((u - 0.10) / 0.40) * (p50 - p10)
    if u < 0.90:
        return p50 + ((u - 0.50) / 0.40) * (p90 - p50)
    return p90 + ((u - 0.90) / 0.10) * wr


def _tri_cdf(x: float, lo: float, mode: float, hi: float) -> float:
    if x <= lo:
        return 0.0
    if x >= hi:
        return 1.0
    if x <= mode:
        return (x - lo) ** 2 / ((hi - lo) * (mode - lo)) if mode > lo else 0.0
    return 1.0 - (hi - x) ** 2 / ((hi - lo) * (hi - mode)) if hi > mode else 1.0


def _tri_from_percentiles(p10: float, p50: float, p90: float) -> tuple:
    """Solve triangular (lo, mode, hi) whose 10th/50th/90th pctiles match.

    Iterative widen-and-rebalance; a few hundred cheap iterations, computed
    once per parameter per MC call. Non-monotone inputs pass through as-is.
    """
    if not (p10 < p50 < p90):
        return p10, p50, p90
    span = p90 - p10
    lo, hi = p10 - 0.6 * span, p90 + 0.6 * span
    mode = p50
    for _ in range(200):
        m_lo, m_hi = lo, hi
        for _ in range(40):
            mode = 0.5 * (m_lo + m_hi)
            if _tri_cdf(p50, lo, mode, hi) > 0.5:
                m_lo = mode
            else:
                m_hi = mode
        c10 = _tri_cdf(p10, lo, mode, hi)
        c90 = _tri_cdf(p90, lo, mode, hi)
        # c10 > 0.10 => too much mass below p10 (fat left tail) => raise lo.
        # c90 > 0.90 => too much mass below p90 (thin right tail) => raise hi.
        lo += (c10 - 0.10) * 0.8 * span
        hi += (c90 - 0.90) * 0.8 * span
        lo = min(lo, p10 - 0.01 * span)
        hi = max(hi, p90 + 0.01 * span)
        if abs(c10 - 0.10) < 0.001 and abs(c90 - 0.90) < 0.001:
            break
    return lo, mode, hi


def _box_muller_pair(rng) -> tuple:
    """Generate two independent standard normals using Box-Muller."""
    while True:
        u1 = rng.random()
        u2 = rng.random()
        if u1 > 0.0:
            break
    mag = math.sqrt(-2.0 * math.log(u1))
    angle = 2.0 * math.pi * u2
    return mag * math.cos(angle), mag * math.sin(angle)


def _correlated_uniforms(rng) -> list:
    """Draw 6 correlated uniforms via Gaussian copula.

    Uses Box-Muller for independent normals, applies Cholesky transform,
    then maps to [0,1] via normal CDF.
    Returns list of 6 uniform values in copula order:
    [rent_growth, expense_growth, vacancy, insurance_mult, exit_cap_spread, turnover_rate]
    """
    # Generate 6 independent standard normals (3 Box-Muller pairs = 6)
    z0, z1 = _box_muller_pair(rng)
    z2, z3 = _box_muller_pair(rng)
    z4, z5 = _box_muller_pair(rng)
    z_ind = [z0, z1, z2, z3, z4, z5]

    # Apply Cholesky: w = L @ z
    n = 6
    w = [sum(_CHOL_L[i][j] * z_ind[j] for j in range(n)) for i in range(n)]

    # Map to uniforms via normal CDF
    return [_norm_cdf(wi) for wi in w]


def _load_fitted_params() -> dict:
    """Load fitted_params.json. Returns dict or None (with stderr warning)."""
    params_path = Path(__file__).parent / "fitted_params.json"
    if not params_path.exists():
        print(
            "WARNING [monte_carlo]: tools/fitted_params.json not found. "
            "Run tools/fit_distributions.py to generate it. "
            "Falling back to hardcoded judgment distributions.",
            file=sys.stderr,
        )
        return None
    try:
        import json as _json
        with open(params_path) as f:
            return _json.load(f)
    except Exception as e:
        print(
            f"WARNING [monte_carlo]: could not parse fitted_params.json ({e}). "
            "Falling back to hardcoded judgment distributions.",
            file=sys.stderr,
        )
        return None


def _load_rentcast_mult(deal_name, base_rent: float) -> tuple:
    """Look for RentCast JSON for the deal and return (tri_lo, tri_mode, tri_hi, note).

    Returns None if no data found (caller skips rent-level uncertainty).
    deal_name: slug like 'eden-church-mhp'. Matched against filenames in _rentcast/.
    base_rent: weighted average rent from unit_mix ($/mo) — used to check for FROM_LISTING.
    """
    root = Path(__file__).resolve().parent.parent
    rentcast_dir = root / "deal-intake" / "_rentcast"
    if not rentcast_dir.exists() or not deal_name:
        return None

    import json as _json
    import glob as _glob

    # Find JSON files matching the deal slug. ALL >3-char tokens must appear:
    # any-token matching let 'church-street-8plex' consume the eden-church
    # cache of a different property (sweep 2026-08-09). Multiple files per
    # property (one per unit spec) are averaged, not first-glob-wins.
    # An explicit hint in deals.json ("rentcast": "<filename substring>") wins.
    # Cache files are named for the street ADDRESS, deal slugs for the
    # neighbourhood/property, so slug-token matching silently missed live deals:
    # 'treme-gov-nicholls' shares no >3-char token with
    # '1429-governor-nicholls-st...' ('treme' is absent), so its paid cache was
    # ignored and the MC ran with no rent-level uncertainty. (sweep 2026-08-24)
    hint = None
    try:
        _rec = _json.loads((root / "portfolio" / "deals.json").read_text()).get(deal_name, {})
        hint = _rec.get("rentcast")
    except Exception:
        hint = None

    matches = []
    if hint:
        for c in sorted(rentcast_dir.glob("*.json")):
            if hint.lower() in c.stem.lower():
                matches.append(c)

    if not matches:
        slug_parts = [p.lower() for p in deal_name.replace("-", " ").split() if len(p) > 3]
        if not slug_parts:
            return None
        for c in sorted(rentcast_dir.glob("*.json")):
            name_lower = c.stem.lower().replace("-", " ").replace("_", " ")
            if all(part in name_lower for part in slug_parts):
                matches.append(c)

    if not matches:
        return None

    mults = []
    for m in matches:
        try:
            with open(m) as f:
                data = _json.load(f)
            rent = data.get("rent")
            lo = data.get("rentRangeLow")
            hi = data.get("rentRangeHigh")
            if rent is None or lo is None or hi is None or rent == 0:
                continue
            mults.append((lo / rent, hi / rent))
        except Exception:
            continue
    if not mults:
        return None
    mult_lo = sum(m[0] for m in mults) / len(mults)
    mult_hi = sum(m[1] for m in mults) / len(mults)
    match_desc = (matches[0].name if len(matches) == 1
                  else f"{len(matches)} files avg ({matches[0].name}, ...)")

    # Check if deal has FROM_LISTING rent roll (narrow the range by half)
    deal_dir = root / "deal-intake" / deal_name
    has_from_listing = any(
        "FROM_LISTING" in f.name.upper()
        for f in deal_dir.glob("*") if f.is_file()
    ) if deal_dir.exists() else False

    if has_from_listing:
        # Halve the spread — actual rent roll narrows uncertainty
        mid_lo = 1.0 - (1.0 - mult_lo) / 2.0
        mid_hi = 1.0 + (mult_hi - 1.0) / 2.0
        note = f"RentCast {match_desc} + FROM_LISTING (range halved): mult [{mid_lo:.3f}, 1.000, {mid_hi:.3f}]"
        return (mid_lo, 1.0, mid_hi, note)
    else:
        note = f"RentCast {match_desc}: mult [{mult_lo:.3f}, 1.000, {mult_hi:.3f}]"
        return (mult_lo, 1.0, mult_hi, note)


def _percentile(sorted_list: list, pct: float) -> float:
    idx = (pct / 100.0) * (len(sorted_list) - 1)
    lo_i = int(idx)
    hi_i = min(lo_i + 1, len(sorted_list) - 1)
    frac = idx - lo_i
    return sorted_list[lo_i] + frac * (sorted_list[hi_i] - sorted_list[lo_i])


def _bootstrap_se_p10(irrs: list, k: int = 200, rng=None) -> float:
    """Bootstrap standard error of P10 from a sorted list of IRR samples."""
    if rng is None:
        rng = random.Random(999)
    n = len(irrs)
    p10s = []
    for _ in range(k):
        sample = sorted(rng.choices(irrs, k=n))
        p10s.append(_percentile(sample, 10))
    mean_p10 = sum(p10s) / k
    var = sum((x - mean_p10) ** 2 for x in p10s) / (k - 1)
    return math.sqrt(var)


def monte_carlo(inputs: dict, n: int = 0, seed: int = 42,
                deal_name: Optional[str] = None,
                year_built: Optional[int] = None,
                effective_age_override: Optional[int] = None) -> dict:
    """Monte Carlo with empirically-fitted distributions and Gaussian copula correlation.

    Draws correlated samples via a 6-variable Gaussian copula (Cholesky, Box-Muller,
    inverse triangular CDF). Fitted params loaded from tools/fitted_params.json;
    falls back to hardcoded judgment values with a stderr warning if missing.

    Auto-scales: runs batches of 1000 until bootstrap SE(P10) < 0.005 or n=20000.
    If n > 0, runs exactly n draws (no auto-scale) — used by tests and --mc flag.

    Optionally applies rent-level uncertainty from RentCast cache (deal_name required).

    FIX 1 — Integer-unit vacancy: turnovers drawn as Bernoulli per unit; days vacant
    drawn from triangular; correlates turnover_rate negatively with rent_growth (rho -0.3).
    Replaces smooth vacancy draws in MC only; deterministic run() is unchanged.

    FIX 2 — Vintage capex: if year_built or effective_age_override is provided, base
    capex is replaced by age-band lookup + lumpy roof/HVAC events (capped at $3k/unit).
    Falls back to flat capex_unit reserve if age is unknown.

    FIX 3 — Exit cap tail widened to +200bps upper bound (from +150bps).

    Returns dict with: p10, p50, p90, p_above_13, n_valid, n_draws, se_p10,
                       irr_samples, rent_level_note, params_provenance.
    """
    rng = random.Random(seed)
    boot_rng = random.Random(seed + 1)

    # --- Load fitted params ---
    fp = _load_fitted_params()

    def _param(key, fallback_p10, fallback_p50, fallback_p90):
        if fp and key in fp:
            p10, p50, p90 = fp[key]["p10"], fp[key]["p50"], fp[key]["p90"]
        else:
            p10, p50, p90 = fallback_p10, fallback_p50, fallback_p90
        # Draws use _pw_icdf (piecewise-uniform matching P10/P50/P90 exactly,
        # with real tails). Passing percentiles straight through - the old
        # triangular-with-P10/P90-endpoints truncated all tail probability.
        # (Fable red-team, 2026-08-03)
        return p10, p50, p90

    rg_p10, rg_p50, rg_p90 = _param("rent_growth", 0.01, 0.02, 0.03)
    eg_p10, eg_p50, eg_p90 = _param("expense_growth", 0.02, 0.025, 0.04)
    vac_p10, vac_p50, vac_p90 = _param("vacancy", 0.05, 0.07, 0.10)
    ins_p10, ins_p50, ins_p90 = _param("insurance_mult", 0.85, 1.0, 1.60)
    # FIX 3: exit cap upper tail widened to +200bps
    cap_p10, cap_p50, cap_p90 = _param("exit_cap_spread", -0.0025, 0.005, 0.020)

    params_provenance = "FITTED" if fp else "JUDGMENT-FALLBACK"

    # --- Base run ---
    base_run = run(inputs)
    base_going_in = base_run["going_in_cap"]
    base_insurance = inputs.get("insurance", 2000)

    # Recenter MC vacancy on the deal's own underwriting (sweep 2026-08-09).
    # Analytic mean of the raw turnover process: E[turnover]x E[days]/365 +
    # E[frictional]; unit count cancels in expectation.
    _proc_mean = ((0.30 + 0.45 + 0.60) / 3) * ((20.0 + 45.0 + 120.0) / 3) / 365.0 + 0.01
    _deal_loss = inputs.get("vacancy", 0.07) + inputs.get("bad_debt", 0.0)
    _vac_shift = _deal_loss - _proc_mean
    vacancy_note = (f"MC vacancy centered on underwriting {_deal_loss:.1%} "
                    f"(raw process mean {_proc_mean:.1%}, shift {_vac_shift:+.1%})")

    # Recenter rent growth and expense growth on the deal's own underwriting
    # (sweep 2026-08-24), exactly as vacancy is recentered above.
    #
    # The 2026-08-09 sweep recentered vacancy but left these two on the fitted
    # FRED marginals: rent growth P50 3.215% against a 2.0% underwrite, expense
    # growth P50 2.147% against 2.5%. Both errors point the same way, so MC P50
    # came out ABOVE the deterministic IRR on all seven deals - eden 12.18% vs
    # det 9.87%, treme 8.12% vs det 6.57% - and every quoted P(IRR>=13%) and
    # "beats index" figure inherited that optimism. The eden record of
    # 2026-08-09 named this as "the remaining known MC optimism"; this closes it.
    #
    # The fitted SPREAD is what the history is evidence for; the CENTER is the
    # deal's own assumption. House rule: the deterministic case is authoritative
    # and MC shows dispersion around it, never a rosier center of its own.
    def _recenter(p10, p50, p90, target, label):
        if target is None:
            return p10, p50, p90, ""
        d = target - p50
        note = (f"MC {label} centered on underwriting {target:.2%} "
                f"(fitted P50 {p50:.2%}, shift {d:+.2%}; fitted spread kept)")
        return p10 + d, p50 + d, p90 + d, note

    rg_p10, rg_p50, rg_p90, rg_note = _recenter(
        rg_p10, rg_p50, rg_p90, inputs.get("rent_growth"), "rent growth")
    eg_p10, eg_p50, eg_p90, eg_note = _recenter(
        eg_p10, eg_p50, eg_p90, inputs.get("expense_growth"), "expense growth")
    growth_note = "; ".join(n for n in (rg_note, eg_note) if n)

    # Weighted avg rent for RentCast lookup
    total_units = sum(g["units"] for g in inputs["unit_mix"])
    wt_rent = (sum(g["units"] * g["rent"] for g in inputs["unit_mix"]) / total_units
               if total_units > 0 else 0.0)
    n_units_int = int(round(total_units))

    # --- Rent-level multiplier from RentCast ---
    rentcast_info = _load_rentcast_mult(deal_name, wt_rent)
    if rentcast_info:
        rl_lo, rl_mode, rl_hi, rent_level_note = rentcast_info
    else:
        rl_lo = rl_mode = rl_hi = None
        # Say so, always. A silent empty note printed nothing, so a deal whose
        # cache simply failed to match looked identical to one with no rent-level
        # risk - and its P10 was optimistic for it. (sweep 2026-08-24)
        rent_level_note = (
            f"no RentCast cache matched '{deal_name}' - rent-level uncertainty NOT "
            "modelled, so the P10 is optimistic. Add a \"rentcast\" filename hint to "
            "this deal in portfolio/deals.json if a cache exists.")

    # --- FIX 2: Compute effective_age for vintage capex ---
    _current_year = 2026
    if effective_age_override is not None:
        _eff_age = effective_age_override
    elif year_built is not None:
        _eff_age = _current_year - year_built
    else:
        _eff_age = None  # unknown — fall back to flat capex reserve
        print("WARNING [monte_carlo]: no year_built/effective_age — vintage "
              "capex is INACTIVE and the P10 is optimistic (flat reserve only). "
              "Pass year_built for honest tails (audit 2026-08-05: this fix "
              "was silently inert in every production call).", file=sys.stderr)

    # Age-band base capex ($/unit/yr) for vintage capex
    def _age_band_base(age):
        if age <= 10:
            return 250.0
        elif age <= 25:
            return 450.0
        elif age <= 40:
            return 600.0
        else:
            return 750.0

    # Capex vintage params from fitted_params.json if present, else hardcoded
    _cap_vintage = None
    if fp and "capex_vintage" in fp:
        _cap_vintage = fp["capex_vintage"]

    def _draw_vintage_capex_annual(age):
        """Average annual capex/unit over the hold at a given effective age.

        Sweep 2026-08-09 rewrite: the prior version used lifetime-cumulative-
        shaped probabilities ((age-10)/30 etc.) as ANNUAL event probabilities
        and charged the resulting single lumpy year for EVERY year of the hold
        — at age 25+ the draw degenerated to the $3,000 cap almost surely.
        Now each hold year is simulated with annual hazards (HVAC ~15-20yr
        life: ramps from age 8, capped 12%/yr; roof ~20-30yr life: ramps from
        age 12, capped 9%/yr), age advances, and the hold-average is returned
        (the engine applies capex_unit uniformly per year).
        """
        hold = int(inputs.get("hold_years", 5) or 5)
        n_bldgs = max(1, round(n_units_int / 4))
        roof_lo, roof_mode, roof_hi = 8000.0, 11500.0, 15000.0
        if _cap_vintage and "roof_event" in _cap_vintage:
            rv = _cap_vintage["roof_event"]
            roof_lo = rv.get("cost_low", roof_lo)
            roof_mode = rv.get("cost_mode", roof_mode)
            roof_hi = rv.get("cost_high", roof_hi)
        hvac_lo, hvac_mode, hvac_hi = 4500.0, 5750.0, 7000.0
        if _cap_vintage and "hvac_event_per_unit" in _cap_vintage:
            hv = _cap_vintage["hvac_event_per_unit"]
            hvac_lo = hv.get("cost_low", hvac_lo)
            hvac_mode = hv.get("cost_mode", hvac_mode)
            hvac_hi = hv.get("cost_high", hvac_hi)
        cap_per_unit = 3000.0
        if _cap_vintage and "cap_per_unit" in _cap_vintage:
            cap_per_unit = _cap_vintage["cap_per_unit"]

        total = 0.0
        for k in range(hold):
            a = age + k
            hvac_prob = min(0.12, max(0.0, (a - 8) * 0.008))
            roof_prob = min(0.09, max(0.0, (a - 12) * 0.006))
            roof_cost = 0.0
            for _ in range(n_bldgs):
                if rng.random() < roof_prob:
                    roof_cost += _tri_icdf(rng.random(), roof_lo, roof_mode, roof_hi)
            hvac_cost = 0.0
            for _ in range(n_units_int):
                if rng.random() < hvac_prob:
                    hvac_cost += _tri_icdf(rng.random(), hvac_lo, hvac_mode, hvac_hi)
            yr_total = _age_band_base(a)
            if n_units_int > 0:
                yr_total += (roof_cost + hvac_cost) / n_units_int
            total += min(yr_total, cap_per_unit)
        return total / max(hold, 1)

    # --- Triangular params (p10/p50/p90 → lo/mode/hi) ---
    # We use p10/p50/p90 as lo/mode/hi for the triangular to keep the implementation
    # simple and the distributions empirically anchored. This is a deliberate
    # approximation: triangular P10 ≈ empirical P10 (exact only in special cases).
    def _draw_one():
        u6 = _correlated_uniforms(rng)
        u_rg, u_eg, u_vac, u_ins, u_cap, u_tr = u6

        rent_growth = _pw_icdf(u_rg, rg_p10, rg_p50, rg_p90)
        expense_growth = _pw_icdf(u_eg, eg_p10, eg_p50, eg_p90)
        ins_mult = _pw_icdf(u_ins, ins_p10, ins_p50, ins_p90)
        cap_delta = _pw_icdf(u_cap, cap_p10, cap_p50, cap_p90)

        # FIX 1: Integer-unit vacancy — physical process per simulated year
        # turnover_rate draws from triangular(0.30, 0.45, 0.60), correlated with
        # rent_growth via copula (rho = -0.3, so high turnover when rents fall)
        turnover_rate = _tri_icdf(u_tr, 0.30, 0.45, 0.60)
        turnovers = sum(1 for _ in range(n_units_int) if rng.random() < turnover_rate)
        days_vacant_total = sum(
            _tri_icdf(rng.random(), 20.0, 45.0, 120.0)
            for _ in range(turnovers)
        )
        physical_vac = days_vacant_total / max(n_units_int * 365, 1)
        frictional = rng.uniform(0.0, 0.02)  # collections / other losses
        # Sweep 2026-08-09: the raw process (mean ~8.6% loss) replaced the
        # deal's own underwriting outright — a 13.7%-actual-loss deal got a
        # generic 8.6% center and MC P50 nearly doubled the deterministic base.
        # Recenter: keep the process's dispersion/correlation, shift its mean
        # onto the deal's underwritten vacancy + bad debt (house rule 08-04).
        vacancy = max(0.0, physical_vac + frictional + _vac_shift)
        p = dict(inputs)
        # frictional replaces the static bad-debt line - zero it to avoid
        # double-counting collections losses (red-team 2026-08-03)
        p["bad_debt"] = 0.0
        p["rent_growth"] = rent_growth
        p["expense_growth"] = expense_growth
        p["vacancy"] = vacancy
        p["insurance"] = base_insurance * ins_mult
        p["exit_cap"] = max(0.001, base_going_in + cap_delta)

        # FIX 2: Vintage capex — override capex_unit for MC draws if age known
        if _eff_age is not None:
            p["capex_unit"] = _draw_vintage_capex_annual(_eff_age)

        # Rent-level uncertainty
        if rl_lo is not None:
            u_rl = rng.random()  # uncorrelated — rent level is a market appraisal error
            rl_mult = _tri_icdf(u_rl, rl_lo, rl_mode, rl_hi)
            p["unit_mix"] = [dict(g, rent=g["rent"] * rl_mult) for g in inputs["unit_mix"]]

        try:
            r = run(p)
            irr = r["levered_irr"]
            # levered_irr None here means the flows never turn positive —
            # equity wiped. Count it as a total-loss draw (-100%), not a
            # silently dropped one that truncates the left tail (sweep 2026-08-09).
            return -1.0 if irr is None else irr
        except Exception:
            return None

    # --- Auto-scale or fixed n ---
    irrs = []
    n_failed = 0
    target_se = 0.005  # 0.5 percentage points
    max_draws = 20000
    se_p10 = float("nan")

    if n > 0:
        # Fixed-n mode (tests, --mc flag)
        for _ in range(n):
            v = _draw_one()
            if v is not None:
                irrs.append(v)
            else:
                n_failed += 1
        irrs.sort()
        if len(irrs) >= 10:
            se_p10 = _bootstrap_se_p10(irrs, rng=boot_rng)
    else:
        # Auto-scale: batches of 1000. Bound by ATTEMPTS, not valid draws —
        # a deal whose every draw errors would otherwise loop forever
        # (sweep 2026-08-09).
        batch = 1000
        attempts = 0
        while len(irrs) < max_draws and attempts < 3 * max_draws:
            for _ in range(batch):
                attempts += 1
                v = _draw_one()
                if v is not None:
                    irrs.append(v)
                else:
                    n_failed += 1
            irrs.sort()
            if len(irrs) >= 10:
                se_p10 = _bootstrap_se_p10(irrs, rng=boot_rng)
                if se_p10 < target_se:
                    break

    if n_failed:
        print(f"WARNING [monte_carlo]: {n_failed} draw(s) raised errors and were "
              "excluded — quoted percentiles cover the surviving draws only.",
              file=sys.stderr)

    n_draws = len(irrs)

    if not irrs:
        return {
            "p10": None, "p50": None, "p90": None,
            "p_above_13": None, "n_valid": 0, "n_draws": 0,
            "n_failed": n_failed,
            "se_p10": float("nan"), "irr_samples": [],
            "rent_level_note": rent_level_note,
            "vacancy_note": vacancy_note,
            "growth_note": growth_note,
            "params_provenance": params_provenance,
        }

    n_valid = n_draws
    p_above = sum(1 for x in irrs if x >= 0.13) / n_valid

    return {
        "p10": _percentile(irrs, 10),
        "p50": _percentile(irrs, 50),
        "p90": _percentile(irrs, 90),
        "p_above_13": p_above,
        "n_valid": n_valid,
        "n_draws": n_draws,
        "se_p10": se_p10,
        "irr_samples": irrs,
        "n_failed": n_failed,
        "rent_level_note": rent_level_note,
        "vacancy_note": vacancy_note,
        "growth_note": growth_note,
        "params_provenance": params_provenance,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _load_deal(deal_name: str) -> dict:
    """Load inputs from a deal workbook's Inputs tab (data_only)."""
    import openpyxl
    root = Path(__file__).resolve().parent.parent
    wb_path = root / "deal-intake" / deal_name / f"{deal_name}_acq.xlsx"
    if not wb_path.exists():
        raise FileNotFoundError(f"No workbook at {wb_path}")

    wb = openpyxl.load_workbook(str(wb_path), data_only=True)
    ws = wb["Inputs"]

    def cell(addr):
        return ws[addr].value

    # Scalars
    price = cell("B2")
    closing_pct = cell("B3")
    loan_fee_pct = cell("B4")
    vacancy = cell("B9")
    bad_debt = cell("B10")
    concessions = cell("B11")
    other_income_unit_mo = cell("B12")
    rent_growth = cell("B13")
    other_income_growth = cell("B14")
    payroll = cell("B17")
    ga = cell("B18")
    marketing = cell("B19")
    rm = cell("B20")
    contract_services = cell("B21")
    utilities = cell("B22")
    other = cell("B23")
    insurance = cell("B24")
    taxes_annual = cell("B25")
    mgmt_pct = cell("B26")
    asset_mgmt_pct = cell("B27")
    capex_unit = cell("B28")
    expense_growth = cell("B29")
    ltv = cell("B32")
    min_dscr = cell("B33")
    rate = cell("B34")
    amort_years = cell("B35")
    io_years = cell("B36")
    hold_years = cell("B40")
    exit_cap = cell("B41")
    cost_of_sale = cell("B42")
    lp_pct = cell("B45")
    pref = cell("B47")
    residual_lp = cell("B48")
    refi_year = cell("F14")
    refi_valuation_cap = cell("F15")
    refi_ltv = cell("F16")
    refi_min_dscr = cell("F17")
    refi_rate = cell("F18")
    refi_amort_years = cell("F19")
    refi_cost_pct = cell("F20")
    reno_units = cell("F31")
    reno_cost_per_unit = cell("F32")
    reno_premium_month = cell("F33")
    reno_units_per_year = cell("F34")
    reno_downtime_months = cell("F35")
    reno_start_year = cell("F36")

    # Unit mix from rent roll (rows 3-10)
    unit_mix = []
    for row in range(3, 11):
        u = ws.cell(row=row, column=5).value  # E
        r = ws.cell(row=row, column=8).value  # H (rent)
        s = ws.cell(row=row, column=7).value  # G (sf)
        if u and u > 0:
            unit_mix.append({"units": u, "sf": s or 0, "rent": r or 0})

    def _v(x, default=0):
        return x if x is not None else default

    out = {
        "price": _v(price),
        "unit_mix": unit_mix,
        "closing_pct": _v(closing_pct, 0.02),
        "loan_fee_pct": _v(loan_fee_pct, 0.01),
        "vacancy": _v(vacancy, 0.07),
        "bad_debt": _v(bad_debt, 0.005),
        "concessions": _v(concessions, 0.0),
        "other_income_unit_mo": _v(other_income_unit_mo, 40),
        "rent_growth": _v(rent_growth, 0.02),
        "other_income_growth": _v(other_income_growth, 0.02),
        "payroll": _v(payroll, 0),
        "ga": _v(ga, 480),
        "marketing": _v(marketing, 415),
        "rm": _v(rm, 515),
        "contract_services": _v(contract_services, 0),
        "utilities": _v(utilities, 755),
        "other": _v(other, 105),
        "insurance": _v(insurance, 2000),
        "taxes_annual": _v(taxes_annual, 0),
        "mgmt_pct": _v(mgmt_pct, 0.08),
        "asset_mgmt_pct": _v(asset_mgmt_pct, 0.0),
        "capex_unit": _v(capex_unit, 250),
        "expense_growth": _v(expense_growth, 0.025),
        "ltv": _v(ltv, 0.75),
        "min_dscr": _v(min_dscr, 1.30),
        "rate": _v(rate, 0.0675),
        "amort_years": _v(amort_years, 25),
        "io_years": _v(io_years, 0),
        "hold_years": _v(hold_years, 5),
        "exit_cap": _v(exit_cap, None),
        "cost_of_sale": _v(cost_of_sale, 0.03),
        "lp_pct": _v(lp_pct, 0.90),
        "pref": _v(pref, 0.08),
        "residual_lp": _v(residual_lp, 0.70),
        "refi_year": _v(refi_year, 0),
        "refi_valuation_cap": _v(refi_valuation_cap, 0.07),
        "refi_ltv": _v(refi_ltv, 0.75),
        "refi_min_dscr": _v(refi_min_dscr, 1.30),
        "refi_rate": _v(refi_rate, 0.0675),
        "refi_amort_years": _v(refi_amort_years, 25),
        "refi_cost_pct": _v(refi_cost_pct, 0.02),
        "reno_units": _v(reno_units, 0),
        "reno_cost_per_unit": _v(reno_cost_per_unit, 0),
        "reno_premium_month": _v(reno_premium_month, 0),
        "reno_units_per_year": _v(reno_units_per_year, 0),
        "reno_downtime_months": _v(reno_downtime_months, 1),
        "reno_start_year": _v(reno_start_year, 1),
    }

    # Location drives per-price tax reassessment in solve_price(). Without it
    # taxes stay frozen at the asking-price basis, which biases every solved
    # price LOW (too much tax at the trial price -> too little NOI -> the
    # solver demands a deeper discount than reality). Baker Trails: $499K
    # frozen vs $528K reassessed, a $29K error in the wrong direction.
    # deals.json is the same source board.py uses, so the CLI and the board
    # can no longer disagree.
    deals_path = root / "portfolio" / "deals.json"
    if deals_path.exists():
        try:
            entry = json.loads(deals_path.read_text()).get(deal_name) or {}
            if entry.get("location"):
                out["location"] = entry["location"]
            if entry.get("commercial_share") is not None:
                out["commercial_share"] = entry["commercial_share"]
        except (ValueError, OSError):
            pass   # a broken deals.json must not stop the model from running

    return out


def main():
    ap = argparse.ArgumentParser(description="Python acquisition model")
    ap.add_argument("--deal", required=True, help="deal name under deal-intake/")
    ap.add_argument("--solve", type=float, metavar="IRR",
                    help="solve for price to hit this IRR (e.g. 0.16)")
    ap.add_argument("--tornado", action="store_true", help="sensitivity table")
    ap.add_argument("--mc", action="store_true", help="Full Monte Carlo report (auto-scale n)")
    ap.add_argument("--year-built", type=int, metavar="YYYY",
                    help="property vintage; activates vintage capex in every MC "
                         "(without it the MC uses the flat reserve and warns)")
    args = ap.parse_args()

    inputs = _load_deal(args.deal)
    r = run(inputs)

    # Quick 1000-draw MC for IRR interval (always shown in base report)
    quick_mc = monte_carlo(inputs, n=1000, seed=42, deal_name=args.deal,
                           year_built=args.year_built)
    irr_pct = r["levered_irr"] * 100 if r["levered_irr"] is not None else float("nan")
    if quick_mc["p10"] is not None:
        irr_interval = (f"  [{quick_mc['p10']*100:.1f}%, P90 {quick_mc['p90']*100:.1f}%]"
                        f"  [{quick_mc['params_provenance']}]")
    else:
        irr_interval = ""

    print(f"\n=== {args.deal.upper()} ===")
    print(f"  Price:          ${inputs['price']:>12,.0f}")
    print(f"  Loan:           ${r['loan_amount']:>12,.0f}")
    print(f"  Equity:         ${r['total_equity']:>12,.0f}")
    print(f"  Going-in cap:   {r['going_in_cap']*100:>10.2f}%")
    print(f"  Y1 NOI:         ${r['noi'][1]:>12,.0f}")
    print(f"  Y1 DSCR:        {r['dscr'][1]:>10.2f}x")
    print(f"  Y1 DSCR(+capex):{r['dscr_post_capex'][1]:>10.2f}x")
    print(f"  Levered IRR:    {irr_pct:>10.1f}%{irr_interval}")
    print(f"  Equity Multiple:{r['equity_multiple']:>10.2f}x")
    if r.get("waterfall_invalid"):
        print(f"  LP/GP IRR:      n/a — negative distributable cash in year(s) "
              f"{r.get('waterfall_neg_years')}; splits would overstate")
    else:
        print(f"  LP IRR:         {r['lp_irr']*100 if r['lp_irr'] else float('nan'):>10.2f}%")
        print(f"  GP IRR:         {r['gp_irr']*100 if r['gp_irr'] else float('nan'):>10.2f}%")
    if quick_mc.get("rent_level_note"):
        print(f"  Rent-level MC:  {quick_mc['rent_level_note']}")
    if quick_mc.get("vacancy_note"):
        print(f"  Vacancy MC:     {quick_mc['vacancy_note']}")
    if quick_mc.get("growth_note"):
        print(f"  Growth MC:      {quick_mc['growth_note']}")

    if args.solve:
        target = args.solve
        print(f"\n--- SOLVE: price to hit {target*100:.1f}% IRR ---")
        res = solve_price(inputs, target)
        if res is None:
            print("  Unreachable at any sensible price")
        elif res.get("note"):
            print(f"  Already clears at ${inputs['price']:,.0f}")
        else:
            disc = (1 - res["price"] / inputs["price"]) * 100
            # Score the rung on the basis it was CERTIFIED on, not the asking
            # price's tax bill and exit cap (2026-09-07 sweep: understated
            # P(IRR>=target) by 3.0-5.5pp across the live ladder).
            solve_inputs = dict(inputs, price=res["price"])
            for k in ("taxes_annual", "exit_cap"):
                if res.get(k) is not None:
                    solve_inputs[k] = res[k]
            solve_mc = monte_carlo(solve_inputs, n=1000, seed=42, deal_name=args.deal,
                                   year_built=args.year_built)
            p_target = sum(1 for x in solve_mc["irr_samples"] if x >= target) / max(len(solve_mc["irr_samples"]), 1)
            print(f"  Price: ${res['price']:,.0f}  ({disc:.0f}% below asking, IRR {res['irr']*100:.1f}%)  "
                  f"P(IRR>={target*100:.0f}%)={p_target*100:.0f}%")

    if args.tornado:
        print("\n--- TORNADO ---")
        print(f"  {'Factor':<25} {'Stressed IRR':>12} {'Delta':>8}")
        for row in tornado(inputs):
            print(f"  {row['factor']:<25} {row['stressed_irr']*100:>11.2f}%"
                  f" {row['delta_irr']*100:>+7.2f}%")

    if args.mc:
        print("\n--- MONTE CARLO (auto-scale, SE<0.5pts) ---")
        mc = monte_carlo(inputs, n=0, seed=42, deal_name=args.deal,
                         year_built=args.year_built)
        print(f"  P10 IRR:        {mc['p10']*100:.2f}%")
        print(f"  P50 IRR:        {mc['p50']*100:.2f}%")
        print(f"  P90 IRR:        {mc['p90']*100:.2f}%")
        print(f"  P(IRR>=13%):    {mc['p_above_13']*100:.1f}%")
        print(f"  Draws:          {mc['n_draws']:,}  SE(P10)={mc['se_p10']*100:.3f}pts")
        print(f"  Provenance:     {mc['params_provenance']}")
        if mc.get("rent_level_note"):
            print(f"  Rent-level:     {mc['rent_level_note']}")


if __name__ == "__main__":
    main()
