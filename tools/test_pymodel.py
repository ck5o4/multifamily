#!/usr/bin/env python3
"""Verification harness for pymodel.py.

Reads each deal workbook's Inputs tab with openpyxl data_only=True to get the
exact input vector, runs pymodel, and asserts against outputs read from Annual CF
and Waterfall tabs.

Tolerance: $1 on dollar figures, 0.01% (0.0001) on rates/IRRs.

Run:
    python3 tools/test_pymodel.py
    python3 tools/test_pymodel.py -v     # verbose diagnostics
"""

import argparse
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import pymodel

ROOT = Path(__file__).resolve().parent.parent
INTAKE = ROOT / "deal-intake"

TOL_DOLLAR = 1.0      # $1 tolerance on dollar figures
TOL_RATE = 0.0001     # 0.01% tolerance on IRR/cap rates


def _cell(ws, addr):
    v = ws[addr].value
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def _load_inputs(wb_path):
    """Read Inputs tab and return inputs dict for pymodel.run()."""
    wb = openpyxl.load_workbook(str(wb_path), data_only=True)
    ws = wb["Inputs"]

    def g(addr):
        v = _cell(ws, addr)
        return v

    def gf(addr, default=0.0):
        v = g(addr)
        return float(v) if v is not None else default

    def gi(addr, default=0):
        v = g(addr)
        return int(v) if v is not None else default

    unit_mix = []
    for row in range(3, 11):
        u = ws.cell(row=row, column=5).value   # E: units
        r = ws.cell(row=row, column=8).value   # H: rent/month
        s = ws.cell(row=row, column=7).value   # G: SF
        if u and float(u) > 0:
            unit_mix.append({
                "units": float(u),
                "sf": float(s) if s else 0,
                "rent": float(r) if r else 0,
            })

    return {
        "price":               gf("B2"),
        "unit_mix":            unit_mix,
        "closing_pct":         gf("B3", 0.02),
        "loan_fee_pct":        gf("B4", 0.01),
        "vacancy":             gf("B9", 0.07),
        "bad_debt":            gf("B10", 0.005),
        "concessions":         gf("B11", 0.0),
        "other_income_unit_mo":gf("B12", 0.0),
        "rent_growth":         gf("B13", 0.02),
        "other_income_growth": gf("B14", 0.02),
        "payroll":             gf("B17", 0.0),
        "ga":                  gf("B18", 480.0),
        "marketing":           gf("B19", 415.0),
        "rm":                  gf("B20", 515.0),
        "contract_services":   gf("B21", 0.0),
        "utilities":           gf("B22", 755.0),
        "other":               gf("B23", 105.0),
        "insurance":           gf("B24", 2000.0),
        "taxes_annual":        gf("B25", 0.0),
        "mgmt_pct":            gf("B26", 0.08),
        "asset_mgmt_pct":      gf("B27", 0.0),
        "capex_unit":          gf("B28", 250.0),
        "expense_growth":      gf("B29", 0.025),
        "ltv":                 gf("B32", 0.75),
        "min_dscr":            gf("B33", 1.30),
        "rate":                gf("B34", 0.0675),
        "amort_years":         gi("B35", 25),
        "io_years":            gi("B36", 0),
        "hold_years":          gi("B40", 5),
        "exit_cap":            gf("B41"),
        "cost_of_sale":        gf("B42", 0.03),
        "lp_pct":              gf("B45", 0.90),
        "pref":                gf("B47", 0.08),
        "residual_lp":         gf("B48", 0.70),
        "refi_year":           gi("F14", 0),
        "refi_valuation_cap":  gf("F15", 0.07),
        "refi_ltv":            gf("F16", 0.75),
        "refi_min_dscr":       gf("F17", 1.30),
        "refi_rate":           gf("F18", 0.0675),
        "refi_amort_years":    gi("F19", 25),
        "refi_cost_pct":       gf("F20", 0.02),
        "reno_units":          gi("F31", 0),
        "reno_cost_per_unit":  gf("F32", 0.0),
        "reno_premium_month":  gf("F33", 0.0),
        "reno_units_per_year": gi("F34", 0),
        "reno_downtime_months":gf("F35", 1.0),
        "reno_start_year":     gi("F36", 1),
    }


def _load_expected(wb_path):
    """Read Annual CF and Waterfall tabs for expected output values."""
    wb = openpyxl.load_workbook(str(wb_path), data_only=True)
    cf = wb["Annual CF"]
    wf = wb["Waterfall"]

    def cfcell(addr):
        return _cell(cf, addr)

    def wfcell(addr):
        return _cell(wf, addr)

    return {
        # Annual CF
        "noi_yr1":        cfcell("D25"),
        "egr_yr1":        cfcell("D11"),
        "opex_yr1":       cfcell("D24"),
        "dscr_yr1":       cfcell("D38"),
        "dscr_yr1_post":  cfcell("D40"),
        "total_ds_yr1":   cfcell("D34"),
        "end_bal_yr1":    cfcell("D35"),
        "lev_cf_yr0":     cfcell("C51"),
        "lev_cf_yr1":     cfcell("D51"),
        "lev_cf_yr5":     cfcell("H51"),   # always hold=5 in these deals
        "unlevered_irr":  cfcell("C54"),
        "levered_irr":    cfcell("C55"),
        "equity_multiple":cfcell("C56"),
        # Waterfall
        "lp_capital":     wfcell("C2"),
        "gp_capital":     wfcell("C3"),
        "lp_irr":         wfcell("C30"),
        "lp_em":          wfcell("C31"),
        "lp_profit":      wfcell("C32"),
        "gp_irr":         wfcell("C33"),
        # Inputs computed
        "loan_amount":    _cell(wb["Inputs"], "B52"),
        "total_equity":   _cell(wb["Inputs"], "B54"),
        "going_in_cap":   _cell(wb["Inputs"], "B57"),
    }


def _check(label, got, expected, tol, verbose, pass_list, fail_list):
    if expected is None:
        if verbose:
            print(f"    SKIP {label}: expected value is None (formula not recalculated)")
        return
    if got is None:
        msg = f"  FAIL {label}: got None, expected {expected}"
        fail_list.append(msg)
        print(msg)
        return

    diff = abs(got - expected)
    if diff <= tol:
        pass_list.append(label)
        if verbose:
            print(f"    OK   {label}: got={got:.6g} expected={expected:.6g} diff={diff:.6g}")
    else:
        msg = f"  FAIL {label}: got={got:.6g} expected={expected:.6g} diff={diff:.6g} (tol={tol})"
        fail_list.append(msg)
        print(msg)


def _run_deal(deal_name, verbose=False):
    wb_path = INTAKE / deal_name / f"{deal_name}_acq.xlsx"
    if not wb_path.exists():
        print(f"  ERROR: {wb_path} not found")
        return False

    print(f"\n{'='*60}")
    print(f"  DEAL: {deal_name}")
    print(f"{'='*60}")

    inputs = _load_inputs(wb_path)
    expected = _load_expected(wb_path)

    try:
        r = pymodel.run(inputs)
    except Exception as e:
        print(f"  FATAL: pymodel.run() raised: {e}")
        import traceback
        traceback.print_exc()
        return False

    passes = []
    fails = []

    def chk(label, got, exp, tol=TOL_DOLLAR):
        _check(label, got, exp, tol, verbose, passes, fails)

    # NOI and operating metrics
    chk("Y1 NOI",            r["noi"][1],            expected["noi_yr1"])
    chk("Y1 EGR",            r["egr"][1],             expected["egr_yr1"])
    chk("Y1 DSCR",           r["dscr"][1],            expected["dscr_yr1"],      tol=TOL_RATE)
    chk("Y1 DSCR post-capex",r["dscr_post_capex"][1], expected["dscr_yr1_post"], tol=TOL_RATE)
    chk("Y1 total DS",       r["total_ds"][1],        expected["total_ds_yr1"])
    chk("End bal Y1",        r["end_bal"][1],          expected["end_bal_yr1"])

    # Loan and equity
    chk("Loan amount",       r["loan_amount"],         expected["loan_amount"])
    chk("Total equity",      r["total_equity"],        expected["total_equity"])
    chk("Going-in cap",      r["going_in_cap"],        expected["going_in_cap"],  tol=TOL_RATE)
    chk("LP capital",        r["lp_capital"],          expected["lp_capital"])
    chk("GP capital",        r["gp_capital"],          expected["gp_capital"])

    # Levered cash flows
    chk("Lev CF year 0",     r["lev_cf"][0],          expected["lev_cf_yr0"])
    chk("Lev CF year 1",     r["lev_cf"][1],          expected["lev_cf_yr1"])
    hold = inputs["hold_years"]
    if hold == 5 and expected["lev_cf_yr5"] is not None:
        chk("Lev CF year 5", r["lev_cf"][5],          expected["lev_cf_yr5"])

    # IRRs and returns
    if expected["levered_irr"] is not None and expected["levered_irr"] != "":
        chk("Levered IRR",       r["levered_irr"],    float(expected["levered_irr"]), tol=TOL_RATE)
    if expected["unlevered_irr"] is not None and expected["unlevered_irr"] != "":
        chk("Unlevered IRR",     r["unlevered_irr"],  float(expected["unlevered_irr"]), tol=TOL_RATE)
    if expected["equity_multiple"] is not None and expected["equity_multiple"] != "":
        chk("Equity multiple",   r["equity_multiple"],float(expected["equity_multiple"]), tol=0.001)

    # Waterfall
    if expected["lp_irr"] is not None and expected["lp_irr"] != "":
        chk("LP IRR",            r["lp_irr"],         float(expected["lp_irr"]),  tol=TOL_RATE)
    if expected["lp_em"] is not None and expected["lp_em"] != "":
        chk("LP EM",             r["lp_em"],          float(expected["lp_em"]),   tol=0.001)
    if expected["lp_profit"] is not None and expected["lp_profit"] != "":
        chk("LP profit",         r["lp_profit"],      float(expected["lp_profit"]))
    if expected["gp_irr"] is not None and expected["gp_irr"] not in ("", None):
        try:
            chk("GP IRR",        r["gp_irr"],         float(expected["gp_irr"]),  tol=TOL_RATE)
        except (TypeError, ValueError):
            pass  # GP IRR can be blank if GP gets nothing

    # Summary
    n_pass = len(passes)
    n_fail = len(fails)
    print(f"\n  Result: {n_pass} passed, {n_fail} failed")
    if n_fail == 0:
        print(f"  ALL PASS for {deal_name}")
    else:
        print(f"  FAILURES:")
        for f in fails:
            print(f"    {f}")

    # Print key numbers regardless
    print(f"\n  Key numbers ({deal_name}):")
    print(f"    Y1 NOI:      ${r['noi'][1]:>12,.2f}    expected ${expected['noi_yr1']:>12,.2f}")
    print(f"    Loan:        ${r['loan_amount']:>12,.2f}    expected ${expected['loan_amount']:>12,.2f}")
    print(f"    Equity:      ${r['total_equity']:>12,.2f}    expected ${expected['total_equity']:>12,.2f}")
    levered_exp = expected['levered_irr']
    levered_got = r['levered_irr']
    print(f"    Levered IRR: {(levered_got or 0)*100:>12.4f}%   expected {(levered_exp or 0)*100:>12.4f}%")
    print(f"    Eq Multiple: {r['equity_multiple']:>12.4f}x    expected {(expected['equity_multiple'] or 0):>12.4f}x")

    return n_fail == 0


def main():
    ap = argparse.ArgumentParser(description="Verify pymodel against deal workbooks")
    ap.add_argument("-v", "--verbose", action="store_true", help="print all individual checks")
    args = ap.parse_args()

    deals = [
        "baker-trails",
        "eden-church-mhp",
        "covington-2nd",
    ]

    results = []
    for deal in deals:
        ok = _run_deal(deal, verbose=args.verbose)
        results.append((deal, ok))

    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    all_pass = True
    for deal, ok in results:
        status = "PASS" if ok else "FAIL"
        print(f"  {deal:<30} {status}")
        if not ok:
            all_pass = False

    if all_pass:
        print("\n  ALL DEALS VERIFIED - pymodel matches workbook outputs")
        sys.exit(0)
    else:
        print("\n  VERIFICATION FAILED - investigate discrepancies above")
        sys.exit(1)


if __name__ == "__main__":
    main()
