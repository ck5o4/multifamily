"""Guarded openpyxl writer. Refuses to touch any cell that is not a documented
blue input cell or inside a declared paste range, and refuses any cell that
currently holds a formula."""

import shutil
from pathlib import Path

import openpyxl

from cellmap import BLUE_RGB


class FormulaGuardError(Exception):
    pass


def parse_ref(ref):
    sheet, cell = ref.split("!")
    return sheet, cell


class ModelWriter:
    def __init__(self, path, spec):
        self.path = Path(path)
        self.spec = spec
        self.wb = openpyxl.load_workbook(self.path)
        self.writes = []
        self._ranges = spec.get("paste_ranges", [])

    def _in_paste_range(self, sheet, row, col):
        for s, r0, r1, c0, c1 in self._ranges:
            if s == sheet and r0 <= row <= r1 and c0 <= col <= c1:
                return True
        return False

    def _is_blue_input(self, cell):
        # Compare on RGB only. openpyxl versions disagree on the leading alpha
        # byte (FF0000FF vs 000000FF) and the workbooks were built across two.
        f = cell.font
        if not f or not f.color:
            return False
        rgb = getattr(f.color, "rgb", None)
        return isinstance(rgb, str) and rgb.upper()[-6:] == BLUE_RGB[-6:]

    def _assert_writable(self, sheet, cell):
        cur = cell.value
        if isinstance(cur, str) and cur.startswith("="):
            raise FormulaGuardError(
                f"{sheet}!{cell.coordinate} holds a formula ({cur[:40]}...) - refusing to write"
            )
        if self._is_blue_input(cell) or self._in_paste_range(sheet, cell.row, cell.column):
            return
        raise FormulaGuardError(
            f"{sheet}!{cell.coordinate} is not a documented input cell - refusing to write"
        )

    def set(self, ref, value):
        sheet, coord = parse_ref(ref)
        ws = self.wb[sheet]
        cell = ws[coord]
        self._assert_writable(sheet, cell)
        old = cell.value
        cell.value = value
        self.writes.append((ref, old, value))

    def set_rc(self, sheet, row, col, value):
        ws = self.wb[sheet]
        cell = ws.cell(row=row, column=col)
        self._assert_writable(sheet, cell)
        old = cell.value
        cell.value = value
        self.writes.append((f"{sheet}!{cell.coordinate}", old, value))

    def clear_block(self, sheet, r0, r1, c0, c1):
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                self.set_rc(sheet, r, c, None)

    def save(self, dest=None):
        out = Path(dest) if dest else self.path
        self.wb.save(out)
        return out


def clone_model(master_path, dest_path):
    dest_path = Path(dest_path)
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(master_path, dest_path)
    return dest_path
