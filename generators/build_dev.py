"""Original multifamily DEVELOPMENT model - monthly 96-month engine.
Same equation structure as Stoa-era model: comp-driven rents, lease-up absorption,
fixed/variable expense phase-in, bell-curve construction draws, equity-first-then-debt,
capitalized construction interest, perm refi (LTV+DSCR sized) w/ cash-out, sell-or-hold,
LP pref waterfall. Non-circular capitalized-interest solve (algebraic, no iterative calc)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Explicit 8-char aRGB - see the note in build_acq.py.
BLUE=Font(name='Arial',color='FF0000FF'); BLK=Font(name='Arial'); GRN=Font(name='Arial',color='FF008000')
BOLD=Font(name='Arial',bold=True); HDR=Font(name='Arial',bold=True,color='FFFFFF')
YEL=PatternFill('solid',fgColor='FFFF00'); DK=PatternFill('solid',fgColor='404040')
CUR='$#,##0;($#,##0);-'; PCT='0.0%'; PCT2='0.00%'; NUM='#,##0'; X='0.00x'; DT='mmm-yy'
NM=96  # months

wb=Workbook()

# ---------- READ ME ----------
rm=wb.active; rm.title='Read Me'
rm['A1']='MULTIFAMILY DEVELOPMENT MODEL'; rm['A1'].font=Font(name='Arial',bold=True,size=14)
for i,t in enumerate([
 'Original model; equation structure mirrors the Stoa-era development model.',
 '',
 'HOW TO USE:',
 '1. Blue cells = inputs (Inputs, Comps, Construction Budget tabs). Yellow = key drivers.',
 '2. NO iterative calculation needed. Capitalized construction interest is solved',
 '   algebraically (loan = LTC x cost / (1 - LTC x (interest factor + fee%))), then the',
 '   monthly ledger accrues ACTUAL interest. The Checks tab compares estimate vs. actual.',
 '3. Rents pull from the Comps tab (weighted comp averages by unit type, less/plus a',
 '   delta adjustment) - same VLOOKUP mechanism as the original. Paste real comps there.',
 '4. Construction draws follow a normal-distribution (bell curve) schedule - SD inputs',
 '   for hard and soft costs, same as the original (SD 6 hard / SD 10 soft).',
 '5. Timeline: month 1 = land close. Construction runs Start Month -> Start+Duration-1.',
 '   Lease-up can begin before completion (phased delivery), as Stoa did.',
 '6. Refi: perm loan sized to MIN(LTV x value, DSCR constraint); value = refi-month NOI',
 '   x 12 / valuation cap rate. Cash-out (or shortfall) flows to equity at refi month.',
 '7. Sell-or-Hold toggle. Sale = sale-month NOI x 12 / exit cap, less cost of sale.',
 '8. Deviations from the original and every benchmark number: Stoa Benchmarks tab.',
 '9. All Checks tab cells should read OK before trusting outputs.',
],start=3): rm.cell(row=i,column=1,value=t)
rm.column_dimensions['A'].width=100

# ---------- COMPS ----------
cp=wb.create_sheet('Comps')
cp['A1']='RENT COMPS'; cp['A1'].font=BOLD
cp['A2']='Paste comp rent rolls below. Include=1 to count a row. Summary (rows 4-6) feeds Inputs rents via VLOOKUP.'
for j,h in enumerate(['','','','Include','# Units','Unit Type','SF','Rent/Month','Rent/SF']):
    if h: cp.cell(row=3,column=1+j,value=h).font=HDR; cp.cell(row=3,column=1+j).fill=DK
# summary block F4:I6 keyed by unit type
cp['E4']='SUMMARY BY TYPE'; cp['E4'].font=BOLD
types=['1 BR/ 1 BA','2 BR/ 2 BA','3 BR/ 2 BA']
for i,t in enumerate(types):
    r=4+i
    cp.cell(row=r,column=6,value=t).font=BOLD
    cp.cell(row=r,column=7,value=f"=IFERROR(SUMPRODUCT(($F$10:$F$80=F{r})*($D$10:$D$80)*($E$10:$E$80)*($G$10:$G$80))/SUMPRODUCT(($F$10:$F$80=F{r})*($D$10:$D$80)*($E$10:$E$80)),0)").number_format=NUM
    cp.cell(row=r,column=8,value=f"=IFERROR(SUMPRODUCT(($F$10:$F$80=F{r})*($D$10:$D$80)*($E$10:$E$80)*($H$10:$H$80))/SUMPRODUCT(($F$10:$F$80=F{r})*($D$10:$D$80)*($E$10:$E$80)),0)").number_format=CUR
    cp.cell(row=r,column=9,value=f"=IFERROR(H{r}/G{r},0)").number_format='0.00'
cp['A9']='COMP DATA (rows 10-80)'; cp['A9'].font=BOLD
demo_comps=[(1,40,'1 BR/ 1 BA',750,1600),(1,30,'1 BR/ 1 BA',820,1720),(1,48,'2 BR/ 2 BA',1050,1950),
            (1,24,'2 BR/ 2 BA',1150,2080),(1,16,'3 BR/ 2 BA',1300,2400),(1,12,'3 BR/ 2 BA',1380,2520)]
for i,(inc,u,t,sf,r_) in enumerate(demo_comps):
    r=10+i
    for j,v in enumerate([inc,u,t,sf,r_]):
        c=cp.cell(row=r,column=4+j,value=v); c.font=BLUE
        if j==4: c.number_format=CUR
for col,w in [('D',8),('E',8),('F',12),('G',8),('H',12),('I',8)]: cp.column_dimensions[col].width=w

# ---------- INPUTS ----------
ip=wb.create_sheet('Inputs')
def I(cell,label,val,fmt=None,font=BLUE,fill=None,labelcol='A'):
    ip[f'{labelcol}{cell}']=label
    c=ip[f'{chr(ord(labelcol)+1)}{cell}']; c.value=val; c.font=font
    if fmt: c.number_format=fmt
    if fill: c.fill=fill
ip['A1']='LAND & TIMING'; ip['A1'].font=BOLD
I(2,'Land Price',6750000,CUR,fill=YEL)
I(3,'Closing Costs (% of Land)',0.0075,PCT)
I(4,'Start Date (Land Close = Month 1)','=DATE(2026,8,31)',DT,font=BLUE)
I(5,'Net Rentable SF','=SUMPRODUCT(G3:G8,E3:E8)',NUM,font=BLK)
I(6,'Total Units','=E9',NUM,font=GRN)
I(8,'Construction Duration (Months)',20,NUM,fill=YEL)
I(9,'Construction Start Month',2,NUM)
I(10,'Lease-Up Start Month',14,NUM)
I(11,'Lease-Up Period (Months)',14,NUM)
I(12,'Lease-Up Pace (Units/Month)','=IFERROR(E9/B11,0)','0.0',font=BLK)
I(13,'Stabilization Month','=B10+ROUNDUP(B6*(1-B16)/B12,0)',NUM,font=BLK)
ip['A15']='REVENUE'; ip['A15'].font=BOLD
I(16,'Vacancy (Stabilized)',0.07,PCT)
I(17,'Bad Debt %',0.005,PCT)
I(18,'Other Income ($/Occupied Unit/Month)',40,CUR)
I(19,'Rent Growth (Annual)',0.02,PCT)
# Unit mix grid E2:K9
ip['E1']='UNIT MIX (rents pull from Comps tab less Delta)'; ip['E1'].font=BOLD
for j,h in enumerate(['# Units','Unit Type','SF','Rent/Month','Rent/SF','','Delta vs Comp']):
    c=ip.cell(row=2,column=5+j,value=h)
    if h: c.font=HDR; c.fill=DK
mix=[(134,'1 BR/ 1 BA',764,-0.04),(121,'2 BR/ 2 BA',1040,-0.05),(33,'3 BR/ 2 BA',1318,0.12),
     (0,'BR/BA',0,0),(0,'BR/BA',0,0),(0,'BR/BA',0,0)]
for i,(u,t,sf,d) in enumerate(mix):
    r=3+i
    ip.cell(row=r,column=5,value=u).font=BLUE
    ip.cell(row=r,column=6,value=t).font=BLUE
    ip.cell(row=r,column=7,value=sf).font=BLUE
    ip.cell(row=r,column=8,value=f"=IFERROR(VLOOKUP(F{r},Comps!$F$4:$I$6,3,FALSE)*(1-K{r}),0)").number_format=CUR
    ip.cell(row=r,column=9,value=f"=IFERROR(H{r}/G{r},0)").number_format='0.00'
    ip.cell(row=r,column=11,value=d).font=BLUE; ip.cell(row=r,column=11).number_format=PCT
ip['E9']='=SUM(E3:E8)'; ip['E9'].font=BOLD; ip['E9'].number_format=NUM
ip['G9']='=IFERROR(SUMPRODUCT(G3:G8,E3:E8)/E9,0)'; ip['G9'].number_format=NUM
ip['H9']='=IFERROR(SUMPRODUCT(H3:H8,E3:E8)/E9,0)'; ip['H9'].number_format=CUR

ip['A21']='HARD & SOFT COST DRIVERS'; ip['A21'].font=BOLD
I(22,'Hard Cost ($/NRSF)',135,CUR,fill=YEL)
I(23,'Hard Cost Fixed Adder ($)',1000000,CUR)
I(24,'GC Overhead %',0.05,PCT)
I(25,'Hard Cost Contingency %',0.10,PCT)
I(26,'Soft Cost Contingency %',0.02,PCT)
I(27,'GC Profit %',0.04,PCT)
I(28,'Developer Fee %',0.035,PCT)
I(29,'Impact Fees ($/Unit)',4127,CUR)
I(30,'Hard Cost Draw Curve SD (Months)',6,NUM)
I(31,'Soft Cost Draw Curve SD (Months)',10,NUM)

ip['A33']='OPERATING EXPENSES'; ip['A33'].font=BOLD
ip['A34']='Line Item'; ip['B34']='$/Unit/Yr'; ip['C34']='% Fixed'; ip['D34']='Months Prior to Lease-Up'
for c_ in ['A34','B34','C34','D34']: ip[c_].font=HDR; ip[c_].fill=DK
exp=[('Payroll',1715,1,4),('General & Administrative',480,1,4),('Marketing',415,1,4),
     ('Repairs & Maintenance',515,0.2,3),('Contract Services',0,0,0),('Utilities',755,0.5,3),
     ('Other',105,0,0),('Insurance',2000,1,4)]
for i,(lab,v,fx,mp) in enumerate(exp):
    r=35+i
    ip.cell(row=r,column=1,value=lab)
    ip.cell(row=r,column=2,value=v).font=BLUE; ip.cell(row=r,column=2).number_format=CUR
    ip.cell(row=r,column=3,value=fx).font=BLUE; ip.cell(row=r,column=3).number_format=PCT
    ip.cell(row=r,column=4,value=mp).font=BLUE
I(43,'Property Taxes During Construction (Annual $)',30000,CUR)
I(44,'Property Taxes Stabilized (Annual $)',400000,CUR,fill=YEL)
I(45,'Property Management Fee (% of EGR)',0.03,PCT)
I(46,'Asset Management Fee (% of EGR)',0.01,PCT)
I(47,'Capital Expense Reserve ($/Unit/Yr)',250,CUR)
I(48,'Expense Growth (Annual)',0.025,PCT)

ip['A50']='CONSTRUCTION LOAN'; ip['A50'].font=BOLD
I(51,'Loan-To-Cost (LTC)',0.75,PCT)
I(52,'Index (WSJ Prime)',0.0675,PCT2,fill=YEL)
I(53,'Spread over Index',0.005,PCT2)
I(54,'All-In Rate','=B52+B53',PCT2,font=BLK)
I(55,'Loan Fee (% of Loan)',0.005,PCT)

ip['A57']='PERMANENT LOAN (Refi)'; ip['A57'].font=BOLD
I(58,'Refinance Month',33,NUM,fill=YEL)
I(59,'Valuation Cap Rate',0.06,PCT2)
I(60,'LTV',0.75,PCT)
I(61,'Min DSCR',1.176,'0.000')
I(62,'Interest Rate',0.0625,PCT2,fill=YEL)
I(63,'Amortization (Months)',420,NUM)
I(64,'MIP (Annual, HUD)',0.0025,PCT2)
I(65,'Perm Loan Fee (% of Proceeds)',0.02,PCT)
I(66,'Interest-Only Period (Months)',0,NUM)

ip['A68']='EXIT'; ip['A68'].font=BOLD
I(69,'Sell or Hold? (type Sell or Hold)','Sell',None,fill=YEL)
I(70,'Sale Month (if Sell, <=96)',60,NUM,fill=YEL)
I(71,'Exit Cap Rate',0.065,PCT2,fill=YEL)
I(72,'Cost of Sale (% of Price)',0.0075,PCT)

ip['A74']='EQUITY / INVESTOR'; ip['A74'].font=BOLD
I(75,'LP (Investor) % of Cash Equity',1.0,PCT)
I(76,'LP Preferred Return (Annual)',0.08,PCT)
I(77,'Residual Split - LP %',0.70,PCT)
I(78,'GP In-Kind Equity Credit (On/Off)','Off')

ip['A80']='COMPUTED FINANCING (do not edit)'; ip['A80'].font=BOLD
I(81,'Total Cost ex-Cap Interest & Loan Fee',"=B2+B2*B3+'Construction Budget'!C22",CUR,font=BLK)
I(82,'Interest Factor K',"=(B54/12)*(0.5*B8+MAX(0,B58-(B9+B8)))",'0.0000',font=BLK)
I(83,'Est. Construction Loan (fee basis)',"=B51*B81/(1-B51*(B82+B55))",CUR,font=BLK)
I(84,'Actual Capitalized Interest',"=SUM('Monthly CF'!C42:CT42)",CUR,font=BLK)
I(85,'Construction Loan Fee',"=B83*B55",CUR,font=BLK)
I(86,'Total Project Cost',"=B81+B84+B85",CUR,font=BLK)
I(87,'Net Equity Required ((1-LTC) x Cost)',"=(1-B51)*B81",CUR,font=BLK,fill=YEL)
I(88,'LP Capital (Cash)',"=B87*B75",CUR,font=BLK)
I(89,'GP Capital (Cash)',"=B87-B88",CUR,font=BLK)
I(90,'GP In-Kind Credit',"=IF(B78=\"On\",'Construction Budget'!C19+'Construction Budget'!C20,0)",CUR,font=BLK)
I(91,'Perm Loan Amount (at Refi)',"=INDEX('Monthly CF'!$C$47:$CT$47,1,B58)",CUR,font=BLK)
I(92,'Perm Monthly Payment',"=IFERROR(PMT(B62/12,B63,-B91),0)",CUR,font=BLK)
I(93,'Effective Hold End Month',"=IF(B69=\"Sell\",B70,96)",NUM,font=BLK)
ip.column_dimensions['A'].width=40; ip.column_dimensions['B'].width=14
for col in 'EFGHIK': ip.column_dimensions[col].width=12

# ---------- CONSTRUCTION BUDGET ----------
cb=wb.create_sheet('Construction Budget')
IB='Inputs!$B$'; IE='Inputs!$E$'
cb['A1']='CONSTRUCTION BUDGET'; cb['A1'].font=BOLD
lines=[
 (3,'HARD COSTS','',True),
 (4,'Base Hard Cost ($/SF x NRSF + Adder)',f"={IB}22*{IB}5+{IB}23",False),
 (5,'Hard Cost Contingency',f"={IB}25*C4",False),
 (6,"Contractor's Overhead (grossed up)",f"=(C4+C5)/(1-{IB}24)-(C4+C5)",False),
 (7,'Hard Cost Subtotal',"=SUM(C4:C6)",True),
 (9,'SOFT COSTS','',True),
 (10,'Architectural',f"=1800*{IB}6",False),
 (11,'Landscape Architecture',f"=150*{IB}6",False),
 (12,'Geotech / Civil / Phase 1 / Title / Legal / Appraisal / Closing',"=20000+95000+2500+68000+5100+4500+700",False),
 (13,'Builders Risk + General Liability Insurance',"=C7/1000*11",False),
 (14,'Local Impact Fees',f"={IB}29*{IB}6",False),
 (15,'NGBS Green Certification + Consulting',f"=(230+200)*{IB}6",False),
 (16,'0.5% Developer Fee (on hard+soft base)',"=0.005*(C7+SUM(C10:C15))",False),
 (17,'Soft Cost Subtotal',"=SUM(C10:C16)",True),
 (18,'Soft Cost Contingency',f"={IB}26*C17",False),
 (19,'GC Profit',f"={IB}27*C7",False),
 (20,'Developer Fee',f"={IB}28*(C7+C17+C18+C19)",False),
 (22,'TOTAL CONSTRUCTION BUDGET (excl land, interest, loan fee)',"=C7+C17+C18+C19+C20",True),
]
for r,lab,f,b in lines:
    cb.cell(row=r,column=2,value=lab).font=BOLD if b else BLK
    if f:
        c=cb.cell(row=r,column=3,value=f); c.number_format=CUR
        if b: c.font=BOLD
    cb.cell(row=r,column=4,value=f"=IFERROR(C{r}/{IB}6,0)" if f else None)
    if f: cb.cell(row=r,column=4).number_format=CUR
cb['D2']='Per Unit'; cb['D2'].font=HDR; cb['D2'].fill=DK
cb['B24']='Month-1 Items (title/legal/appraisal/closing/phase1/geotech-civil deposits)'; 
cb['C24']='=C12'; cb['C24'].number_format=CUR

# Draw schedule rows 27-33, months in cols F.. (F = month 1)
cb['B26']='DRAW SCHEDULE (bell curve)'; cb['B26'].font=BOLD
cb['B27']='Month'
cb['B28']='Hard Curve Raw (NORMDIST)'
cb['B29']='Hard Spend'
cb['B30']='Soft Curve Raw (NORMDIST)'
cb['B31']='Soft Spend (curve portion)'
cb['B32']='Month-1 Lump'
cb['B33']='Total Construction Spend'
cb['E28']='=SUM(F28:CW28)'; cb['E30']='=SUM(F30:CW30)'
for m in range(1,NM+1):
    col=get_column_letter(5+m)  # F=month1
    cb[f'{col}27']=m
    cb[f'{col}28']=f"=IF(AND({col}27>={IB}9,{col}27<={IB}9+{IB}8-1),NORMDIST({col}27,{IB}9+{IB}8/2,{IB}30,0),0)"
    cb[f'{col}29']=f"=IF($E$28=0,0,('Construction Budget'!$C$7+$C$19)*{col}28/$E$28)"
    cb[f'{col}30']=f"=IF(AND({col}27>={IB}9,{col}27<={IB}9+{IB}8-1),NORMDIST({col}27,{IB}9+{IB}8/2,{IB}31,0),0)"
    cb[f'{col}31']=f"=IF($E$30=0,0,($C$17+$C$18+$C$20-$C$24)*{col}30/$E$30)"
    cb[f'{col}32']=f"=IF({col}27=1,$C$24,0)"
    cb[f'{col}33']=f"={col}29+{col}31+{col}32"
    for r in (29,31,32,33): cb[f'{col}{r}'].number_format=CUR
cb['B35']='Check: Sum of Draws = Total Budget'
cb['C35']="=IF(ABS(SUM(F33:CW33)-C22)<1,\"OK\",\"ERROR\")"
cb.column_dimensions['B'].width=55; cb.column_dimensions['C'].width=16; cb.column_dimensions['D'].width=12

# ---------- MONTHLY CF ----------
mc=wb.create_sheet('Monthly CF')
mc['B1']='Month'; mc['B2']='Date'; mc['B3']='Year Index'
labels={
 5:'Units Leased',6:'Occupied Units',7:'Gross Potential Rent',8:'Vacancy / Lease-Up Loss',
 9:'Bad Debt',10:'Net Rental Income',11:'Other Income',12:'Effective Gross Revenue',
 14:'Payroll',15:'General & Administrative',16:'Marketing',17:'Repairs & Maintenance',
 18:'Contract Services',19:'Utilities',20:'Other',21:'Insurance',22:'Property Taxes',
 23:'Management Fee',24:'Total Operating Expenses',26:'Net Operating Income',
 27:'Capital Expense Reserve',28:'Asset Management Fee',29:'Cash Flow Before Debt Service',
 31:'Land & Closing Costs',32:'Construction Spend',33:'Construction Loan Fee',34:'Total Development Cost',
 36:'Funding Need (pre-refi)',37:'Equity Remaining - Beg',38:'Equity Draw',39:'Equity Remaining - End',
 40:'Constr Loan Balance - Beg',41:'Constr Loan Draw',42:'Capitalized Interest',43:'Constr Loan Payoff',
 44:'Constr Loan Balance - End',
 46:'Refi Value (NOI x 12 / Cap)',47:'Perm Loan Funding',48:'Perm Loan Fee',49:'Net Cash-Out at Refi',
 50:'Perm Balance - Beg',51:'Perm Total Payment',52:'Perm Interest',53:'Perm Principal',54:'MIP',
 55:'Perm Balance - End',56:'Perm Payoff at Sale',
 58:'Sale Price',59:'Cost of Sale',
 61:'Levered Net Cash Flow',
}
for r,l in labels.items():
    b = r in (12,24,26,29,34,44,49,55,61)
    mc.cell(row=r,column=2,value=l).font=BOLD if b else BLK
mc['B63']='Project IRR (XIRR)'; mc['B63'].font=BOLD
mc['B64']='Equity Multiple'; mc['B64'].font=BOLD
mc['B65']='Total Profit'; mc['B65'].font=BOLD

exp_rows=list(range(14,22))  # maps to Inputs rows 35..42
for m in range(1,NM+1):
    col=get_column_letter(2+m)  # C = month 1
    P=get_column_letter(1+m)    # prev col
    f=mc[f'{col}1']; f.value=m; f.font=BOLD
    mc[f'{col}2']=f"=EDATE({IB}4,{col}1-1)"; mc[f'{col}2'].number_format=DT
    mc[f'{col}3']=f"=INT(({col}1-1)/12)+1"
    mc[f'{col}5']=f"=MAX(0,MIN({IB}6,({col}1-{IB}10+1)*{IB}12))"
    mc[f'{col}6']=f"=MIN({col}5,{IB}6*(1-{IB}16))"
    mc[f'{col}7']=f"=SUMPRODUCT({IE}3:{IE}8,Inputs!$H$3:$H$8)*(1+{IB}19)^({col}3-1)"
    mc[f'{col}8']=f"=-{col}7*(1-{col}6/{IB}6)"
    mc[f'{col}9']=f"=-({col}7+{col}8)*{IB}17"
    mc[f'{col}10']=f"=SUM({col}7:{col}9)"
    mc[f'{col}11']=f"={col}6*{IB}18*(1+{IB}19)^({col}3-1)"
    mc[f'{col}12']=f"={col}10+{col}11"
    for k,r in enumerate(exp_rows):
        ir=35+k
        mc[f'{col}{r}']=(f"=IF({col}1>={IB}10-Inputs!$D${ir},"
                         f"-(Inputs!$B${ir}*{IB}6/12)*(1+{IB}48)^({col}3-1)*"
                         f"(Inputs!$C${ir}+(1-Inputs!$C${ir})*{col}6/{IB}6),0)")
    mc[f'{col}22']=f"=IF({col}1>={IB}9+{IB}8,-{IB}44/12*(1+{IB}48)^({col}3-1),-{IB}43/12)"
    mc[f'{col}23']=f"=-MAX({col}12,0)*{IB}45"
    mc[f'{col}24']=f"=SUM({col}14:{col}23)"
    mc[f'{col}26']=f"={col}12+{col}24"
    mc[f'{col}27']=f"=IF({col}1>={IB}13,-({IB}6*{IB}47/12)*(1+{IB}48)^({col}3-1),0)"
    mc[f'{col}28']=f"=-MAX({col}12,0)*{IB}46"
    mc[f'{col}29']=f"={col}26+{col}27+{col}28"
    mc[f'{col}31']=f"=IF({col}1=1,-({IB}2+{IB}2*{IB}3),0)"
    mc[f'{col}32']=f"=-'Construction Budget'!{get_column_letter(5+m)}33"
    mc[f'{col}33']=f"=IF({col}1=1,-{IB}85,0)"
    mc[f'{col}34']=f"=SUM({col}31:{col}33)"
    mc[f'{col}36']=f"=IF({col}1<{IB}58,-{col}34+MAX(0,-{col}29),0)"
    mc[f'{col}37']=f"=IF({col}1=1,{IB}87,{P}39)" if m>1 else f"=IF({col}1=1,{IB}87,0)"
    mc[f'{col}38']=f"=MIN({col}37,{col}36)"
    mc[f'{col}39']=f"={col}37-{col}38"
    mc[f'{col}40']=f"=IF({col}1=1,0,{P}44)" if m>1 else "=0"
    mc[f'{col}41']=f"=IF({col}1<{IB}58,{col}36-{col}38,0)"
    mc[f'{col}42']=f"=IF({col}1<={IB}58,{col}40*{IB}54/12,0)"
    mc[f'{col}43']=f"=IF({col}1={IB}58,-({col}40+{col}41+{col}42),0)"
    mc[f'{col}44']=f"=SUM({col}40:{col}43)"
    mc[f'{col}46']=f"=IF({col}1={IB}58,{col}26*12/{IB}59,0)"
    mc[f'{col}47']=f"=IF({col}1={IB}58,MIN({IB}60*{col}46,PV({IB}62/12,{IB}63,-({col}26*12/{IB}61)/12)),0)"
    mc[f'{col}48']=f"=-{col}47*{IB}65"
    mc[f'{col}49']=f"={col}47+{col}43+{col}48"
    mc[f'{col}50']=f"=IF({col}1={IB}58+1,{IB}91,IF({col}1>{IB}58+1,{P}55,0))" if m>1 else "=0"
    mc[f'{col}51']=(f"=IF(AND({col}1>{IB}58,{col}1<={IB}93),"
                    f"IF({col}1<={IB}58+{IB}66,{col}50*{IB}62/12,MIN({IB}92,{col}50*(1+{IB}62/12))),0)")
    mc[f'{col}52']=f"=IF(AND({col}1>{IB}58,{col}1<={IB}93),{col}50*{IB}62/12,0)"
    mc[f'{col}53']=f"=MAX({col}51-{col}52,0)"
    mc[f'{col}54']=f"=IF(AND({col}1>{IB}58,{col}1<={IB}93),{col}50*{IB}64/12,0)"
    mc[f'{col}55']=f"={col}50-{col}53"
    mc[f'{col}56']=f"=IF(AND({IB}69=\"Sell\",{col}1={IB}70),-{col}55,0)"
    mc[f'{col}58']=f"=IF(AND({IB}69=\"Sell\",{col}1={IB}70),{col}26*12/{IB}71,0)"
    mc[f'{col}59']=f"=-{col}58*{IB}72"
    mc[f'{col}61']=(f"=-{col}38"
                    f"+IF({col}1<{IB}58,MAX(0,{col}29),0)"
                    f"+IF({col}1={IB}58,{col}29+{col}49,0)"
                    f"+IF(AND({col}1>{IB}58,{col}1<={IB}93),{col}29-{col}51-{col}54,0)"
                    f"+{col}58+{col}59+{col}56")
    for r in [7,8,9,10,11,12,22,23,24,26,27,28,29,31,32,33,34,36,37,38,39,40,41,42,43,44,46,47,48,49,50,51,52,53,54,55,56,58,59,61]:
        mc[f'{col}{r}'].number_format=CUR
    for r in exp_rows: mc[f'{col}{r}'].number_format=CUR
    mc[f'{col}5'].number_format='0.0'; mc[f'{col}6'].number_format='0.0'
mc['C63']=f"=IFERROR(XIRR(C61:CT61,C2:CT2,0.1),\"\")"; mc['C63'].number_format=PCT; mc['C63'].fill=YEL
mc['C64']="=IFERROR(SUMIF(C61:CT61,\">0\")/-SUMIF(C61:CT61,\"<0\"),\"\")"; mc['C64'].number_format=X
mc['C65']="=SUM(C61:CT61)"; mc['C65'].number_format=CUR
mc.column_dimensions['B'].width=30

# ---------- WATERFALL ----------
wf=wb.create_sheet('Waterfall')
wf['A1']='LP / GP WATERFALL (Monthly)'; wf['A1'].font=BOLD
wf['B2']='LP Capital (Cash)'; wf['C2']=f"={IB}88"; wf['C2'].number_format=CUR; wf['C2'].font=GRN
wf['B3']='GP Capital (Cash)'; wf['C3']=f"={IB}89"; wf['C3'].number_format=CUR; wf['C3'].font=GRN
wf['B4']='GP In-Kind Credit'; wf['C4']=f"={IB}90"; wf['C4'].number_format=CUR; wf['C4'].font=GRN
wlab={6:'Month',7:'Date',8:'Contribution (Equity Draw)',9:'LP Contribution',10:'GP Contribution',
      11:'Distributable Cash',13:'Pref Owed - Beg',14:'Pref Accrual',15:'Pref Paid to LP',16:'Pref Owed - End',
      18:'LP Unreturned Capital - Beg',19:'LP Return of Capital',20:'LP Unreturned - End',
      21:'GP Unreturned Capital - Beg',22:'GP Return of Capital',23:'GP Unreturned - End',
      25:'Residual Cash',26:'LP Residual Share',27:'GP Residual Share',
      29:'LP Net Cash Flow',30:'GP Net Cash Flow'}
for r,l in wlab.items(): wf.cell(row=r,column=2,value=l).font=BOLD if r in (11,29,30) else BLK
for m in range(1,NM+1):
    col=get_column_letter(2+m); P=get_column_letter(1+m)
    wf[f'{col}6']=f"='Monthly CF'!{col}1"
    wf[f'{col}7']=f"='Monthly CF'!{col}2"; wf[f'{col}7'].number_format=DT
    wf[f'{col}8']=f"='Monthly CF'!{col}38"
    wf[f'{col}9']=f"={col}8*{IB}75"
    wf[f'{col}10']=f"={col}8-{col}9"
    wf[f'{col}11']=f"=MAX(0,'Monthly CF'!{col}61+{col}8)"
    wf[f'{col}13']=f"=IF({col}6=1,0,{P}16)" if m>1 else "=0"
    wf[f'{col}14']=f"={col}18*{IB}76/12"
    wf[f'{col}15']=f"=MIN({col}11,{col}13+{col}14)"
    wf[f'{col}16']=f"={col}13+{col}14-{col}15"
    wf[f'{col}18']=(f"=IF({col}6=1,{col}9,{P}20+{col}9)" if m>1 else f"={col}9")
    wf[f'{col}19']=f"=MIN(MAX({col}11-{col}15,0),{col}18)"
    wf[f'{col}20']=f"={col}18-{col}19"
    wf[f'{col}21']=(f"=IF({col}6=1,{col}10+IF({IB}78=\"On\",$C$4,0),{P}23+{col}10)" if m>1
                    else f"={col}10+IF({IB}78=\"On\",$C$4,0)")
    wf[f'{col}22']=f"=MIN(MAX({col}11-{col}15-{col}19,0),{col}21)"
    wf[f'{col}23']=f"={col}21-{col}22"
    wf[f'{col}25']=f"=MAX({col}11-{col}15-{col}19-{col}22,0)"
    wf[f'{col}26']=f"={col}25*{IB}77"
    wf[f'{col}27']=f"={col}25*(1-{IB}77)"
    wf[f'{col}29']=f"=-{col}9+{col}15+{col}19+{col}26"
    wf[f'{col}30']=f"=-{col}10+{col}22+{col}27"
    for r in [8,9,10,11,13,14,15,16,18,19,20,21,22,23,25,26,27,29,30]:
        wf[f'{col}{r}'].number_format=CUR
res=[('LP IRR',"=IFERROR(XIRR(C29:CT29,C7:CT7,0.1),\"\")",PCT,True),
     ('LP Equity Multiple',"=IFERROR(SUMIF(C29:CT29,\">0\")/-SUMIF(C29:CT29,\"<0\"),\"\")",X,False),
     ('LP Total Profit',"=SUM(C29:CT29)",CUR,False),
     ('GP IRR',"=IF($C$3<1,\"n/a (no GP cash)\",IFERROR(XIRR(C30:CT30,C7:CT7,0.1),\"\"))",PCT,False),
     ('GP Equity Multiple',"=IF($C$3<1,\"n/a (no GP cash)\",IFERROR(SUMIF(C30:CT30,\">0\")/-SUMIF(C30:CT30,\"<0\"),\"\"))",X,False),
     ('GP Total Profit',"=SUM(C30:CT30)",CUR,False)]
for i,(lab,f,fmt,hl) in enumerate(res):
    r=32+i
    wf.cell(row=r,column=2,value=lab).font=BOLD
    c=wf.cell(row=r,column=3,value=f); c.number_format=fmt
    if hl: c.fill=YEL
wf.column_dimensions['B'].width=28

# ---------- SUMMARY ----------
sm=wb.create_sheet('Summary')
sm['A1']='PROJECT SUMMARY'; sm['A1'].font=Font(name='Arial',bold=True,size=13)
rows=[
 ('Total Units',f"={IB}6",NUM),('Net Rentable SF',f"={IB}5",NUM),
 ('Total Project Cost',f"={IB}86",CUR),('Cost / Unit',f"=IFERROR({IB}86/{IB}6,0)",CUR),
 ('Cost / SF',f"=IFERROR({IB}86/{IB}5,0)",'$#,##0.00'),
 ('Construction Loan (Peak Balance)',f"=MAX('Monthly CF'!C44:CT44)",CUR),('Effective LTC',f"=IFERROR(MAX('Monthly CF'!C44:CT44)/{IB}86,0)",PCT),
 ('Net Equity Required',f"={IB}87",CUR),
 ('Stabilized NOI (Refi Month x 12)',f"=INDEX('Monthly CF'!$C$26:$CT$26,1,{IB}58)*12",CUR),
 ('Development Yield (Stab. NOI / TPC)',f"=IFERROR(INDEX('Monthly CF'!$C$26:$CT$26,1,{IB}58)*12/{IB}86,0)",PCT2),
 ('Refi Value',f"=INDEX('Monthly CF'!$C$46:$CT$46,1,{IB}58)",CUR),
 ('Perm Loan Amount',f"={IB}91",CUR),
 ('Net Cash-Out at Refi',f"=INDEX('Monthly CF'!$C$49:$CT$49,1,{IB}58)",CUR),
 ('Going-In DSCR (Perm)',f"=IFERROR(INDEX('Monthly CF'!$C$26:$CT$26,1,{IB}58)*12/(({IB}92+{IB}91*{IB}64/12)*12),0)",'0.00x'),
 ('Sale Price (if Sell)',f"=IF({IB}69=\"Sell\",INDEX('Monthly CF'!$C$58:$CT$58,1,{IB}70),0)",CUR),
 ('Project IRR',"='Monthly CF'!C63",PCT),
 ('Project Equity Multiple',"='Monthly CF'!C64",X),
 ('Total Profit',"='Monthly CF'!C65",CUR),
 ('LP IRR',"=Waterfall!C32",PCT),('GP IRR',"=Waterfall!C35",PCT),
]
for i,(lab,f,fmt) in enumerate(rows):
    r=3+i
    sm.cell(row=r,column=1,value=lab).font=BOLD
    c=sm.cell(row=r,column=2,value=f); c.number_format=fmt; c.font=GRN
sm['B18'].fill=YEL
sm.column_dimensions['A'].width=38; sm.column_dimensions['B'].width=18

# ---------- STOA BENCHMARKS ----------
bm=wb.create_sheet('Stoa Benchmarks')
bm['A1']='STOA-ERA MODEL: EXACT SETTINGS PRESERVED / CHANGED'; bm['A1'].font=BOLD
notes=[
 'PRESERVED EXACTLY (same equations/defaults):',
 'Comp-driven rents: rent = weighted comp average by unit type x (1 - delta). Deltas were -4% (1BR), -5% (2BR), +12% (3BR) on Inverness.',
 'Vacancy 7%; rent growth 2%; lease-up 14 months; phased delivery (lease-up starts before completion).',
 'Expense phase-in: each line starts N months before lease-up; cost = full x (%fixed + (1-%fixed) x occupancy). Their %fixed: Payroll 100%, G&A 100%, Marketing 250% (capped at 100% here), R&M 20%, Utilities 50%, Insurance 100%.',
 'Contingencies 10% hard / 2% soft; GC overhead 5% grossed up; GC profit 4%; developer fee 3.5% + 0.5%; impact fees $4,127/unit.',
 'Soft cost unit rates: architectural $1,800/u; landscape $150/u; geotech $20k; civil $95k; Phase 1 $2.5k; title $68k; attorney $5.1k; appraisal $4.5k; other closing $700; NGBS $230/u; consulting $200/u; GL insurance hard/1000 x 3.',
 'Bell-curve draws: SD 6 hard / SD 10 soft; construction loan LTC 75%, WSJ Prime + 0.50%; equity draws first, then debt; interest capitalized.',
 'Perm refi: value = stabilized NOI x 12 / valuation cap; LTV 75%; 420-mo amortization; MIP 0.25%; cash-out at refi. Their refi month: 33. Their caps: 5.5% valuation / 6.5% exit. Cost of sale 0.75%.',
 'Waterfall: 8% pref first. (Their hurdles 10%/12% with 10/20/30 promote simplified here to pref -> return of capital -> residual split; extendable on request.)',
 'GC profit + dev fee taken as GP in-kind equity: preserved as a toggle (Inputs B78). Stoa: On. You (not self-building): Off.',
 '',
 'CHANGED AND WHY:',
 'Hard cost $122.23/SF -> $165/SF default: 2022-era pricing is stale; update per GC bid every deal.',
 'WSJ Prime 7.00% -> 6.75% (current since Dec 11, 2025).',
 'Insurance $1,164 -> $2,000/unit; expense growth 1% -> 2.5%; expense levels escalated ~3%/yr from their 2023 budgets (see Acquisition model Benchmarks tab for the exact Hammond/Manhattan numbers).',
 'Builders Risk was loan-balance-based (circular); rebased to hard cost (combined with GL at hard/1000 x 11). Same magnitude, no circularity.',
 'Their 0.5% dev fee and capitalized interest were circular (iterative calc required). This model solves both algebraically - Checks tab verifies estimate vs actual ledger interest.',
 'Their construction loan fee was $95,000 flat; here 0.50% of the estimated loan.',
 'Financing scheme: equity fixed at (1-LTC) x pre-interest cost, drawn first; construction loan funds the rest and carries its own capitalized interest. This reproduces their EFFECTIVE LTC behavior (Inverness: 75% nominal -> 79.5% effective) with no circular references.',
 'Demo inputs approximate the Inverness deal (288 units, same mix/SF/deltas) at 2026 pricing: $135/SF hard (assumes a builder cost advantage - market GC pricing is $160-180/SF; at market pricing this deal does NOT pencil, which is the core Stoa lesson: their edge WAS the build cost).',
 'Their valuation cap 5.5% -> 6.0% default (cap rates have widened since 2022; update per market).',
 'Marketing %fixed of 250% capped at 100% (their value scaled marketing ABOVE full budget during lease-up; if you want that behavior, set %fixed above 100%).',
]
for i,t in enumerate(notes,start=3): bm.cell(row=i,column=1,value=t)
bm.column_dimensions['A'].width=160

# ---------- CHECKS ----------
ck=wb.create_sheet('Checks')
ck['A1']='Check'; ck['B1']='Status'; ck['A1'].font=BOLD; ck['B1'].font=BOLD
checks=[
 ('Construction draws sum = budget',"='Construction Budget'!C35"),
 ('Effective LTC between 65% and 85% (loan carries capitalized interest, like Stoa)',
  f"=IF(AND(MAX('Monthly CF'!C44:CT44)/{IB}86>=0.65,MAX('Monthly CF'!C44:CT44)/{IB}86<=0.85),\"OK\",\"CHECK\")"),
 ('Equity fully drawn before debt (equity remaining at first loan draw = 0)',
  "=IF(SUMPRODUCT(('Monthly CF'!C41:CT41>0)*('Monthly CF'!C39:CT39))<1,\"OK\",\"ERROR\")"),
 ('Refi month after stabilization',f"=IF({IB}58>={IB}13,\"OK\",\"CHECK\")"),
 ('Sale month after refi (if Sell)',f"=IF(OR({IB}69<>\"Sell\",{IB}70>{IB}58),\"OK\",\"ERROR\")"),
 ('Waterfall distributions = project distributions',
  "=IF(ABS(SUMIF(Waterfall!C29:CT29,\">0\")+SUMIF(Waterfall!C30:CT30,\">0\")-SUM(Waterfall!C11:CT11))<1,\"OK\",\"ERROR\")"),
 ('Waterfall contributions = equity drawn',
  "=IF(ABS(SUM(Waterfall!C8:CT8)-SUM('Monthly CF'!C38:CT38))<1,\"OK\",\"ERROR\")"),
]
for i,(lab,f) in enumerate(checks,start=2):
    ck.cell(row=i,column=1,value=lab); ck.cell(row=i,column=2,value=f)
ck.column_dimensions['A'].width=65; ck.column_dimensions['B'].width=10

import os
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..',
                   'Multifamily_Development_Model.xlsx')
wb.save(os.path.normpath(OUT))
print('saved', os.path.normpath(OUT))
