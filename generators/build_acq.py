"""Original multifamily ACQUISITION model - annual, 11-yr forecast, DSCR/LTV loan sizing,
sale at exit, LP/GP waterfall (pref + ROC + split), sensitivity, Stoa benchmark tab."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Explicit 8-char aRGB. openpyxl versions differ on the alpha byte they infer
# from a 6-char value, and tools/safe_writer.py identifies inputs by this colour.
BLUE = Font(name='Arial', color='FF0000FF')        # inputs
BLK  = Font(name='Arial', color='FF000000')        # formulas
GRN  = Font(name='Arial', color='FF008000')        # cross-sheet links
BOLD = Font(name='Arial', bold=True)
HDR  = Font(name='Arial', bold=True, color='FFFFFF')
YEL  = PatternFill('solid', fgColor='FFFF00')
DK   = PatternFill('solid', fgColor='404040')
CUR  = '$#,##0;($#,##0);-'
PCT  = '0.0%'
PCT2 = '0.00%'
NUM  = '#,##0'
X    = '0.00x'

wb = Workbook()

# ---------------- READ ME ----------------
rm = wb.active; rm.title = 'Read Me'
rm['A1'] = 'MULTIFAMILY ACQUISITION MODEL'; rm['A1'].font = Font(name='Arial', bold=True, size=14)
notes = [
 'Original model. Mechanics mirror standard institutional development/acquisition underwriting',
 '(same equation structure as the Stoa-era model you provided), rebuilt from scratch.',
 '',
 'HOW TO USE:',
 '1. Enter values ONLY in blue-font cells on the Inputs tab. Yellow cells are the key drivers.',
 '2. All other cells are formulas - do not overwrite.',
 '3. Check the Checks tab: all cells in column B should read OK.',
 '4. Forecast runs 11 years so a 10-year hold still has forward NOI for exit pricing.',
 '5. Sale price convention: forward (next-year) NOI divided by exit cap rate.',
 '6. Loan is sized to the LESSER of Max LTV and Min DSCR constraint - matches bank sizing.',
 '7. Waterfall: LP preferred return (cumulative, non-compounding annually) -> LP return of',
 '   capital -> GP return of capital -> residual split LP/GP. Set LP% = 100% if the investor',
 '   funds all equity.',
 '',
 'Assumption provenance and every change vs. the Stoa-era model: see Stoa Benchmarks tab.',
]
for i, t in enumerate(notes, start=3):
    rm.cell(row=i, column=1, value=t).font = BLK
rm.column_dimensions['A'].width = 100

# ---------------- INPUTS ----------------
inp = wb.create_sheet('Inputs')
def put(ws, cell, val, font=BLK, fmt=None, fill=None, bold=False):
    c = ws[cell]; c.value = val
    c.font = Font(name='Arial', bold=bold, color=font.color.rgb[2:] if hasattr(font.color,'rgb') and font.color.rgb else '000000') if False else font
    if bold: c.font = Font(name='Arial', bold=True, color=(font.color.rgb[2:] if font.color and font.color.rgb and len(font.color.rgb)==8 else '000000'))
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    return c

# Section: Acquisition
inp['A1'] = 'ACQUISITION'; inp['A1'].font = BOLD
inp['A2'] = 'Purchase Price'; inp['B2'] = 600000; inp['B2'].font = BLUE; inp['B2'].number_format = CUR; inp['B2'].fill = YEL
inp['A3'] = 'Closing Costs (% of Price)'; inp['B3'] = 0.02; inp['B3'].font = BLUE; inp['B3'].number_format = PCT
inp['A4'] = 'Loan Fee (% of Loan)'; inp['B4'] = 0.01; inp['B4'].font = BLUE; inp['B4'].number_format = PCT
inp['A5'] = 'Total Units'; inp['B5'] = '=E11'; inp['B5'].font = GRN; inp['B5'].number_format = NUM
inp['A6'] = 'Price / Unit'; inp['B6'] = '=IFERROR(B2/B5,0)'; inp['B6'].number_format = CUR

# Rent roll
inp['E1'] = 'RENT ROLL (in-place)'; inp['E1'].font = BOLD
hdrs = ['# Units','Unit Type','SF','Rent/Month','Rent/SF']
for j,h in enumerate(hdrs):
    c = inp.cell(row=2, column=5+j, value=h); c.font = HDR; c.fill = DK
demo = [(4,'1 BR/ 1 BA',700,900),(4,'2 BR/ 2 BA',1000,1100),(0,'BR/BA',0,0),(0,'BR/BA',0,0),
        (0,'BR/BA',0,0),(0,'BR/BA',0,0),(0,'BR/BA',0,0),(0,'BR/BA',0,0)]
for i,(u,t,sf,r) in enumerate(demo):
    rr = 3+i
    inp.cell(row=rr, column=5, value=u).font = BLUE
    inp.cell(row=rr, column=6, value=t).font = BLUE
    inp.cell(row=rr, column=7, value=sf).font = BLUE
    inp.cell(row=rr, column=8, value=r).font = BLUE
    inp.cell(row=rr, column=8).number_format = CUR
    c = inp.cell(row=rr, column=9, value=f'=IFERROR(H{rr}/G{rr},0)'); c.number_format='0.00'
inp['E11'] = '=SUM(E3:E10)'; inp['E11'].number_format = NUM; inp['E11'].font = BOLD
inp['G11'] = '=IFERROR(SUMPRODUCT(G3:G10,E3:E10)/E11,0)'; inp['G11'].number_format = NUM
inp['H11'] = '=IFERROR(SUMPRODUCT(H3:H10,E3:E10)/E11,0)'; inp['H11'].number_format = CUR
inp['I11'] = '=IFERROR(H11/G11,0)'; inp['I11'].number_format = '0.00'

# ---- Refinance (Refi Year 0 = no refinance) ----
# Placed in columns E/F so the row numbering of every other section is unchanged.
inp['E13'] = 'REFINANCE  (Refi Year 0 = none)'; inp['E13'].font = BOLD
refi_inputs = [
    (14, 'Refi Year (0 = none)', 0, NUM),
    (15, 'Refi Valuation Cap Rate', 0.07, PCT2),
    (16, 'Refi LTV', 0.75, PCT),
    (17, 'Refi Min DSCR', 1.25, '0.00'),
    (18, 'Refi Interest Rate', 0.0675, PCT2),
    (19, 'Refi Amortization (Years)', 25, NUM),
    (20, 'Refi Cost (% of New Loan)', 0.02, PCT),
]
for r, lab, val, fmt in refi_inputs:
    inp.cell(row=r, column=5, value=lab)
    c = inp.cell(row=r, column=6, value=val); c.font = BLUE; c.number_format = fmt
inp['F14'].fill = YEL

# NOI at the refi year drives both the value and the DSCR test.
NOI_AT_REFI = "INDEX('Annual CF'!C25:N25,MATCH($F$14,'Annual CF'!C2:N2,0))"
inp['E22'] = 'COMPUTED REFI (do not edit)'; inp['E22'].font = BOLD
inp['E23'] = 'Refi Value (NOI / Val Cap)'
inp['F23'] = f"=IFERROR(IF($F$14=0,0,{NOI_AT_REFI}/$F$15),0)"
inp['E24'] = 'New Loan Amount'
inp['F24'] = (f"=IFERROR(IF($F$14=0,0,MIN($F$16*$F$23,"
              f"PV($F$18/12,$F$19*12,-({NOI_AT_REFI}/$F$17)/12))),0)")
inp['E25'] = 'Old Loan Payoff at Refi'
inp['F25'] = "=IFERROR(IF($F$14=0,0,INDEX('Annual CF'!C35:N35,MATCH($F$14,'Annual CF'!C2:N2,0))),0)"
inp['E26'] = 'Refi Costs'
inp['F26'] = '=$F$24*$F$20'
inp['E27'] = 'Net Cash-Out at Refi'
inp['F27'] = '=$F$24-$F$25-$F$26'
inp['E28'] = 'New Monthly Payment'
inp['F28'] = '=IFERROR(IF($F$14=0,0,PMT($F$18/12,$F$19*12,-$F$24)),0)'
for r in (23, 24, 25, 26, 27, 28):
    inp.cell(row=r, column=6).number_format = CUR
inp['F27'].font = BOLD

# ---- Value-add / renovation (0 units = none) ----
# The budget is funded at close, so it lands in Total Equity Required. That is
# the conservative treatment and it answers the only question that matters on a
# small deal: is there enough cash to buy AND renovate.
inp['E30'] = 'VALUE-ADD / RENOVATION  (0 units = none)'; inp['E30'].font = BOLD
va_inputs = [
    (31, 'Units to Renovate', 0, NUM),
    (32, 'Renovation Cost per Unit', 0, CUR),
    (33, 'Rent Premium per Unit / Month', 0, CUR),
    (34, 'Units Renovated per Year', 0, NUM),
    (35, 'Downtime Months per Turn', 1, NUM),
    (36, 'Renovation Start Year', 1, NUM),
]
for r, lab, val, fmt in va_inputs:
    inp.cell(row=r, column=5, value=lab)
    c = inp.cell(row=r, column=6, value=val); c.font = BLUE; c.number_format = fmt
inp['F31'].fill = YEL; inp['F33'].fill = YEL
inp['E38'] = 'Total Renovation Budget'; inp['F38'] = '=$F$31*$F$32'; inp['F38'].number_format = CUR
inp['F38'].font = BOLD
inp['E39'] = 'Years to Complete'; inp['F39'] = '=IFERROR(ROUNDUP($F$31/$F$34,0),0)'
inp['F39'].number_format = NUM

# Income
inp['A8']  = 'INCOME'; inp['A8'].font = BOLD
rows_inc = [('Vacancy %',0.07,PCT),('Bad Debt %',0.005,PCT),('Concessions %',0.0,PCT),
            ('Other Income ($/Unit/Month)',40,CUR),('Rent Growth (Annual)',0.02,PCT),
            ('Other Income Growth (Annual)',0.02,PCT)]
for i,(lab,val,fmt) in enumerate(rows_inc):
    r = 9+i
    inp.cell(row=r, column=1, value=lab)
    c = inp.cell(row=r, column=2, value=val); c.font = BLUE; c.number_format = fmt

# Expenses
inp['A16'] = 'OPERATING EXPENSES ($/Unit/Year)'; inp['A16'].font = BOLD
exp_rows = [('Payroll (0 for small props - see Benchmarks)',0),('General & Administrative',480),('Marketing',415),
            ('Repairs & Maintenance',515),('Contract Services',0),('Utilities',755),
            ('Other',105),('Insurance',2000)]
for i,(lab,val) in enumerate(exp_rows):
    r = 17+i
    inp.cell(row=r, column=1, value=lab)
    c = inp.cell(row=r, column=2, value=val); c.font = BLUE; c.number_format = CUR
inp['A25'] = 'Property Taxes (Annual $, total)'; inp['B25'] = 7000; inp['B25'].font = BLUE; inp['B25'].number_format = CUR; inp['B25'].fill = YEL
inp['A26'] = 'Property Management Fee (% of EGR)'; inp['B26'] = 0.08; inp['B26'].font = BLUE; inp['B26'].number_format = PCT
inp['A27'] = 'Asset Management Fee (% of EGR)'; inp['B27'] = 0.0; inp['B27'].font = BLUE; inp['B27'].number_format = PCT
inp['A28'] = 'Capital Expense Reserve ($/Unit/Year)'; inp['B28'] = 250; inp['B28'].font = BLUE; inp['B28'].number_format = CUR
inp['A29'] = 'Expense Growth (Annual)'; inp['B29'] = 0.025; inp['B29'].font = BLUE; inp['B29'].number_format = PCT

# Debt
inp['A31'] = 'DEBT'; inp['A31'].font = BOLD
inp['A32'] = 'Max Loan-To-Value (LTV)'; inp['B32'] = 0.75; inp['B32'].font = BLUE; inp['B32'].number_format = PCT
inp['A33'] = 'Min DSCR (Lender Constraint)'; inp['B33'] = 1.25; inp['B33'].font = BLUE; inp['B33'].number_format = '0.00'
inp['A34'] = 'Interest Rate (Fixed)'; inp['B34'] = 0.0675; inp['B34'].font = BLUE; inp['B34'].number_format = PCT2; inp['B34'].fill = YEL
inp['A35'] = 'Amortization (Years)'; inp['B35'] = 25; inp['B35'].font = BLUE; inp['B35'].number_format = NUM
inp['A36'] = 'Interest-Only Period (Years)'; inp['B36'] = 0; inp['B36'].font = BLUE; inp['B36'].number_format = NUM

# Exit
inp['A39'] = 'EXIT'; inp['A39'].font = BOLD
inp['A40'] = 'Hold Period (Years, 1-10)'; inp['B40'] = 5; inp['B40'].font = BLUE; inp['B40'].number_format = NUM; inp['B40'].fill = YEL
inp['A41'] = 'Exit Cap Rate'; inp['B41'] = 0.075; inp['B41'].font = BLUE; inp['B41'].number_format = PCT2; inp['B41'].fill = YEL
inp['A42'] = 'Cost of Sale (% of Sale Price)'; inp['B42'] = 0.03; inp['B42'].font = BLUE; inp['B42'].number_format = PCT

# Equity
inp['A44'] = 'EQUITY / INVESTOR STRUCTURE'; inp['A44'].font = BOLD
inp['A45'] = 'LP (Investor) % of Equity'; inp['B45'] = 0.90; inp['B45'].font = BLUE; inp['B45'].number_format = PCT
inp['A46'] = 'GP (You) % of Equity'; inp['B46'] = '=1-B45'; inp['B46'].number_format = PCT
inp['A47'] = 'LP Preferred Return (Annual)'; inp['B47'] = 0.08; inp['B47'].font = BLUE; inp['B47'].number_format = PCT
inp['A48'] = 'Residual Split - LP %'; inp['B48'] = 0.70; inp['B48'].font = BLUE; inp['B48'].number_format = PCT
inp['A49'] = 'Residual Split - GP %'; inp['B49'] = '=1-B48'; inp['B49'].number_format = PCT

# Computed financing
inp['A51'] = 'COMPUTED (do not edit)'; inp['A51'].font = BOLD
inp['A52'] = 'Loan Amount'
inp['B52'] = "=MIN(B32*B2, PV(B34/12, B35*12, -('Annual CF'!D25/B33)/12))"
inp['B52'].number_format = CUR
inp['A53'] = 'Monthly Debt Service Payment'; inp['B53'] = '=PMT(B34/12,B35*12,-B52)'; inp['B53'].number_format = CUR
inp['A54'] = 'Total Equity Required'; inp['B54'] = '=B2+B2*B3+B52*B4-B52+$F$38'; inp['B54'].number_format = CUR; inp['B54'].font = BOLD
inp['A55'] = 'LP Capital'; inp['B55'] = '=B54*B45'; inp['B55'].number_format = CUR
inp['A56'] = 'GP Capital'; inp['B56'] = '=B54-B55'; inp['B56'].number_format = CUR
inp['A57'] = 'Going-In Cap Rate (Yr1 NOI / Price)'; inp['B57'] = "=IFERROR('Annual CF'!D25/B2,0)"; inp['B57'].number_format = PCT2

inp.column_dimensions['A'].width = 36
inp.column_dimensions['B'].width = 14
for col in 'EFGHI': inp.column_dimensions[col].width = 12

# ---------------- ANNUAL CF ----------------
cf = wb.create_sheet('Annual CF')
IB = "Inputs!$B$"   # shorthand
cf['A1'] = 'ANNUAL CASH FLOW'; cf['A1'].font = BOLD
cf['B2'] = 'Year'; cf['B2'].font = BOLD
for j in range(12):  # C..N = years 0..11
    col = get_column_letter(3+j)
    cf[f'{col}2'] = j
    cf[f'{col}2'].font = BOLD
    cf.column_dimensions[col].width = 12

def rowfill(r, label, formula_fn, fmt=CUR, start_year=1, bold=False, font=BLK):
    cf.cell(row=r, column=2, value=label).font = BOLD if bold else BLK
    for j in range(12):
        col = get_column_letter(3+j)
        if j < start_year:
            continue
        c = cf[f'{col}{r}']
        c.value = formula_fn(col)
        c.number_format = fmt
        c.font = BOLD if bold else font

cf['B4'] = 'Revenue'; cf['B4'].font = BOLD
# Renovation phasing. VA31 total units, VA34 per year, VA36 start year.
VA = "Inputs!$F$"
CUM   = lambda c: f"MIN({VA}31,MAX(0,({c}$2-{VA}36+1)*{VA}34))"   # renovated by end of year
PRIOR = lambda c: f"MIN({VA}31,MAX(0,({c}$2-{VA}36)*{VA}34))"      # renovated before it
GRW   = lambda c: f"(1+{IB}13)^({c}$2-1)"
rowfill(5,'Gross Potential Rent',
        lambda c: f"=(SUMPRODUCT(Inputs!$E$3:$E$10,Inputs!$H$3:$H$10)+{CUM(c)}*{VA}33)*12*{GRW(c)}")
# Units being turned sit empty for the downtime period - a real cost of renovating.
rowfill(6,'Physical Vacancy + Reno Downtime',
        lambda c: f"=-({c}5*{IB}9+({CUM(c)}-{PRIOR(c)})*{VA}35*Inputs!$H$11*{GRW(c)})")
rowfill(7,'Bad Debt',             lambda c: f"=-{c}5*{IB}10")
rowfill(8,'Concessions',          lambda c: f"=-{c}5*{IB}11")
rowfill(9,'Net Rental Income',    lambda c: f"=SUM({c}5:{c}8)", bold=True)
rowfill(10,'Other Income',        lambda c: f"={IB}5*{IB}12*12*(1-{IB}9)*(1+{IB}14)^({c}$2-1)")
rowfill(11,'Effective Gross Revenue', lambda c: f"={c}9+{c}10", bold=True)

cf['B13'] = 'Expenses'; cf['B13'].font = BOLD
exp_map = [(14,17,'Payroll'),(15,18,'General & Administrative'),(16,19,'Marketing'),
           (17,20,'Repairs & Maintenance'),(18,21,'Contract Services'),(19,22,'Utilities'),
           (20,23,'Other'),(21,24,'Insurance')]
for r, ir, lab in exp_map:
    rowfill(r, lab, lambda c, ir=ir: f"=-Inputs!$B${ir}*{IB}5*(1+{IB}29)^({c}$2-1)")
rowfill(22,'Property Taxes', lambda c: f"=-{IB}25*(1+{IB}29)^({c}$2-1)")
rowfill(23,'Property Management Fee', lambda c: f"=-{c}11*{IB}26")
rowfill(24,'Total Operating Expenses', lambda c: f"=SUM({c}14:{c}23)", bold=True)

rowfill(25,'Net Operating Income', lambda c: f"={c}11+{c}24", bold=True)
rowfill(26,'Capital Expense Reserve', lambda c: f"=-{IB}5*{IB}28*(1+{IB}29)^({c}$2-1)")
rowfill(27,'Asset Management Fee', lambda c: f"=-{c}11*{IB}27")
rowfill(28,'Cash Flow Before Debt Service', lambda c: f"={c}25+{c}26+{c}27", bold=True)

cf['B30'] = 'Debt Service'; cf['B30'].font = BOLD
# After the refi year the ledger runs on the new loan's rate and payment.
IF_ = "Inputs!$F$"
RATE_T = lambda c: f"IF(AND({IF_}14>0,{c}$2>{IF_}14),{IF_}18,{IB}34)"
PMT_T  = lambda c: f"IF(AND({IF_}14>0,{c}$2>{IF_}14),{IF_}28,{IB}53)"
for j in range(1,12):
    col = get_column_letter(3+j); prev = get_column_letter(2+j)
    # Year 1 draws the original loan; the year after the refi draws the new one.
    cf[f'{col}31'] = f"=IF({col}$2=1,{IB}52,IF(AND({IF_}14>0,{col}$2={IF_}14+1),{IF_}24,{prev}35))"
    cf[f'{col}31'].number_format = CUR
cf['B31'] = 'Beginning Loan Balance'
rowfill(32,'Interest Payment', lambda c: f"=IF({c}$2<={IB}36,{c}31*{RATE_T(c)},12*{PMT_T(c)}-({c}31-{c}35))")
rowfill(33,'Principal Payment', lambda c: f"={c}31-{c}35")
rowfill(34,'Total Debt Service', lambda c: f"={c}32+{c}33", bold=True)
rowfill(35,'Ending Loan Balance', lambda c: f"=IF({c}$2<={IB}36,{c}31,MAX(0,FV({RATE_T(c)}/12,12,{PMT_T(c)},-{c}31)))")
rowfill(37,'Cash Flow After Debt Service', lambda c: f"={c}28-{c}34", bold=True)
rowfill(38,'DSCR', lambda c: f"=IFERROR({c}25/{c}34,\"\")", fmt='0.00x')
rowfill(39,'Cash-on-Cash Return', lambda c: f"=IFERROR({c}37/{IB}54,\"\")", fmt=PCT)
# Row 26 (capex reserve) is stored negative, so 25+26 is NOI net of reserves.
# Agency/HUD underwrite DSCR on this basis; small banks often use row 38.
rowfill(40,'DSCR (after capex reserve)', lambda c: f"=IFERROR(({c}25+{c}26)/{c}34,\"\")", fmt='0.00x')

cf['B41'] = 'Acquisition & Sale'; cf['B41'].font = BOLD
# Year 0 column C
cf['C42'] = f"=-{IB}2"; cf['B42']='Purchase Price'
cf['C43'] = f"=-{IB}2*{IB}3"; cf['B43']='Closing Costs'
cf['C44'] = f"=-{IB}52*{IB}4"; cf['B44']='Loan Fees'
cf['C45'] = f"={IB}52"; cf['B45']='Loan Proceeds'
for r in range(42,46): cf[f'C{r}'].number_format = CUR
# Sale rows across years 1..10 (need forward NOI so cap at year 10 = col M)
cf['B46'] = 'Sale Price (Fwd NOI / Exit Cap)'
cf['B47'] = 'Cost of Sale'
cf['B48'] = 'Loan Payoff'
for j in range(1,11):
    col = get_column_letter(3+j); nxt = get_column_letter(4+j)
    cf[f'{col}46'] = f"=IF({col}$2={IB}40,{nxt}25/{IB}41,0)"
    cf[f'{col}47'] = f"=-{col}46*{IB}42"
    cf[f'{col}48'] = f"=IF({col}$2={IB}40,-{col}35,0)"
    for r in (46,47,48): cf[f'{col}{r}'].number_format = CUR

cf['B49'] = 'Refi Proceeds (net of costs)'
for j in range(1,12):
    col = get_column_letter(3+j)
    cf[f'{col}49'] = f"=IF(AND({IF_}14>0,{col}$2={IF_}14),{IF_}27,0)"
    cf[f'{col}49'].number_format = CUR

cf['B50'] = 'Unlevered Net Cash Flow'; cf['B50'].font = BOLD
cf['B51'] = 'Levered Net Cash Flow'; cf['B51'].font = BOLD
cf['B52'] = 'Renovation Budget (funded at close)'
cf['C52'] = f"=-{VA}38"; cf['C52'].number_format = CUR
cf['C50'] = "=C42+C43+C52"; cf['C50'].number_format = CUR; cf['C50'].font = BOLD
cf['C51'] = "=C42+C43+C44+C45+C52"; cf['C51'].number_format = CUR; cf['C51'].font = BOLD
for j in range(1,12):
    col = get_column_letter(3+j)
    s46 = f"{col}46+{col}47" if j<=10 else "0"
    s48 = f"{col}48" if j<=10 else "0"
    cf[f'{col}50'] = f"=IF({col}$2<={IB}40,{col}28,0)+{s46}"
    # Refi proceeds are a financing event: levered cash flow only.
    cf[f'{col}51'] = f"=IF({col}$2<={IB}40,{col}37,0)+{s46}+{s48}+{col}49"
    cf[f'{col}50'].number_format = CUR; cf[f'{col}50'].font = BOLD
    cf[f'{col}51'].number_format = CUR; cf[f'{col}51'].font = BOLD

cf['B53'] = 'RETURNS (Project Level)'; cf['B53'].font = BOLD
cf['B54'] = 'Unlevered IRR'; cf['C54'] = "=IFERROR(IRR(C50:N50,0.1),\"\")"; cf['C54'].number_format = PCT
cf['B55'] = 'Levered IRR';   cf['C55'] = "=IFERROR(IRR(C51:N51,0.1),\"\")"; cf['C55'].number_format = PCT; cf['C55'].fill = YEL
cf['B56'] = 'Equity Multiple'; cf['C56'] = "=IFERROR(1+SUM(C51:N51)/(-C51),\"\")"; cf['C56'].number_format = X
cf['B57'] = 'Total Profit'; cf['C57'] = "=SUM(C51:N51)"; cf['C57'].number_format = CUR
cf['B58'] = 'Avg Cash-on-Cash (Hold Yrs)'; cf['C58'] = f"=IFERROR(AVERAGEIFS(D37:N37,D2:N2,\"<=\"&{IB}40)/{IB}54,\"\")"; cf['C58'].number_format = PCT
cf['B59'] = 'Avg DSCR (Hold Yrs)'; cf['C59'] = f"=IFERROR(AVERAGEIFS(D25:N25,D2:N2,\"<=\"&{IB}40)/AVERAGEIFS(D34:N34,D2:N2,\"<=\"&{IB}40),\"\")"; cf['C59'].number_format = '0.00x'; cf['C59'].fill = YEL

cf['B61'] = 'RENOVATION DETAIL'; cf['B61'].font = BOLD
rowfill(62,'Units Renovated (cumulative)', lambda c: f"={CUM(c)}", fmt=NUM)
rowfill(63,'Rent Premium Captured', lambda c: f"={CUM(c)}*{VA}33*12*{GRW(c)}")
rowfill(64,'Downtime Loss', lambda c: f"=-({CUM(c)}-{PRIOR(c)})*{VA}35*Inputs!$H$11*{GRW(c)}")
cf.column_dimensions['B'].width = 32

# ---------------- WATERFALL ----------------
wf = wb.create_sheet('Waterfall')
wf['A1'] = 'LP / GP DISTRIBUTION WATERFALL (Annual)'; wf['A1'].font = BOLD
wf['B2'] = 'LP Capital'; wf['C2'] = f"={IB}55"; wf['C2'].number_format = CUR; wf['C2'].font = GRN
wf['B3'] = 'GP Capital'; wf['C3'] = f"={IB}56"; wf['C3'].number_format = CUR; wf['C3'].font = GRN
wf['B5'] = 'Year'; wf['B5'].font = BOLD
for j in range(12):
    col = get_column_letter(3+j)
    wf[f'{col}5'] = f"='Annual CF'!{col}2"; wf[f'{col}5'].font = BOLD
    wf.column_dimensions[col].width = 12
def wrow(r, label, fn, fmt=CUR, bold=False, years=range(1,12)):
    wf.cell(row=r, column=2, value=label).font = BOLD if bold else BLK
    for j in years:
        col = get_column_letter(3+j)
        c = wf[f'{col}{r}']; c.value = fn(col, j); c.number_format = fmt
        if bold: c.font = BOLD
wrow(6,'Distributable Cash Flow', lambda c,j: f"='Annual CF'!{c}51", bold=True)
wf['B8']='TIER 1: LP Preferred Return'; wf['B8'].font=BOLD
for j in range(1,12):
    col = get_column_letter(3+j); prev = get_column_letter(2+j)
    wf[f'{col}9']  = f"=IF({col}$5=1,0,{prev}12)"
    # Pref accrues on UNRETURNED capital (row 15), not original contribution.
    # Returned capital stops earning the preferred return - market convention.
    wf[f'{col}10'] = f"=IF({col}$5<={IB}40,{col}15*{IB}47,0)"
    wf[f'{col}11'] = f"=MIN(MAX({col}6,0),{col}9+{col}10)"
    wf[f'{col}12'] = f"={col}9+{col}10-{col}11"
    for r in range(9,13): wf[f'{col}{r}'].number_format = CUR
wf['B9']='Pref Owed - Beginning'; wf['B10']='Pref Accrual'; wf['B11']='Pref Paid to LP'; wf['B12']='Pref Owed - Ending'
wf['B14']='TIER 2: Return of Capital'; wf['B14'].font=BOLD
for j in range(1,12):
    col = get_column_letter(3+j); prev = get_column_letter(2+j)
    wf[f'{col}15'] = f"=IF({col}$5=1,$C$2,{prev}17)"
    wf[f'{col}16'] = f"=MIN(MAX({col}6-{col}11,0),{col}15)"
    wf[f'{col}17'] = f"={col}15-{col}16"
    wf[f'{col}18'] = f"=IF({col}$5=1,$C$3,{prev}20)"
    wf[f'{col}19'] = f"=MIN(MAX({col}6-{col}11-{col}16,0),{col}18)"
    wf[f'{col}20'] = f"={col}18-{col}19"
    for r in range(15,21): wf[f'{col}{r}'].number_format = CUR
wf['B15']='LP Unreturned Capital - Beg'; wf['B16']='LP Return of Capital'; wf['B17']='LP Unreturned Capital - End'
wf['B18']='GP Unreturned Capital - Beg'; wf['B19']='GP Return of Capital'; wf['B20']='GP Unreturned Capital - End'
wf['B22']='TIER 3: Residual Split'; wf['B22'].font=BOLD
for j in range(1,12):
    col = get_column_letter(3+j)
    wf[f'{col}23'] = f"=MAX({col}6-{col}11-{col}16-{col}19,0)"
    wf[f'{col}24'] = f"={col}23*{IB}48"
    wf[f'{col}25'] = f"={col}23*{IB}49"
    for r in range(23,26): wf[f'{col}{r}'].number_format = CUR
wf['B23']='Residual Cash Flow'; wf['B24']='LP Residual Share'; wf['B25']='GP Residual Share'
wf['B27']='LP Net Cash Flow'; wf['B27'].font=BOLD
wf['B28']='GP Net Cash Flow'; wf['B28'].font=BOLD
wf['C27'] = "=-$C$2"; wf['C28'] = "=-$C$3"
wf['C27'].number_format = CUR; wf['C28'].number_format = CUR
for j in range(1,12):
    col = get_column_letter(3+j)
    wf[f'{col}27'] = f"={col}11+{col}16+{col}24"
    wf[f'{col}28'] = f"={col}19+{col}25"
    wf[f'{col}27'].number_format = CUR; wf[f'{col}28'].number_format = CUR
wf['B30']='LP IRR'; wf['C30']="=IFERROR(IRR(C27:N27,0.1),\"\")"; wf['C30'].number_format = PCT; wf['C30'].fill = YEL
wf['B31']='LP Equity Multiple'; wf['C31']="=IFERROR(1+SUM(C27:N27)/$C$2,\"\")"; wf['C31'].number_format = X
wf['B32']='LP Total Profit'; wf['C32']="=SUM(C27:N27)"; wf['C32'].number_format = CUR
wf['B33']='GP IRR'; wf['C33']="=IFERROR(IRR(C28:N28,0.1),\"\")"; wf['C33'].number_format = PCT
wf['B34']='GP Equity Multiple'; wf['C34']="=IFERROR(1+SUM(C28:N28)/$C$3,\"\")"; wf['C34'].number_format = X
wf['B35']='GP Total Profit'; wf['C35']="=SUM(C28:N28)"; wf['C35'].number_format = CUR
wf.column_dimensions['B'].width = 30

# ---------------- SENSITIVITY ----------------
sn = wb.create_sheet('Sensitivity')
sn['A1'] = 'EXIT CAP RATE SENSITIVITY (at chosen Hold Period)'; sn['A1'].font = BOLD
sn['A3'] = 'Forward NOI at Exit'; sn['B3'] = f"=INDEX('Annual CF'!C25:N25,MATCH({IB}40+1,'Annual CF'!C2:N2,0))"; sn['B3'].number_format = CUR
sn['A4'] = 'Loan Payoff at Exit'; sn['B4'] = f"=INDEX('Annual CF'!C35:N35,MATCH({IB}40,'Annual CF'!C2:N2,0))"; sn['B4'].number_format = CUR
sn['A5'] = 'Cum. CF After Debt Svc (Hold Yrs)'; sn['B5'] = f"=SUMIFS('Annual CF'!D37:N37,'Annual CF'!D2:N2,\"<=\"&{IB}40)"; sn['B5'].number_format = CUR
sn['A6'] = 'Total Equity'; sn['B6'] = f"={IB}54"; sn['B6'].number_format = CUR
hdr2 = ['Exit Cap','Sale Price','Net Sale Proceeds','Total Profit','Equity Multiple']
for j,h in enumerate(hdr2):
    c = sn.cell(row=8, column=1+j, value=h); c.font = HDR; c.fill = DK
caps = [0.05+0.0025*i for i in range(13)]
for i,cap in enumerate(caps):
    r = 9+i
    sn.cell(row=r, column=1, value=cap).number_format = PCT2
    sn.cell(row=r, column=1).font = BLUE
    sn.cell(row=r, column=2, value=f"=IFERROR($B$3/A{r},0)").number_format = CUR
    sn.cell(row=r, column=3, value=f"=B{r}*(1-{IB}42)-$B$4").number_format = CUR
    sn.cell(row=r, column=4, value=f"=C{r}+$B$5-$B$6").number_format = CUR
    sn.cell(row=r, column=5, value=f"=IFERROR((C{r}+$B$5)/$B$6,0)").number_format = X
for col,w in [('A',12),('B',14),('C',16),('D',14),('E',14)]: sn.column_dimensions[col].width = w

# ---------------- STOA BENCHMARKS ----------------
bm = wb.create_sheet('Stoa Benchmarks')
bm['A1'] = 'STOA-ERA MODEL ASSUMPTIONS (exact) AND WHAT THIS MODEL CHANGED'; bm['A1'].font = BOLD
data = [
 ('','','',''),
 ('OPERATING EXPENSES ($/Unit/Yr) - from Waters at Hammond & Manhattan budgets','','',''),
 ('Line Item','Hammond','Manhattan','Average (their formula: =AVERAGE)'),
 ('Payroll',-1661.03,-1476.90,-1568.96),
 ('General & Administrative',-476.09,-400.82,-438.45),
 ('Marketing',-362.42,-395.13,-378.78),
 ('Repairs & Maintenance',-572.19,-372.59,-472.39),
 ('Utilities',-641.47,-742.20,-691.84),
 ('Other',-105.04,-85.33,-95.19),
 ('Insurance',-1094.69,-1233.33,-1164.01),
 ('','','',''),
 ('THEIR MODEL SETTINGS (exact)','','',''),
 ('Vacancy 7% | at Refi 7%','Rent Growth 2%/yr','Expense Growth 1%/yr','Capex Reserve $250/unit/yr'),
 ('Mgmt Fee 3% of EGR (self-managed at scale)','Asset Mgmt Fee 1% of NRI','Bad Debt/Concessions inside vacancy','Other income itemized (valet trash, late fees, pet rent/fees, admin, carport, termination)'),
 ('Construction: LTC 75%, WSJ Prime + 0.50%','Loan fee $95,000 flat','Hard cost = $122.23/SF x NRSF + $1,000,000','Contingency: 10% hard / 2% soft; GC Overhead 5%'),
 ('GC Profit 4% | Developer Fee 3.5% + 0.5% (taken as in-kind equity)','Impact fees $4,127/unit','Draw curve: normal distribution, SD 6 (hard) / 10 (soft)','Lease-up 14 months; ~20.6 units/month (288 units)'),
 ('Perm: FHA 223(f) refi at month 33','Valuation cap 5.5%; LTV 75%; DSCR ~1.185 going-in','420-month amortization; MIP 0.25%','Perm fee ~2.14% of proceeds'),
 ('Exit: sell at 6.5% cap on forward NOI','Cost of sale 0.75%','Sell-or-Hold toggle','Waterfall: 8% pref; hurdles 10%/12%; promote 10/20/30; GP/LP splits'),
 ('','','',''),
 ('CHANGES MADE IN THIS MODEL AND WHY','','',''),
 ('Insurance: $1,164 -> $2,000/unit default','Louisiana insurance costs escalated sharply post-2022; LA is the 3rd most expensive insurance state. Update per actual quote every deal.','',''),
 ('Expense growth: 1% -> 2.5%','1% was aggressive; recent expense inflation (insurance, labor) runs above rents. Conservative default.','',''),
 ('Mgmt fee: 3% -> 8%','Stoa self-managed 300+ unit assets. Third-party management on <50 units typically runs 7-10% of collections.','',''),
 ('Cost of sale: 0.75% -> 3%','Stoa sold $30M+ assets off-market. Small multifamily sells with brokerage: 2-5% typical.','',''),
 ('Expense levels escalated ~3%/yr from their ~2023 budgets','Payroll 1569->1715, G&A 438->480, Mktg 379->415, R&M 472->515, Util 692->755, Other 95->105.','Note: on very small properties (<10 units) payroll is usually $0 (no on-site staff; covered by mgmt fee). Zero it out when appropriate.',''),
 ('Rate defaults: WSJ Prime = 6.75% (as of Dec 11, 2025)','Bank acquisition debt default 6.75% fixed. Replace with your actual quote.','',''),
 ('All other equation structures preserved: forward-NOI exit pricing, DSCR/LTV loan sizing, pref-then-ROC-then-split waterfall, fixed/variable expense logic (Development model), bell-curve draws (Development model).','','',''),
]
for i,rowv in enumerate(data, start=2):
    for j,v in enumerate(rowv):
        c = bm.cell(row=i, column=1+j, value=v)
        if isinstance(v,(int,float)): c.number_format = CUR
bm.column_dimensions['A'].width = 55
for col in 'BCD': bm.column_dimensions[col].width = 45

# ---------------- CHECKS ----------------
ck = wb.create_sheet('Checks')
ck['A1'] = 'Check'; ck['B1'] = 'Status'; ck['A1'].font = BOLD; ck['B1'].font = BOLD
checks = [
 ('Sources = Uses (Loan + Equity = Price + Closing + Fees + Renovation)',
  f"=IF(ABS({IB}52+{IB}54-({IB}2+{IB}2*{IB}3+{IB}52*{IB}4+Inputs!$F$38))<0.01,\"OK\",\"ERROR\")"),
 ('Year 1 DSCR meets lender minimum',
  f"=IF('Annual CF'!D38>={IB}33-0.005,\"OK\",\"CHECK\")"),
 ('Hold period within 1-10 years',
  f"=IF(AND({IB}40>=1,{IB}40<=10),\"OK\",\"ERROR\")"),
 ('Waterfall distributions = project levered CF (yrs 1+)',
  f"=IF(ABS(SUM(Waterfall!D27:N27)+SUM(Waterfall!D28:N28)-SUMIFS('Annual CF'!D51:N51,'Annual CF'!D2:N2,\">=1\"))<1,\"OK\",\"ERROR\")"),
 ('Loan balance non-negative at exit',
  f"=IF(INDEX('Annual CF'!C35:N35,MATCH({IB}40,'Annual CF'!C2:N2,0))>=0,\"OK\",\"ERROR\")"),
 ('Post-reserve DSCR >= 1.15 (lender view, net of replacement reserves)',
  f"=IF('Annual CF'!D40>=1.15,\"OK\",\"CHECK\")"),
 ('No balloon at maturity (an interest-only period leaves one)',
  f"=IF({IB}36=0,\"OK\",\"CHECK\")"),
 ('Exit cap not below going-in cap (compression is an unearned assumption)',
  f"=IF({IB}41>={IB}57-0.0001,\"OK\",\"CHECK\")"),
 ('Refi year (if any) falls inside the hold period',
  f"=IF(OR(Inputs!$F$14=0,AND(Inputs!$F$14>=1,Inputs!$F$14<{IB}40)),\"OK\",\"CHECK\")"),
 ('Refi loan does not exceed the refi valuation',
  "=IF(OR(Inputs!$F$14=0,Inputs!$F$24<=Inputs!$F$23*Inputs!$F$16+1),\"OK\",\"CHECK\")"),
 ('Renovation unit count does not exceed the property',
  f"=IF(Inputs!$F$31<={IB}5,\"OK\",\"ERROR\")"),
 ('Renovation finishes inside the hold period',
  f"=IF(OR(Inputs!$F$31=0,Inputs!$F$36+Inputs!$F$39-1<={IB}40),\"OK\",\"CHECK\")"),
]
for i,(lab,f) in enumerate(checks, start=2):
    ck.cell(row=i, column=1, value=lab)
    ck.cell(row=i, column=2, value=f)
ck.column_dimensions['A'].width = 60; ck.column_dimensions['B'].width = 10

import os
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..',
                   'Multifamily_Acquisition_Model.xlsx')
wb.save(os.path.normpath(OUT))
print('saved', os.path.normpath(OUT))
